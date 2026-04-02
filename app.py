"""在庫管理・自動発注システム"""
import sys, os, csv, io, hashlib, hmac, threading, queue, json, time as _time, logging, math

# Windows CP932 環境でも UTF-8 で出力できるよう stdout/stderr を再設定
if hasattr(sys.stdout, 'reconfigure'):
    try:
        sys.stdout.reconfigure(encoding='utf-8', errors='replace')
        sys.stderr.reconfigure(encoding='utf-8', errors='replace')
    except Exception:
        pass

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s [%(levelname)s] %(name)s: %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S',
)
logger = logging.getLogger('inventory.app')
from pathlib import Path
from datetime import date, timedelta, datetime
from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, Response, session, g
from dotenv import load_dotenv

BASE_DIR = Path(__file__).parent
# .env を utf-8 / utf-8-sig / shift-jis の順で試して読み込む
_env_path = BASE_DIR / '.env'
for _enc in ('utf-8-sig', 'utf-8', 'shift_jis', 'cp932'):
    try:
        _env_path.read_text(encoding=_enc)
        load_dotenv(_env_path, override=True, encoding=_enc)
        break
    except (UnicodeDecodeError, Exception):
        continue

from database import init_db
from mail_service import send_order_mail, send_expiry_alert
from auto_check import (run_order_check, run_expiry_check, run_csv_import,
                        run_month_end_import, is_month_end,
                        start_scheduler, update_reorder_points, create_inventory_count,
                        get_pending_orders)

app = Flask(__name__, template_folder='templates', static_folder='static')
app.config['TEMPLATES_AUTO_RELOAD'] = True
_sk = os.getenv('SECRET_KEY', '')
if not _sk:
    import warnings
    warnings.warn("SECRET_KEY が未設定です。.env に設定してください。", stacklevel=1)
    _sk = 'inventory-secret-key-change-this-in-production'
app.secret_key = _sk

# ── CSRF保護 ──────────────────────────────────────────────────────────
try:
    from flask_wtf.csrf import CSRFProtect
    csrf = CSRFProtect(app)
except ImportError:
    csrf = None

# ── レート制限（ブルートフォース対策） ──────────────────────────────────
try:
    from flask_limiter import Limiter
    from flask_limiter.util import get_remote_address
    limiter = Limiter(
        get_remote_address,
        app=app,
        default_limits=[],
        storage_uri='memory://',
    )
except ImportError:
    limiter = None

def _rate_limit(limit_string):
    """limiter が利用可能な場合のみレート制限を適用する条件付きデコレータ"""
    def decorator(f):
        if limiter:
            return limiter.limit(limit_string)(f)
        return f
    return decorator

# ── DB接続管理（リクエストスコープ） ─────────────────────────────────
# get_db()をFlask gオブジェクトで一元管理し、リクエスト終了時に必ずcloseする
# これにより PostgreSQL の接続が確実に解放され、接続枯渇を防ぐ
def get_db():
    if 'db' not in g:
        from database import get_db as _get_db
        g.db = _get_db()
    return g.db

@app.teardown_appcontext
def close_db(error):
    db = g.pop('db', None)
    if db is not None:
        try:
            db.close()
        except Exception:
            pass


# ── 予測キャッシュ & バックグラウンド事前計算 ──────────────────────────
# ・TTL内は即座にメモリから返す（最速パス）
# ・キャッシュ失効／無効化時は背景スレッドで再構築し、ユーザーを待たせない
# ・起動時・CSV取込後にも背景スレッドで事前計算してウォームアップ
_fc_lock      = threading.Lock()
_fc_store:    dict  = {}           # {'rows': list, 'ts': datetime, 'mode_key': str}
_FC_TTL       = 86400              # 24時間（サーバー再起動がなければほぼ再計算不要）
_fc_computing = False              # 背景計算進行中フラグ
_fc_event     = threading.Event()  # 計算完了通知
_fc_event.set()                    # 初期値: 「計算中でない」= set 済み

_agg_lock    = threading.Lock()
_agg_running = False               # sales_daily_agg 更新進行中フラグ


def _refresh_sales_daily_agg(db):
    """sales_history を日次集計して sales_daily_agg へ UPSERT（過去365日分）"""
    db.execute("""
        INSERT INTO sales_daily_agg (jan, sale_dt, dow, qty)
        SELECT
            jan,
            sale_date::date                               AS sale_dt,
            EXTRACT(ISODOW FROM sale_date::date)::int     AS dow,
            SUM(quantity)                                 AS qty
        FROM sales_history
        WHERE sale_date::date >= CURRENT_DATE - INTERVAL '365 days'
        GROUP BY jan, sale_date::date, EXTRACT(ISODOW FROM sale_date::date)::int
        ON CONFLICT (jan, sale_dt) DO UPDATE
            SET qty = EXCLUDED.qty, dow = EXCLUDED.dow
    """)
    db.commit()
    logger.info('[PerfOpt] sales_daily_agg UPSERT 完了')


def _bg_refresh_sales_daily_agg():
    """バックグラウンドで sales_daily_agg を更新する（重複実行防止付き）"""
    global _agg_running
    with _agg_lock:
        if _agg_running:
            return
        _agg_running = True
    try:
        from database import get_dsn, DBConn
        import psycopg2 as _pg2
        conn = _pg2.connect(**get_dsn(long_timeout=True))
        conn.autocommit = False
        _db = DBConn(conn)
        _refresh_sales_daily_agg(_db)
        _db.close()
    except Exception as _e:
        logger.warning(f'[PerfOpt] sales_daily_agg 更新エラー: {_e}')
    finally:
        with _agg_lock:
            _agg_running = False


def _bg_rebuild_forecast_cache():
    """バックグラウンドで予測キャッシュを再構築（ユーザーを待たせない）"""
    global _fc_computing, _fc_store
    with _fc_lock:
        if _fc_computing:
            return
        _fc_computing = True
        _fc_event.clear()
    try:
        from database import get_dsn, DBConn
        import psycopg2 as _pg2
        conn = _pg2.connect(**get_dsn(long_timeout=True))
        conn.autocommit = False
        _db = DBConn(conn)
        flags    = _get_forecast_feature_flags(_db)
        mode_key = f"{flags['forecast_ai_mode']}_{flags['forecast_reorder_mode']}"
        rows     = _build_forecast_rows_raw(_db, flags)
        _db.close()
        with _fc_lock:
            _fc_store = {'rows': rows, 'ts': datetime.now(), 'mode_key': mode_key}
        logger.info(f'[ForecastCache] 背景再構築完了: {len(rows)}商品')
    except Exception as _e:
        logger.warning(f'[ForecastCache] 背景再構築エラー: {_e}')
    finally:
        with _fc_lock:
            _fc_computing = False
        _fc_event.set()


def invalidate_forecast_cache(background_refresh: bool = True):
    """商品・在庫・設定変更時に呼び出してキャッシュを破棄し、背景で再構築する"""
    global _fc_store
    with _fc_lock:
        _fc_store = {}
    if background_refresh:
        threading.Thread(target=_bg_rebuild_forecast_cache, daemon=True).start()


# ── Blueprint 登録 ─────────────────────────────────────────────
from blueprints.auth import bp as auth_bp
from blueprints.dashboard import bp as dashboard_bp
from blueprints.inventory import bp as inventory_bp
from blueprints.orders import bp as orders_bp
from blueprints.products import bp as products_bp
from blueprints.chains import bp as chains_bp
from blueprints.forecast import bp as forecast_bp

app.register_blueprint(auth_bp)
app.register_blueprint(dashboard_bp)
app.register_blueprint(inventory_bp)
app.register_blueprint(orders_bp)
app.register_blueprint(products_bp)
app.register_blueprint(chains_bp)
app.register_blueprint(forecast_bp)
@app.template_filter('format_number')
def format_number_filter(value):
    """数値をカンマ区切りで表示するJinjaフィルター"""
    try:
        if value is None or value == '':
            return '-'
        n = int(float(str(value)))
        return f'{n:,}'
    except (ValueError, TypeError):
        return str(value) if value is not None else '-'



def _normalize_jan(val):
    """JANコード正規化: 指数表記(4.90123E+12)を整数文字列に変換"""
    if val is None:
        return ''
    s = str(val).strip()
    if not s:
        return ''
    try:
        # 指数表記の場合は数値として解釈して整数化
        f = float(s)
        if 'e' in s.lower() or '.' in s:
            return str(int(f))
        return s
    except (ValueError, OverflowError):
        return s

def _normalize_date(val):
    """日付正規化: 各種形式をYYYY-MM-DDに統一
    対応: 2026/3/30, 2026/03/30, 2026-3-30, 20260330, datetime等
    """
    if val is None:
        return ''
    import datetime as _dt
    # すでにdatetime/dateオブジェクト
    if isinstance(val, (_dt.datetime, _dt.date)):
        return val.strftime('%Y-%m-%d')
    s = str(val).strip()
    if not s:
        return ''
    # YYYY-MM-DD（そのまま）
    if len(s) == 10 and s[4] == '-' and s[7] == '-':
        return s
    # YYYY/MM/DD or YYYY/M/D
    if '/' in s:
        parts = s.split('/')
        if len(parts) == 3:
            try:
                y, m, d = int(parts[0]), int(parts[1]), int(parts[2])
                return f'{y:04d}-{m:02d}-{d:02d}'
            except ValueError:
                pass
    # YYYYMMDD
    if len(s) == 8 and s.isdigit():
        return f'{s[:4]}-{s[4:6]}-{s[6:8]}'
    # YYYY-M-D
    if '-' in s:
        parts = s.split('-')
        if len(parts) == 3:
            try:
                y, m, d = int(parts[0]), int(parts[1]), int(parts[2])
                return f'{y:04d}-{m:02d}-{d:02d}'
            except ValueError:
                pass
    return s


def _to_int(v, default=0):
    try:
        return int(float(str(v)))
    except Exception:
        return default

def _safe_date(s):
    try:
        return datetime.strptime(str(s), '%Y-%m-%d').date()
    except Exception:
        return None

def _resolve_product_by_code(db, code):
    code = (code or '').strip()
    if not code:
        return None
    return db.execute("SELECT jan, product_cd, product_name, supplier_name FROM products WHERE (jan=%s OR product_cd=%s) AND is_active=1 ORDER BY jan LIMIT 1", [code, code]).fetchone()

def _record_receipt(db, product, qty, expiry, lot_no='', location_code='', source='manual', note=''):
    before = db.execute("SELECT COALESCE(SUM(quantity),0) AS _sum FROM stocks WHERE jan=%s", [product['jan']]).fetchone()['_sum']
    db.execute("""
        INSERT INTO stocks
        (product_id,jan,product_name,supplier_cd,supplier_name,
         product_cd,unit_qty,quantity,expiry_date,lot_no,location_code)
        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
    """, [product['id'],product['jan'],product['product_name'],
          product['supplier_cd'],product['supplier_name'],product['product_cd'],
          product['unit_qty'],qty,expiry,lot_no,location_code or product.get('location_code','') or ''])
    db.execute("""
        INSERT INTO stock_movements
        (jan,product_name,move_type,quantity,before_qty,after_qty,note,source_file,move_date,expiry_date)
        VALUES (%s,%s,'receipt',%s,%s,%s,%s,%s,%s,%s)
    """, [product['jan'],product['product_name'],qty,before,before+qty,
          note or '', source, str(date.today()), expiry])


def _build_promotion_calendar(db, horizon_days=35):
    rows = db.execute("""
        SELECT pp.jan,
               TO_CHAR(pp.promo_date, 'YYYY-MM-DD') AS promo_date,
               pp.uplift_factor, pp.promo_name
        FROM promotion_plans pp
        WHERE pp.promo_date BETWEEN CURRENT_DATE - INTERVAL '1 day'
                                AND CURRENT_DATE + (%s || ' days')::interval
    """, [horizon_days]).fetchall()
    mp = {}
    for r in rows:
        mp.setdefault(r['jan'], {})[r['promo_date']] = {
            'uplift_factor': float(r.get('uplift_factor') or 1.0),
            'promo_name': r.get('promo_name') or ''
        }
    return mp


def _build_demand_plan_map(db, horizon_days=35):
    rows = db.execute("""
        SELECT jan,
               TO_CHAR(demand_date, 'YYYY-MM-DD') AS demand_date,
               SUM(demand_qty) AS demand_qty
        FROM demand_plans
        WHERE demand_date BETWEEN CURRENT_DATE - INTERVAL '1 day'
                              AND CURRENT_DATE + (%s || ' days')::interval
        GROUP BY jan, demand_date
    """, [horizon_days]).fetchall()
    mp = {}
    for r in rows:
        mp.setdefault(r['jan'], {})[r['demand_date']] = int(r.get('demand_qty') or 0)
    return mp


def _abc_rank_from_ratio(ratio):
    if ratio <= 0.7:
        return 'A'
    if ratio <= 0.9:
        return 'B'
    return 'C'

def _get_settings_all(db) -> dict:
    """settingsテーブルを丸ごとリクエストスコープでキャッシュ（1リクエスト中の重複クエリを防ぐ）"""
    try:
        if not hasattr(g, '_settings_cache'):
            rows = db.execute("SELECT key, value FROM settings").fetchall()
            g._settings_cache = {r['key']: (r['value'] or '') for r in rows}
        return g._settings_cache
    except RuntimeError:
        # Flaskアプリコンテキスト外（スケジューラ等）ではキャッシュしない
        rows = db.execute("SELECT key, value FROM settings").fetchall()
        return {r['key']: (r['value'] or '') for r in rows}


def _get_forecast_feature_flags(db):
    s = _get_settings_all(db)
    # forecast_ai_mode: '1'=AIモードON(全機能) / '0'=前年実績モード(シンプル)
    ai_mode      = str(s.get('forecast_ai_mode', '1')).strip() in ('1','true','True','on','yes')
    # P2発注点モード: sf/p80/p90（AIモードON/OFF共通で有効）
    reorder_mode = s.get('forecast_reorder_mode', 'sf') or 'sf'
    return {
        'forecast_ai_mode':      ai_mode,
        'forecast_reorder_mode': reorder_mode,
    }



def _build_forecast_rows(db, q=''):
    """キャッシュラッパー: TTL内ならメモリから即返却。背景計算中は完了を待機。"""
    global _fc_store
    flags    = _get_forecast_feature_flags(db)
    mode_key = f"{flags['forecast_ai_mode']}_{flags['forecast_reorder_mode']}"
    now      = datetime.now()

    def _apply_q(rows):
        if not q:
            return rows
        ql = q.strip().lower()
        return [r for r in rows if any(ql in str(r.get(k) or '').lower()
                                       for k in ('jan','product_cd','product_name','supplier_cd','supplier_name'))]

    # ── 最速パス: キャッシュヒット ──────────────────────────────────────
    with _fc_lock:
        c = _fc_store
        if (c and c.get('mode_key') == mode_key
                and (now - c['ts']).total_seconds() < _FC_TTL):
            return _apply_q(c['rows'])

    # ── 背景スレッドが計算中なら最大 60 秒待つ ────────────────────────
    if _fc_computing:
        _fc_event.wait(timeout=60)
        with _fc_lock:
            c = _fc_store
            if c and c.get('mode_key') == mode_key:
                return _apply_q(c['rows'])

    # ── フォールバック: 同期計算（通常は起動直後のごく短時間のみ）──────
    rows = _build_forecast_rows_raw(db, flags)
    with _fc_lock:
        _fc_store = {'rows': rows, 'ts': datetime.now(), 'mode_key': mode_key}
    return _apply_q(rows)


def _build_forecast_rows_raw(db, flags=None):
    """実際の予測計算（重いCTEクエリ）。キャッシュから呼び出されるため q フィルタなし"""
    if flags is None:
        flags = _get_forecast_feature_flags(db)
    ai_mode      = flags['forecast_ai_mode']      # True=AIモード / False=前年実績モード
    reorder_mode = flags['forecast_reorder_mode']  # sf/p80/p90

    # AIモード時のみ販促・受注予定マップを取得
    promo_map  = _build_promotion_calendar(db, 35) if ai_mode else {}
    demand_map = _build_demand_plan_map(db, 35)  # AIモードON/OFF共通で受注予定を反映

    # 前年実績モード用: 前年同月の日次平均を取得
    if not ai_mode:
        import calendar as _cal
        _today = date.today()
        _last_year  = _today.year - 1
        _this_month = _today.month
        _days_in_month = _cal.monthrange(_last_year, _this_month)[1]
        # sale_date は TEXT 型 'YYYY-MM-DD' → キャスト不要で文字列 BETWEEN により ix_sales_history_sale_date を使用
        _ly_start = f'{_last_year}-{_this_month:02d}-01'
        _ly_end   = f'{_last_year}-{_this_month:02d}-{_days_in_month:02d}'
        _ly_rows = db.execute("""
            SELECT jan, COALESCE(SUM(quantity),0) AS total_qty
            FROM sales_history
            WHERE sale_date BETWEEN %s AND %s
            GROUP BY jan
        """, [_ly_start, _ly_end]).fetchall()
        _ly_map = {r['jan']: float(r['total_qty'] or 0) / _days_in_month for r in _ly_rows}
    else:
        _ly_map = {}

    # ── daily CTE のデータソースを選択 ───────────────────────────────────
    # sales_daily_agg に直近7日分のデータがあれば使用（高速インデックス読み込み）
    # なければ sales_history を直接スキャンしつつ、背景スレッドで集計テーブルを構築する
    try:
        _agg_cnt = db.execute(
            "SELECT COUNT(*) AS c FROM sales_daily_agg "
            "WHERE sale_dt >= CURRENT_DATE - INTERVAL '7 days'"
        ).fetchone()['c']
        _use_agg = int(_agg_cnt or 0) > 0
    except Exception:
        _use_agg = False

    if _use_agg:
        _daily_cte = """daily AS (
            SELECT jan, sale_dt, dow, qty
            FROM   sales_daily_agg
            WHERE  sale_dt >= CURRENT_DATE - INTERVAL '180 days'
        )"""
    else:
        # 初回起動 or 集計テーブル未構築 → sales_history を直接使い、背景で集計
        threading.Thread(target=_bg_refresh_sales_daily_agg, daemon=True).start()
        _daily_cte = """daily AS (
            SELECT sh.jan,
                   sh.sale_date::date                            AS sale_dt,
                   EXTRACT(ISODOW FROM sh.sale_date::date)::int  AS dow,
                   SUM(sh.quantity)                              AS qty
            FROM sales_history sh
            WHERE sh.sale_date::date >= CURRENT_DATE - INTERVAL '180 days'
            GROUP BY sh.jan, sh.sale_date::date, dow
        )"""

    rows = db.execute(f"""
        WITH {_daily_cte},
        monthly AS (
            SELECT jan,
                   EXTRACT(MONTH FROM sale_dt)::int AS mon,
                   DATE_TRUNC('month', sale_dt)      AS ym,
                   SUM(qty)                          AS qty
            FROM daily
            WHERE sale_dt >= DATE_TRUNC('month', CURRENT_DATE) - INTERVAL '12 months'
            GROUP BY jan, mon, ym
        ), avg_all AS (
            SELECT jan, AVG(qty) AS avg_monthly
            FROM monthly GROUP BY jan
        ), season AS (
            SELECT m.jan, m.mon,
                   CASE WHEN COALESCE(a.avg_monthly,0)=0 THEN 1
                        ELSE AVG(m.qty)/a.avg_monthly END AS season_idx
            FROM monthly m
            JOIN avg_all a ON a.jan=m.jan
            GROUP BY m.jan, m.mon, a.avg_monthly
        ), dow_avg AS (
            SELECT jan, dow, AVG(qty) AS dow_qty
            FROM daily
            WHERE sale_dt >= CURRENT_DATE - INTERVAL '84 days'
            GROUP BY jan, dow
        ), wma_daily AS (
            SELECT jan,
                   SUM(qty * day_weight) / NULLIF(SUM(day_weight), 0) AS avg_daily
            FROM (
                SELECT jan, qty,
                       ROW_NUMBER() OVER (PARTITION BY jan ORDER BY sale_dt) AS day_weight
                FROM daily
                WHERE sale_dt >= CURRENT_DATE - INTERVAL '30 days'
            ) w
            GROUP BY jan
        ), base_daily AS (
            SELECT d.jan,
                   COALESCE(w.avg_daily, plain.avg_daily) AS avg_daily
            FROM (SELECT DISTINCT jan FROM daily) d
            LEFT JOIN wma_daily w ON w.jan = d.jan
            LEFT JOIN (
                SELECT jan, AVG(qty) AS avg_daily FROM daily
                WHERE sale_dt >= CURRENT_DATE - INTERVAL '84 days'
                GROUP BY jan
            ) plain ON plain.jan = d.jan
        ), dow_pivot AS (
            -- 7曜日インデックスを1回の GROUP BY で横展開（従来の7 LEFT JOIN を削減）
            SELECT d.jan,
                   COALESCE(MAX(CASE WHEN d.dow=1 THEN
                       CASE WHEN COALESCE(b.avg_daily,0)=0 THEN 1.0
                            ELSE d.dow_qty/b.avg_daily END END), 1.0) AS dow_idx_1,
                   COALESCE(MAX(CASE WHEN d.dow=2 THEN
                       CASE WHEN COALESCE(b.avg_daily,0)=0 THEN 1.0
                            ELSE d.dow_qty/b.avg_daily END END), 1.0) AS dow_idx_2,
                   COALESCE(MAX(CASE WHEN d.dow=3 THEN
                       CASE WHEN COALESCE(b.avg_daily,0)=0 THEN 1.0
                            ELSE d.dow_qty/b.avg_daily END END), 1.0) AS dow_idx_3,
                   COALESCE(MAX(CASE WHEN d.dow=4 THEN
                       CASE WHEN COALESCE(b.avg_daily,0)=0 THEN 1.0
                            ELSE d.dow_qty/b.avg_daily END END), 1.0) AS dow_idx_4,
                   COALESCE(MAX(CASE WHEN d.dow=5 THEN
                       CASE WHEN COALESCE(b.avg_daily,0)=0 THEN 1.0
                            ELSE d.dow_qty/b.avg_daily END END), 1.0) AS dow_idx_5,
                   COALESCE(MAX(CASE WHEN d.dow=6 THEN
                       CASE WHEN COALESCE(b.avg_daily,0)=0 THEN 1.0
                            ELSE d.dow_qty/b.avg_daily END END), 1.0) AS dow_idx_6,
                   COALESCE(MAX(CASE WHEN d.dow=7 THEN
                       CASE WHEN COALESCE(b.avg_daily,0)=0 THEN 1.0
                            ELSE d.dow_qty/b.avg_daily END END), 1.0) AS dow_idx_7
            FROM dow_avg d
            JOIN base_daily b ON b.jan = d.jan
            GROUP BY d.jan
        ), stock AS (
            SELECT jan, SUM(quantity) AS stock_qty FROM stocks WHERE quantity>0 GROUP BY jan
        )
        SELECT p.id AS product_id, p.supplier_cd, p.supplier_name, p.product_cd, p.jan, p.product_name,
               p.reorder_point, p.order_unit, p.order_qty, p.unit_qty, p.lead_time_days, p.safety_factor,
               p.reorder_auto, p.shelf_face_qty, p.shelf_replenish_point,
               COALESCE(s.stock_qty,0)              AS stock_qty,
               ROUND(COALESCE(a.avg_monthly,0),1)   AS avg_monthly,
               ROUND(COALESCE(b.avg_daily,0),2)     AS avg_daily,
               ROUND(COALESCE(b.avg_daily,0),2)     AS avg_daily_wma,
               ROUND(COALESCE(se_this.season_idx,1),4) AS season_idx_this,
               ROUND(COALESCE(se_next.season_idx,1),4) AS season_idx_next,
               ROUND(COALESCE(dp.dow_idx_1,1),4)   AS dow_idx_1,
               ROUND(COALESCE(dp.dow_idx_2,1),4)   AS dow_idx_2,
               ROUND(COALESCE(dp.dow_idx_3,1),4)   AS dow_idx_3,
               ROUND(COALESCE(dp.dow_idx_4,1),4)   AS dow_idx_4,
               ROUND(COALESCE(dp.dow_idx_5,1),4)   AS dow_idx_5,
               ROUND(COALESCE(dp.dow_idx_6,1),4)   AS dow_idx_6,
               ROUND(COALESCE(dp.dow_idx_7,1),4)   AS dow_idx_7
        FROM products p
        LEFT JOIN avg_all   a        ON a.jan=p.jan
        LEFT JOIN season    se_this  ON se_this.jan=p.jan
                                    AND se_this.mon=EXTRACT(MONTH FROM CURRENT_DATE)::int
        LEFT JOIN season    se_next  ON se_next.jan=p.jan
                                    AND se_next.mon=EXTRACT(MONTH FROM CURRENT_DATE
                                                            + INTERVAL '1 month')::int
        LEFT JOIN base_daily b       ON b.jan=p.jan
        LEFT JOIN dow_pivot  dp      ON dp.jan=p.jan
        LEFT JOIN stock      s       ON s.jan=p.jan
        WHERE p.is_active=1
        ORDER BY p.supplier_cd, p.product_cd
    """).fetchall()

    import calendar, statistics as _stats

    reorder_mode = flags.get('forecast_reorder_mode', 'sf')  # P2: 'sf'/'p80'/'p90'

    # P2: 分位点計算のために過去30日の日次売上を取得
    # sale_date は TEXT 型 → ::date キャストを除去し文字列比較でインデックスを活用
    quantile_map = {}
    _q30_start = (date.today() - timedelta(days=30)).strftime('%Y-%m-%d')
    try:
        qrows = db.execute("""
            SELECT jan,
                   array_agg(qty ORDER BY sale_dt) AS daily_qtys
            FROM (
                SELECT jan,
                       sale_date AS sale_dt,
                       SUM(quantity) AS qty
                FROM sales_history
                WHERE sale_date >= %s
                GROUP BY jan, sale_date
            ) d
            GROUP BY jan
        """, [_q30_start]).fetchall()
        for qr in qrows:
            qtys = sorted([int(x) for x in (qr['daily_qtys'] or []) if x is not None])
            if len(qtys) >= 5:  # 前日データのみ取込環境を考慮して5件以上で算出
                n = len(qtys)
                quantile_map[qr['jan']] = {
                    'p80':  qtys[int(n * 0.80)],
                    'p90':  qtys[int(n * 0.90)],
                    'mean': _stats.mean(qtys),
                    'std':  _stats.stdev(qtys) if len(qtys) >= 2 else 0.0,
                }
    except Exception:
        pass

    out = []
    today = date.today()
    for r in rows:
        r = dict(r)

        manual_adj = 1.0  # デフォルト値（前年実績モード時はP3無効のため1.0固定）
        if not ai_mode:
            # == 前年実績モード ==
            # 前年同月の日次平均をそのまま使用（WMA/季節/曜日/販促/受注/P3 全て無効）
            ly_daily = _ly_map.get(r['jan'], 0.0)
            base_next_30   = ly_daily * 30.0
            adjusted30       = base_next_30   # 販促・受注予定は加算しない
            base_next_daily_display = ly_daily
            avg_daily           = ly_daily
            season_idx_display  = 1.0
            dow_idx_display     = 1.0
            promo_count         = 0
            promo_uplift_qty    = 0.0
            # AIモードOFF時でも受注予定は反映する
            direct_demand_qty   = 0
            direct_demand_days  = 0
            demand_days_ly = demand_map.get(r['jan'], {})
            for _i in range(35):
                _tday = today + timedelta(days=_i)
                _planned = int(demand_days_ly.get(str(_tday), 0) or 0)
                if _planned > 0:
                    direct_demand_qty += _planned
                    direct_demand_days += 1
            p80_daily = p90_daily = daily_std = None
        else:
            # == AIモード（全機能） ==
            # P1: WMAベースのavg_daily
            avg_daily = float(r.get('avg_daily') or 0)

            # 曜日指数（全7曜日・常時ON）
            dow_idx_map = {
                d: float(r.get(f'dow_idx_{d}') or 1)
                for d in range(1, 8)
            }

            # 季節指数（今月・来月按分・常時ON）
            season_idx_this = float(r.get('season_idx_this') or 1)
            season_idx_next = float(r.get('season_idx_next') or 1)

            season_idx_display = season_idx_next
            dow_idx_display = dow_idx_map[today.isoweekday()]

            # P3: 手動調整係数（商品ごとの係数で制御）
            manual_adj = max(0.1, float(r.get('manual_adj_factor') or 1.0))

            base_next_30 = 0.0
            promo_days  = promo_map.get(r['jan'], {})
            demand_days = demand_map.get(r['jan'], {})

            promo_uplift_qty = 0.0
            promo_count = 0
            direct_demand_qty = 0
            direct_demand_days = 0

            for i in range(30):
                target_day = today + timedelta(days=i)
                d_idx = dow_idx_map[target_day.isoweekday()]
                s_idx = season_idx_this if target_day.month == today.month else season_idx_next
                daily_fc = avg_daily * s_idx * d_idx * manual_adj
                base_next_30 += daily_fc

                ds = str(target_day)
                promo = promo_days.get(ds)
                if promo and float(promo.get('uplift_factor', 1.0) or 1.0) > 1:
                    promo_count += 1
                    promo_uplift_qty += daily_fc * (float(promo['uplift_factor']) - 1.0)
                planned = int(demand_days.get(ds, 0) or 0)
                if planned > 0:
                    direct_demand_qty += planned
                    direct_demand_days += 1

            adjusted30 = base_next_30 + promo_uplift_qty + direct_demand_qty
            base_next_daily_display = base_next_30 / 30.0

            # P2: 分位点データ（AIモード時のみ）
            qdata = quantile_map.get(r['jan'], {})
            p80_daily = qdata.get('p80', None)
            p90_daily = qdata.get('p90', None)
            daily_std = qdata.get('std', None)


        lt = max(int(r.get('lead_time_days') or 1), 1)
        sf = max(float(r.get('safety_factor') or 1.0), 1.0)
        next_daily = round(adjusted30 / 30.0, 2) if adjusted30 else 0

        # P2: 発注点モード（AIモード・前年実績モード共通）
        if not ai_mode:
            # 前年実績モード: 前年日次平均を基準に発注点計算
            ly_daily_for_rp = _ly_map.get(r['jan'], 0.0)
            if reorder_mode == 'p90' and ly_daily_for_rp > 0:
                suggested_rp = int(max(0, ly_daily_for_rp * lt * 1.1 + 0.9999))  # P90相当: +10%
                rp_mode_label = 'P90'
            elif reorder_mode == 'p80' and ly_daily_for_rp > 0:
                suggested_rp = int(max(0, ly_daily_for_rp * lt * 1.05 + 0.9999))  # P80相当: +5%
                rp_mode_label = 'P80'
            else:
                suggested_rp = int(max(0, next_daily * lt * sf + 0.9999))
                rp_mode_label = 'SF'
        else:
            # AIモード: WMA分位点ベースで発注点計算
            if reorder_mode == 'p90' and p90_daily is not None:
                suggested_rp = int(max(0, p90_daily * lt + 0.9999))
                rp_mode_label = 'P90'
            elif reorder_mode == 'p80' and p80_daily is not None:
                suggested_rp = int(max(0, p80_daily * lt + 0.9999))
                rp_mode_label = 'P80'
            else:
                suggested_rp = int(max(0, next_daily * lt * sf + 0.9999))
                rp_mode_label = 'SF'

        # ロットサイズ考慮: order_unit / unit_qty の倍数に切り上げ
        _order_unit = max(1, int(r.get('order_unit') or 1))
        _unit_qty   = max(1, int(r.get('unit_qty')   or 1))

        # 発注点: order_unit の倍数に切り上げ
        if _order_unit > 1 and suggested_rp > 0:
            suggested_rp = math.ceil(suggested_rp / _order_unit) * _order_unit
        else:
            suggested_rp = max(0, suggested_rp)

        # 推奨発注数 = (LT + 14) × 日次予測 → unit_qty の倍数に切り上げ
        raw_oq = max(0, next_daily * (lt + 14))
        if _unit_qty > 1:
            suggested_oq = max(_unit_qty, math.ceil(raw_oq / _unit_qty) * _unit_qty)
        else:
            suggested_oq = int(max(0, raw_oq + 0.9999))

        r['season_idx']         = round(season_idx_display, 2)
        r['avg_dow_idx']        = round(dow_idx_display, 2)
        r['manual_adj_factor']  = round(manual_adj, 2)          # P3
        r['promo_days']         = promo_count
        r['promo_uplift_qty']   = round(promo_uplift_qty, 1)
        r['direct_demand_qty']  = int(direct_demand_qty)
        r['direct_demand_days'] = int(direct_demand_days)
        r['next_30d_forecast']  = round(adjusted30, 1)
        r['next_daily_forecast']= next_daily
        r['weighted_daily_forecast'] = round(base_next_daily_display, 2)
        r['next_month_forecast']= round(adjusted30, 1)
        r['cover_days']         = round(float(r.get('stock_qty') or 0) / next_daily, 1) if next_daily else None
        r['suggested_reorder_point'] = suggested_rp
        r['rp_mode_label']      = rp_mode_label                 # P2
        r['p80_daily']          = p80_daily                     # P2
        r['p90_daily']          = p90_daily                     # P2
        r['daily_std']          = round(daily_std, 2) if daily_std is not None else None  # P2
        r['suggested_order_qty']= suggested_oq
        out.append(r)
    return out

def _build_picking_plan(db, days=7, q=''):
    rows = db.execute("""
        WITH demand AS (
            SELECT jan,
                   CEIL(SUM(quantity) / GREATEST(COUNT(DISTINCT sale_date::date),1) * %s) AS need_qty
            FROM sales_history
            WHERE sale_date::date >= CURRENT_DATE - (%s || ' days')::interval
            GROUP BY jan
        )
        SELECT p.supplier_cd, p.supplier_name, p.product_cd, p.jan, p.product_name,
               COALESCE(d.need_qty, 0) AS need_qty,
               s.id AS stock_id, s.expiry_date, s.lot_no,
               COALESCE(NULLIF(s.location_code,''), NULLIF(p.location_code,''), '') AS location_code,
               s.quantity
        FROM products p
        LEFT JOIN demand d ON d.jan=p.jan
        JOIN stocks s ON s.jan=p.jan AND s.quantity>0
        WHERE p.is_active=1 AND COALESCE(d.need_qty,0) > 0
        ORDER BY p.supplier_cd, p.product_cd,
                 CASE WHEN s.expiry_date='' THEN '9999-99-99' ELSE s.expiry_date END ASC,
                 COALESCE(NULLIF(s.location_code,''), NULLIF(p.location_code,''), '') ASC
    """, [days, days]).fetchall()
    q = (q or '').strip().lower()
    if q:
        rows=[r for r in rows if q in (r['jan'] or '').lower() or q in (r['product_cd'] or '').lower() or q in (r['product_name'] or '').lower() or q in (r['location_code'] or '').lower()]
    plan=[]
    current_jan=None
    remaining=0
    for r in rows:
        if r['jan'] != current_jan:
            current_jan=r['jan']
            remaining=int(r['need_qty'] or 0)
        if remaining <= 0:
            continue
        pick=min(int(r['quantity'] or 0), remaining)
        if pick <= 0:
            continue
        remaining -= pick
        x=dict(r)
        x['pick_qty']=pick
        x['remaining_after']=remaining
        x['location_code']=x['location_code'] or '未設定'
        plan.append(x)
    return plan


def _build_shortage_rows(db, q=''):
    forecast_rows = _build_forecast_rows(db, q='')
    forecast_map = {r['jan']: r for r in forecast_rows}
    today = date.today()
    inbound_rows = db.execute("""
        SELECT oh.jan,
               COALESCE(NULLIF(oh.expected_receipt_date,''), TO_CHAR((oh.order_date::date + (COALESCE(p.lead_time_days,3) || ' days')::interval)::date, 'YYYY-MM-DD')) AS eta,
               GREATEST(oh.order_qty - COALESCE(rcpt.total_received, 0), 0) AS outstanding_qty
        FROM order_history oh
        JOIN products p ON p.jan = oh.jan
        LEFT JOIN (
            SELECT order_history_id, SUM(received_qty) AS total_received
            FROM order_receipts
            GROUP BY order_history_id
        ) rcpt ON rcpt.order_history_id = oh.id
        WHERE oh.order_qty - COALESCE(rcpt.total_received, 0) > 0
    """).fetchall()
    inbound = {}
    for r in inbound_rows:
        eta = r['eta'] or str(today + timedelta(days=3))
        inbound.setdefault(r['jan'], {}).setdefault(eta, 0)
        inbound[r['jan']][eta] += int(r['outstanding_qty'] or 0)

    abc_rows = db.execute("""
        WITH sales AS (
            SELECT sh.jan, SUM(sh.quantity * COALESCE(p.cost_price,0)) AS sales_value
            FROM sales_history sh
            JOIN products p ON p.jan=sh.jan
            WHERE sh.sale_date::date >= CURRENT_DATE - INTERVAL '365 days'
            GROUP BY sh.jan
        ), ranked AS (
            SELECT jan, sales_value,
                   SUM(sales_value) OVER () AS total_sales,
                   SUM(sales_value) OVER (ORDER BY sales_value DESC NULLS LAST, jan) AS running_sales
            FROM sales
        )
        SELECT jan, COALESCE(sales_value,0) AS sales_value,
               CASE WHEN COALESCE(total_sales,0)=0 THEN 'C'
                    WHEN (running_sales/NULLIF(total_sales,0)) <= 0.7 THEN 'A'
                    WHEN (running_sales/NULLIF(total_sales,0)) <= 0.9 THEN 'B'
                    ELSE 'C' END AS abc_class
        FROM ranked
    """).fetchall()
    abc_map = {r['jan']: {'abc_class': r['abc_class'], 'sales_value': float(r['sales_value'] or 0)} for r in abc_rows}

    delay_rows = db.execute("""
        SELECT oh.jan,
               COUNT(*) FILTER (
                   WHERE oh.order_qty - COALESCE(rcpt.total_received, 0) > 0
                   AND COALESCE(NULLIF(oh.expected_receipt_date,''), oh.order_date) < CURRENT_DATE::text
               ) AS overdue_count
        FROM order_history oh
        LEFT JOIN (
            SELECT order_history_id, SUM(received_qty) AS total_received
            FROM order_receipts
            GROUP BY order_history_id
        ) rcpt ON rcpt.order_history_id = oh.id
        GROUP BY oh.jan
    """).fetchall()
    delay_map = {r['jan']: r for r in delay_rows}

    stock_rows = db.execute("""
        SELECT p.id AS product_id,p.supplier_cd,p.supplier_name,p.product_cd,p.jan,p.product_name,p.reorder_point,p.order_qty,p.lead_time_days,
               COALESCE(SUM(s.quantity),0) AS stock_qty,
               COALESCE(p.cost_price,0) AS cost_price,
               p.ordered_at
        FROM products p
        LEFT JOIN stocks s ON s.jan=p.jan
        WHERE p.is_active=1
        GROUP BY p.id
        ORDER BY p.supplier_cd,p.product_cd
    """).fetchall()
    out=[]
    risk_order = {'入荷前欠品': 4, '予測欠品': 3, '要注意': 2, '安全': 1}
    abc_weight = {'A': 3, 'B': 2, 'C': 1}
    for r in stock_rows:
        fr = forecast_map.get(r['jan'], {})
        daily = float(fr.get('next_daily_forecast') or 0)
        if q and not any(q in str(r.get(k) or '').lower() for k in ('jan','product_cd','product_name','supplier_cd','supplier_name')):
            continue
        projected = float(r.get('stock_qty') or 0)
        stockout_date = None
        worst_stock = projected
        first_inbound = None
        first_inbound_qty = 0
        for i in range(0, 31):
            d = str(today + timedelta(days=i))
            if d in inbound.get(r['jan'], {}):
                projected += inbound[r['jan']][d]
                if first_inbound is None:
                    first_inbound = d
                    first_inbound_qty = inbound[r['jan']][d]
            projected -= daily
            worst_stock = min(worst_stock, projected)
            if stockout_date is None and projected < 0:
                stockout_date = d
        risk = '安全'
        if stockout_date:
            if first_inbound and stockout_date < first_inbound:
                risk = '入荷前欠品'
            else:
                risk = '予測欠品'
        elif projected <= int(r.get('reorder_point') or 0):
            risk = '要注意'
        abc = abc_map.get(r['jan'], {'abc_class': 'C', 'sales_value': 0})
        overdue_count = int((delay_map.get(r['jan']) or {}).get('overdue_count') or 0)
        days_to_stockout = ((_safe_date(stockout_date) - today).days if stockout_date else None)
        priority_score = risk_order.get(risk, 0) * 100 + abc_weight.get(abc['abc_class'], 1) * 10 + (20 - min(max(days_to_stockout or 20, 0), 20))
        action = '監視'
        if risk == '入荷前欠品':
            action = '即発注候補'
        elif risk == '予測欠品':
            action = '追加発注候補'
        elif risk == '要注意' and abc['abc_class'] == 'A':
            action = '前倒し確認'
        out.append({**dict(r),
            'daily_forecast': round(daily,2),
            'forecast_30d': round(daily*30,1),
            'expected_receipt_date': first_inbound or '',
            'expected_receipt_qty': first_inbound_qty,
            'stockout_date': stockout_date or '',
            'projected_30d_stock': round(projected,1),
            'worst_projected_stock': round(worst_stock,1),
            'risk_level': risk,
            'days_to_stockout': days_to_stockout,
            'abc_class': abc['abc_class'],
            'annual_sales_value': round(abc['sales_value'], 0),
            'priority_score': int(priority_score),
            'recommended_action': action,
            'overdue_order_count': overdue_count,
        })
    out.sort(key=lambda x:(-x['priority_score'], x['days_to_stockout'] if x['days_to_stockout'] is not None else 9999, x['supplier_cd'], x['product_cd']))
    return out


def _build_replenishment_rows(db, q=''):
    rows = db.execute("""
        WITH shelf AS (
            SELECT p.jan,
                   COALESCE(SUM(CASE WHEN COALESCE(s.location_code,'') = COALESCE(NULLIF(p.location_code,''), '__none__') THEN s.quantity ELSE 0 END),0) AS shelf_qty,
                   COALESCE(SUM(CASE WHEN COALESCE(s.location_code,'') <> COALESCE(NULLIF(p.location_code,''), '__none__') THEN s.quantity ELSE 0 END),0) AS reserve_qty,
                   MIN(CASE WHEN COALESCE(s.location_code,'') <> COALESCE(NULLIF(p.location_code,''), '__none__') THEN s.expiry_date END) AS reserve_oldest_expiry
            FROM products p
            LEFT JOIN stocks s ON s.jan=p.jan AND s.quantity>0
            WHERE p.is_active=1
            GROUP BY p.jan, p.location_code
        )
        SELECT p.id AS product_id,p.supplier_cd,p.supplier_name,p.product_cd,p.jan,p.product_name,
               COALESCE(NULLIF(p.location_code,''),'未設定') AS shelf_location,
               p.shelf_face_qty,p.shelf_replenish_point,
               COALESCE(sh.shelf_qty,0) AS shelf_qty,
               COALESCE(sh.reserve_qty,0) AS reserve_qty,
               sh.reserve_oldest_expiry
        FROM products p
        LEFT JOIN shelf sh ON sh.jan=p.jan
        WHERE p.is_active=1
        ORDER BY p.supplier_cd,p.product_cd
    """).fetchall()
    out=[]
    for r in rows:
        if q and not any(q in str(r.get(k) or '').lower() for k in ('jan','product_cd','product_name','supplier_cd','supplier_name','shelf_location')):
            continue
        target = int(r.get('shelf_face_qty') or 0)
        trigger = int(r.get('shelf_replenish_point') or 0)
        if target <= 0:
            target = max(int(r.get('shelf_qty') or 0), 0)
        if trigger <= 0:
            trigger = max(int(round(target * 0.4)), 1 if target > 0 else 0)
        shelf_qty = int(r.get('shelf_qty') or 0)
        reserve_qty = int(r.get('reserve_qty') or 0)
        need = max(target - shelf_qty, 0)
        suggested = min(need, reserve_qty)
        status = '十分'
        if shelf_qty <= 0 and reserve_qty <= 0:
            status = '欠品'
        elif shelf_qty <= trigger:
            status = '要補充'
        elif shelf_qty < target:
            status = '補充推奨'
        if status != '十分':
            out.append({**dict(r), 'shelf_target':target, 'shelf_trigger':trigger, 'suggested_replenish_qty':suggested, 'status':status})
    order={'欠品':0,'要補充':1,'補充推奨':2,'十分':3}
    out.sort(key=lambda x:(order.get(x['status'],9), x['supplier_cd'], x['product_cd']))
    return out


def _excel_bytes_from_rows(title, headers, rows):
    from io import BytesIO
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]
    ws.append(headers)
    for row in rows:
        ws.append(row)
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            val = '' if cell.value is None else str(cell.value)
            if len(val) > max_len:
                max_len = len(val)
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 40)
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio.getvalue()

# ─── CSV進捗ストア ────────────────────────────────────────────────
_csv_progress = {}   # job_id -> list of progress events
_csv_lock = threading.Lock()

def _csv_progress_push(job_id, event):
    with _csv_lock:
        if job_id not in _csv_progress:
            _csv_progress[job_id] = []
        _csv_progress[job_id].append(event)


# ─── 認証ヘルパー ────────────────────────────────────────────────
def _hash(pw): return hashlib.sha256(pw.encode()).hexdigest()

def current_user():
    return session.get('user')

def login_required(f):
    from functools import wraps
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get('user'):
            flash('ログインしてください。', 'warning')
            return redirect(url_for('auth.login'))
        return f(*args, **kwargs)
    return decorated

def admin_required(f):
    from functools import wraps
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get('user'):
            flash('ログインしてください。', 'warning')
            return redirect(url_for('auth.login'))
        if session.get('role') != 'admin':
            flash('管理者権限が必要です。', 'danger')
            return redirect(url_for('dashboard.dashboard'))
        return f(*args, **kwargs)
    return decorated

app.jinja_env.globals['current_user'] = current_user

# ─── ページ権限リスト ────────────────────────────────────────────────
PAGE_PERMISSIONS = [
    ('dashboard',     'ホーム'),
    ('inventory',     '在庫一覧'),
    ('receipt',       '入庫'),
    ('orders',        '発注'),
    ('order_history', '発注履歴'),
    ('stocktake',     '棚卸'),
    ('reports',       'レポート'),
    ('forecast',      'AI予測'),
    ('products',      '商品管理'),
    ('csv',           'CSV取込'),
    ('chains',        'チェーン管理'),
    ('recipients',    'メール宛先'),
    ('users',         'ユーザー管理'),
    ('settings',      '設定'),
]

def has_permission(perm):
    if session.get('role') == 'admin':
        return True
    return perm in session.get('permissions', '').split(',')

def permission_required(perm):
    from functools import wraps
    def decorator(f):
        @wraps(f)
        def decorated(*args, **kwargs):
            if not session.get('user'):
                return redirect(url_for('auth.login'))
            if has_permission(perm):
                return f(*args, **kwargs)
            flash('このページへのアクセス権限がありません', 'danger')
            _ep = {
                'dashboard': 'dashboard.dashboard', 'inventory': 'inventory.inventory',
                'orders': 'orders.orders', 'order_history': 'orders.order_history',
                'products': 'products.products', 'chains': 'chains.chains',
            }
            if perm == 'dashboard' or not has_permission('dashboard'):
                # ダッシュボード権限がない場合は最初に許可されたページへ
                for p, _ in PAGE_PERMISSIONS:
                    if p != 'dashboard' and has_permission(p):
                        return redirect(url_for(_ep.get(p, p)))
                return redirect(url_for('auth.login'))
            return redirect(url_for('dashboard.dashboard'))
        return decorated
    return decorator

@app.context_processor
def inject_permissions():
    return dict(has_permission=has_permission, PAGE_PERMISSIONS=PAGE_PERMISSIONS)


# --- グローバルエラーハンドラー ---
@app.errorhandler(500)
def internal_error(e):
    import traceback, sys
    tb = traceback.format_exc()
    print("500 ERROR:", tb, flush=True)
    logger.error("500 ERROR: %s", tb)
    orig = getattr(e, 'original_exception', e)
    err_str = str(orig)
    is_db = any(w in err_str.lower() for w in
                ['connection refused','could not connect','authentication failed',
                 'pg_host','password','psycopg2'])
    _style = (
        "<!DOCTYPE html><html lang=\"ja\"><head><meta charset=\"utf-8\">"
        "<title>エラー</title>"
        "<style>body{font-family:sans-serif;background:#fef2f2;display:flex;"
        "align-items:center;justify-content:center;min-height:100vh;margin:0;padding:20px}"
        ".box{background:#fff;border:2px solid #fca5a5;border-radius:12px;"
        "padding:40px;max-width:620px;width:100%}"
        "h2{color:#dc2626;margin-top:0}"
        ".steps{background:#eff6ff;border-radius:8px;padding:16px 20px;margin-top:16px}"
        "li{margin:10px 0;font-size:14px;line-height:1.6}"
        "a.r{display:inline-block;margin-top:20px;padding:10px 24px;"
        "background:#2563eb;color:#fff;border-radius:8px;text-decoration:none}"
        "</style></head><body><div class=\"box\">"
    )
    _footer = '<a href=\"/\" class=\"r\">再読み込み</a></div></body></html>'
    is_file = isinstance(orig, (PermissionError, FileNotFoundError)) or \
              any(w in err_str.lower() for w in ['no such file'])
    is_encode = isinstance(orig, UnicodeDecodeError) or \
                any(w in err_str.lower() for w in ['codec', 'encoding'])
    if is_db:
        html = (
            _style
            + "<h2>PostgreSQL 接続エラー</h2>"
            "<p style=\"color:#6b7280;font-size:14px\">データベースに接続できませんでした。<br>"
            "サーバーの .env ファイルの接続情報を確認してください。</p>"
            '<div class=\"steps\"><strong>修正手順:</strong><ol>'
            "<li>サーバー上の .env ファイルをテキストエディタで開く</li>"
            "<li>PG_HOST・PG_PORT・PG_DBNAME・PG_USER・PG_PASSWORD を確認・修正する</li>"
            "<li>PostgreSQL サービスが起動しているか確認する</li>"
            "<li>下の「再読み込み」をクリック（サーバー再起動は不要）</li>"
            "</ol></div>"
            + _footer
        )
        return html, 500
    if is_file:
        html = (
            _style
            + "<h2>ファイルアクセスエラー</h2>"
            "<p style=\"color:#6b7280;font-size:14px\">ファイルの読み書き中にエラーが発生しました。<br>"
            "アップロードしたファイルのパスやアクセス権限を確認してください。</p>"
            '<div class=\"steps\"><strong>確認事項:</strong><ol>'
            "<li>対象ファイルが存在するか確認する</li>"
            "<li>ファイルが他のアプリケーションで開かれていないか確認する</li>"
            "<li>サーバーのファイルアクセス権限を確認する</li>"
            "<li>問題が解決しない場合はシステム管理者に連絡してください</li>"
            "</ol></div>"
            + _footer
        )
        return html, 500
    if is_encode:
        html = (
            _style
            + "<h2>文字コードエラー</h2>"
            "<p style=\"color:#6b7280;font-size:14px\">ファイルの文字コード変換中にエラーが発生しました。<br>"
            "インポートするファイルの文字コードを確認してください。</p>"
            '<div class=\"steps\"><strong>確認事項:</strong><ol>'
            "<li>CSVファイルの文字コードが UTF-8 または Shift-JIS であることを確認する</li>"
            "<li>Excelで開いて「名前を付けて保存」→「CSV UTF-8(BOM付き)」で保存し直す</li>"
            "<li>特殊文字・絵文字が含まれていないか確認する</li>"
            "<li>問題が解決しない場合はシステム管理者に連絡してください</li>"
            "</ol></div>"
            + _footer
        )
        return html, 500
    html = (
        _style
        + "<h2>内部エラーが発生しました</h2>"
        f"<p style=\"color:#6b7280;font-size:14px\">予期しないエラーが発生しました。<br>"
        f"エラー内容: <code style=\"background:#f3f4f6;padding:2px 6px;border-radius:4px\">"
        f"{err_str[:200]}</code></p>"
        '<div class=\"steps\"><strong>対処方法:</strong><ol>'
        "<li>ページを再読み込みして再度お試しください</li>"
        "<li>操作の内容とエラーメッセージをメモしてシステム管理者に連絡してください</li>"
        "<li>サーバーのログファイルで詳細なエラー情報を確認してください</li>"
        "</ol></div>"
        + _footer
    )
    return html, 500





# ─── 在庫一覧 ───────────────────────────────────────────────────


# ─── 商品管理 ───────────────────────────────────────────────────


# ─── フォルダ参照API ──────────────────────────────────────────────
@app.route('/api/browse_folders')
@admin_required
def api_browse_folders():
    """サーバー側のフォルダ一覧を返す（フォルダ参照ダイアログ用）"""
    import os as _os
    base = request.args.get('path', '')

    # デフォルト: Windowsのドライブ一覧 or Linux root
    if not base:
        import platform
        if platform.system() == 'Windows':
            drives = []
            for d in 'CDEFGHIJKLMNOPQRSTUVWXYZ':
                if _os.path.exists(f'{d}:'):
                    drives.append({'name': f'{d}:', 'path': f'{d}:\\', 'type': 'drive'})
            # UNCパスも候補に（よく使うルート）
            return jsonify({'current': '', 'parent': None, 'items': drives})
        else:
            base = '/'

    base = base.replace('/', _os.sep)
    try:
        entries = []
        with _os.scandir(base) as it:
            for entry in sorted(it, key=lambda e: e.name.lower()):
                if entry.is_dir(follow_symlinks=False):
                    entries.append({
                        'name': entry.name,
                        'path': entry.path,
                        'type': 'folder'
                    })
        parent = str(Path(base).parent) if Path(base).parent != Path(base) else None
        return jsonify({'current': base, 'parent': parent, 'items': entries})
    except PermissionError:
        return jsonify({'error': 'アクセス権限がありません', 'current': base, 'parent': str(Path(base).parent), 'items': []})
    except Exception as e:
        return jsonify({'error': str(e), 'current': base, 'parent': None, 'items': []})


@app.route('/api/net_use_browse', methods=['POST'])
@admin_required
def api_net_use_browse():
    """net use でUNC接続してフォルダ一覧を返す"""
    import subprocess as _sp
    data     = request.get_json() or {}
    unc_base = (data.get('path') or '').strip()
    net_user = (data.get('net_user') or '').strip()
    net_pass = (data.get('net_pass') or '').strip()

    if not unc_base:
        return jsonify({'error': 'パスを入力してください'})

    # サーバー\共有 部分だけ接続
    import os as _os
    from auto_check import _unc_server, _net_use_connect, _net_use_disconnect
    ok, unc_or_err = _net_use_connect(unc_base, net_user, net_pass)
    if not ok:
        return jsonify({'error': unc_or_err})

    try:
        entries = []
        with _os.scandir(unc_base) as it:
            for entry in sorted(it, key=lambda e: e.name.lower()):
                if entry.is_dir(follow_symlinks=False):
                    entries.append({'name': entry.name, 'path': entry.path, 'type': 'folder'})
        parent = str(Path(unc_base).parent) if Path(unc_base).parent != Path(unc_base) else None
        return jsonify({'current': unc_base, 'parent': parent, 'items': entries})
    except PermissionError:
        return jsonify({'error': 'アクセス権限がありません（ユーザー名・パスワードを確認）', 'current': unc_base, 'parent': None, 'items': []})
    except Exception as e:
        return jsonify({'error': str(e), 'current': unc_base, 'parent': None, 'items': []})
    finally:
        _net_use_disconnect(_unc_server(unc_base) if not ok else unc_or_err)

@app.route('/api/csv_filter_values', methods=['POST'])
@admin_required
def api_csv_filter_values():
    """フォルダ内の最新CSVから指定列の値一覧を返す"""
    import glob as _glob
    data    = request.get_json() or {}
    folder  = (data.get('folder') or '').strip()
    col     = (data.get('col') or '').strip()
    enc     = (data.get('encoding') or 'utf-8-sig').strip()
    if not folder or not col:
        return jsonify({'error': 'フォルダパスとフィルター列名を入力してください'})
    # フォルダ内の最新CSVを1つ取得
    import os as _os
    patterns = [os.path.join(folder, '*.csv'), os.path.join(folder, '*.CSV')]
    files = []
    for p in patterns:
        files.extend(_glob.glob(p))
    if not files:
        return jsonify({'error': f'CSVファイルが見つかりません: {folder}'})
    latest = max(files, key=_os.path.getmtime)
    try:
        values = set()
        for enc_try in [enc, 'cp932', 'utf-8-sig', 'utf-8']:
            try:
                with open(latest, encoding=enc_try, errors='replace') as f:
                    reader = csv.DictReader(f)
                    if col not in (reader.fieldnames or []):
                        continue
                    for row in reader:
                        v = str(row.get(col, '') or '').strip()
                        if v:
                            values.add(v)
                break
            except Exception:
                continue
        if not values:
            return jsonify({'error': f'列「{col}」が見つかりません（ファイル: {_os.path.basename(latest)}）'})
        return jsonify({'values': sorted(values), 'file': _os.path.basename(latest)})
    except Exception as e:
        return jsonify({'error': str(e)})


# ─── チェーン管理 ───────────────────────────────────────────────────────


# ─── 商品マスタ インポート・エクスポート ──────────────────────────
# エクスポート列定義（JANがキー、この順でExcel/CSV出力）
_PRODUCT_COLS = [
    ('jan',            'JANコード',          'text'),
    ('product_cd',     '商品コード',          'text'),
    ('product_name',   '商品名',              'text'),
    ('supplier_cd',    '仕入先コード',        'text'),
    ('supplier_name',  '仕入先名',            'text'),
    ('unit_qty',       '入数',                'int'),
    ('order_unit',     '発注単位',            'int'),
    ('order_qty',      '発注数量',            'int'),
    ('reorder_point',  '発注点',              'int'),
    ('reorder_auto',   '発注点自動更新',      'int'),
    ('lead_time_days', 'リードタイム日数',    'int'),
    ('safety_factor',  '安全係数',            'float'),
    ('lot_size',       'メーカーロット数',    'int'),
    ('shelf_life_days','賞味期限日数',        'int'),
    ('expiry_alert_days','期限アラート日数',  'int'),
    ('mixed_group',    '混載グループ名',      'text'),
    ('mixed_lot_mode', '混載ロットルール',    'text'),
    ('mixed_lot_cases','混載ケース数',        'int'),
    ('mixed_force_days','強制発注日数',       'int'),
    ('cost_price',     '原価',                'float'),
    ('sell_price',     '売価',                'float'),
]

_EXPORT_ROW_LIMIT = 10_000  # エクスポート上限（OOM防止）



# ─── 入庫登録 ───────────────────────────────────────────────────



@app.route('/receipt/import/template')
@login_required
def receipt_import_template():
    import io, csv
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(['JANコード', '商品CD', '数量', '賞味期限', 'ロット番号'])

    output.seek(0)
    _db4 = get_db()
    _rcpt_name = _db4.execute("SELECT value FROM settings WHERE key='receipt_template_name'").fetchone()
    _rcpt_name = (_rcpt_name['value'] if _rcpt_name else '入庫一括インポート_テンプレート') + '.csv'
    from urllib.parse import quote
    return Response(
        output.getvalue().encode('utf-8-sig'),
        mimetype='text/csv',
        headers={'Content-Disposition': f"attachment; filename*=UTF-8''{quote(_rcpt_name)}"}
    )


@app.route('/receipt/import', methods=['GET','POST'])
@login_required
def receipt_import():
    if request.method == 'GET':
        return render_template('receipt_import.html')
    
    f = request.files.get('file')
    if not f:
        flash('ファイルを選択してください', 'danger')
        return redirect(url_for('receipt_import'))

    filename = f.filename.lower()
    db = get_db()
    ok_count = 0
    err_rows = []

    try:
        if filename.endswith('.csv'):
            import csv, io
            content = f.read().decode('utf-8-sig', errors='ignore')
            reader = csv.DictReader(io.StringIO(content))
            rows = list(reader)
        elif filename.endswith(('.xlsx', '.xls')):
            import openpyxl
            wb = openpyxl.load_workbook(f, data_only=True)
            ws = wb.active
            headers = [str(ws.cell(1, c).value or '').strip() for c in range(1, ws.max_column+1)]
            rows = []
            for r in range(2, ws.max_row+1):
                row = {headers[c-1]: ws.cell(r, c).value for c in range(1, ws.max_column+1)}
                if any(v for v in row.values()):
                    rows.append(row)
        else:
            flash('CSV または Excel ファイルを選択してください', 'danger')
            return redirect(url_for('receipt_import'))

        for i, row in enumerate(rows, 2):
            jan_or_cd  = str(row.get('JANコード') or row.get('JAN') or row.get('商品CD') or row.get('商品コード') or '').strip()
            qty_raw    = row.get('数量') or row.get('入庫数量') or ''
            expiry     = str(row.get('賞味期限') or '').strip()
            lot_no     = str(row.get('ロット番号') or row.get('ロットNo') or '').strip()

            if not jan_or_cd:
                continue
            try:
                qty = int(float(str(qty_raw)))
            except Exception:
                err_rows.append(f'{i}行目: 数量が不正 ({qty_raw})')
                continue
            if qty <= 0:
                err_rows.append(f'{i}行目: 数量が0以下')
                continue

            product = db.execute(
                "SELECT * FROM products WHERE (jan=%s OR product_cd=%s) AND is_active=1",
                [jan_or_cd, jan_or_cd]
            ).fetchone()
            if not product:
                err_rows.append(f'{i}行目: JAN/商品CD「{jan_or_cd}」未登録')
                continue

            # 賞味期限フォーマット整形
            if expiry:
                expiry = expiry.replace('/', '-')
                parts = expiry.split('-')
                if len(parts) == 3:
                    expiry = f'{parts[0]}-{int(parts[1]):02d}-{int(parts[2]):02d}'
                elif len(expiry) == 8 and expiry.isdigit():
                    expiry = f'{expiry[:4]}-{expiry[4:6]}-{expiry[6:]}'
            else:
                err_rows.append(f'{i}行目: 賞味期限が未入力（{product["product_name"]}）')
                continue

            # 重複チェック（ファイル名＋行番号＋JAN＋数量＋賞味期限）
            import hashlib as _hl
            row_hash = _hl.md5(
                f"{f.filename}|{i}|{product['jan']}|{qty}|{expiry}|{lot_no}".encode()
            ).hexdigest()
            dup = db.execute(
                "SELECT 1 FROM stock_movements WHERE note=%s",
                [f'receipt_hash:{row_hash}']
            ).fetchone()
            if dup:
                err_rows.append(f'{i}行目: 重複のためスキップ（{product["product_name"]}）')
                continue

            location_code = str(row.get('ロケーション') or row.get('棚番') or row.get('location_code') or product.get('location_code') or '').strip()
            _record_receipt(db, product, qty, expiry, lot_no, location_code, f.filename, f'receipt_hash:{row_hash}')
            db.execute("UPDATE products SET ordered_at='' WHERE jan=%s", [product['jan']])
            db.execute("DELETE FROM order_pending WHERE jan=%s AND status='pending'", [product['jan']])
            ok_count += 1

        db.commit()
    except Exception as e:
        db.rollback()
        flash(f'インポートエラー: {e}', 'danger')
        return redirect(url_for('receipt_import'))

    msg = f'{ok_count}件の入庫を登録しました。'
    if err_rows:
        flash(msg + f' （エラー/スキップ {len(err_rows)}件）', 'warning')
        # セッションにエラー詳細を保存（4KB制限対策：最大20件・120文字以内）
        session['import_errors'] = [str(e)[:120] for e in err_rows[:20]]
    else:
        flash(msg, 'success')
        session.pop('import_errors', None)
    return redirect(url_for('receipt_import'))


@app.route('/receipt/import/clear_errors')
@login_required
def receipt_import_clear_errors():
    session.pop('import_errors', None)
    return redirect(url_for('receipt_import'))

@app.route('/receipt', methods=['GET','POST'])
@permission_required('receipt')
def receipt():
    db = get_db()
    if request.method == 'POST':
        f = request.form
        product = db.execute("SELECT * FROM products WHERE jan=%s AND is_active=1",[f['jan']]).fetchone()
        if not product:
            flash('JANコードが見つかりません。', 'danger')
        else:
            qty = int(f['quantity'])
            expiry = f.get('expiry_date','').strip()
            lot_no = f.get('lot_no','').strip()
            location_code = f.get('location_code','').strip() or product.get('location_code','')

            if not expiry:
                flash('賞味期限は必須です。入力してください。', 'danger')
                return redirect(url_for('receipt'))

            # 重複チェック（同日・同JAN・同数量・同賞味期限・同ロットの手動入庫）
            import hashlib as _hl
            row_hash = _hl.md5(
                f"manual|{str(date.today())}|{product['jan']}|{qty}|{expiry}|{lot_no}".encode()
            ).hexdigest()
            dup = db.execute(
                "SELECT 1 FROM stock_movements WHERE note=%s",
                [f'receipt_hash:{row_hash}']
            ).fetchone()
            if dup:
                flash(f"⚠️ 同じ内容がすでに入庫済みです（{product['product_name']} {qty}個）。重複のためスキップしました。", 'warning')
                return redirect(url_for('receipt'))

            _record_receipt(db, product, qty, expiry, lot_no, location_code, 'manual', f'receipt_hash:{row_hash}')
            # 入庫で発注済みフラグをクリア
            db.execute("UPDATE products SET ordered_at='' WHERE jan=%s", [product['jan']])
            db.execute("DELETE FROM order_pending WHERE jan=%s AND status='pending'", [product['jan']])
            db.commit()
            flash(f"{product['product_name']} を {qty} 個入庫しました。", 'success')
        return redirect(url_for('receipt'))
    prods = db.execute("SELECT * FROM products WHERE is_active=1 ORDER BY CAST(NULLIF(regexp_replace(supplier_cd,'[^0-9]','','g'),'') AS BIGINT) NULLS LAST, CAST(NULLIF(regexp_replace(product_cd,'[^0-9]','','g'),'') AS BIGINT) NULLS LAST").fetchall()
    return render_template('receipt.html', products=prods)

@app.route('/receipt/history')
@permission_required('receipt')
def receipt_history():
    db = get_db()
    q = request.args.get('q','').strip()
    page = max(1, int(request.args.get('page', 1)))
    per = 50
    rows = db.execute("""
        SELECT m.*, p.supplier_cd, p.supplier_name, p.product_cd
        FROM stock_movements m
        LEFT JOIN products p ON m.jan=p.jan
        WHERE m.move_type='receipt'
        ORDER BY m.created_at DESC
        LIMIT %s OFFSET %s
    """, [per, (page-1)*per]).fetchall()
    total = db.execute(
        "SELECT COUNT(*) AS c FROM stock_movements WHERE move_type='receipt'"
    ).fetchone()['c']
    if q:
        rows = [r for r in rows if q.lower() in (r['jan'] or '').lower()
                or q.lower() in (r['product_cd'] or '').lower()
                or q.lower() in (r['product_name'] or '').lower()
                or q.lower() in (r['supplier_cd'] or '').lower()
                or q.lower() in (r['supplier_name'] or '').lower()]
    pages = (total + per - 1) // per
    return render_template('receipt_history.html', rows=rows, q=q,
                           page=page, pages=pages, total=total)


@app.route('/receipt/history/<int:mid>/delete', methods=['POST'])
@permission_required('receipt')
def receipt_history_delete(mid):
    db = get_db()
    mv = db.execute("SELECT * FROM stock_movements WHERE id=%s AND move_type='receipt'", [mid]).fetchone()
    if not mv:
        flash('該当する入庫履歴が見つかりません。', 'danger')
        return redirect(url_for('receipt_history'))
    # 在庫から数量を戻す（FIFOの逆：賞味期限が新しいものから引く）
    qty_to_restore = mv['quantity']
    stocks = db.execute("""
        SELECT * FROM stocks WHERE jan=%s AND quantity>0
        ORDER BY CASE WHEN expiry_date='' THEN '0000-00-00' ELSE expiry_date END DESC
    """, [mv['jan']]).fetchall()
    for s in stocks:
        if qty_to_restore <= 0:
            break
        deduct = min(s['quantity'], qty_to_restore)
        db.execute("UPDATE stocks SET quantity=quantity-%s WHERE id=%s", [deduct, s['id']])
        qty_to_restore -= deduct
    # 在庫0のレコードを削除
    db.execute("DELETE FROM stocks WHERE jan=%s AND quantity<=0", [mv['jan']])
    # 入庫履歴を削除
    db.execute("DELETE FROM stock_movements WHERE id=%s", [mid])
    db.commit()
    flash(f"{mv['product_name']} の入庫履歴を削除し、在庫を {mv['quantity']} 個戻しました。", 'success')
    return redirect(url_for('receipt_history'))


# ─── 発注チェック ────────────────────────────────────────────────




# ─── 発注履歴 ────────────────────────────────────────────────────




# ─── 混載ペンディング管理 ──────────────────────────────────────












# ─── フォルダパス候補API ─────────────────────────────────────────

# ─── CSV取込設定 ─────────────────────────────────────────────────
@app.route('/csv')
@permission_required('csv')
def csv_settings():
    db = get_db()
    settings = db.execute("SELECT * FROM csv_import_settings ORDER BY id").fetchall()
    # CSV取込ログ保持月数
    csv_log_row = db.execute("SELECT value FROM settings WHERE key='csv_log_months'").fetchone()
    csv_log_months = int(csv_log_row['value']) if csv_log_row else 6

    logs = db.execute("""
        SELECT l.*, s.name as setting_name FROM import_logs l
        LEFT JOIN csv_import_settings s ON l.setting_id=s.id
        ORDER BY l.imported_at DESC LIMIT 200
    """).fetchall()
    return render_template('csv_settings.html', settings=settings, logs=logs,
                           today=str(date.today()),
                           today_ym=date.today().strftime('%Y-%m'),
                           csv_log_months=csv_log_months)

@app.route('/csv/logs/delete', methods=['POST'])
@admin_required
def csv_logs_delete():
    db = get_db()
    mode = request.form.get('mode', 'all')
    try:
        if mode == 'all':
            db.execute("DELETE FROM import_logs")
            flash('CSV取込ログをすべて削除しました', 'success')
        elif mode == 'months':
            months = int(request.form.get('months', 3))
            db.execute(
                "DELETE FROM import_logs WHERE imported_at < NOW() - (INTERVAL '1 month' * %s)",
                [months]
            )
            flash(f'{months}ヶ月以上前のCSV取込ログを削除しました', 'success')
        db.commit()
    except Exception as e:
        db.rollback()
        flash(f'エラー: {e}', 'error')
    return redirect(url_for('csv_settings'))

@app.route('/csv/logs/retention', methods=['POST'])
@admin_required
def csv_logs_retention():
    db = get_db()
    months = request.form.get('csv_log_months', '6').strip()
    try:
        months = str(int(months))
        existing = db.execute("SELECT id FROM settings WHERE key='csv_log_months'").fetchone()
        if existing:
            db.execute("UPDATE settings SET value=%s WHERE key='csv_log_months'", [months])
        else:
            db.execute("INSERT INTO settings (key, value) VALUES ('csv_log_months',%s)", [months])
        db.commit()
        flash(f'CSV取込ログの表示・保持期間を{months}ヶ月に設定しました', 'success')
    except Exception as e:
        db.rollback()
        flash(f'エラー: {e}', 'error')
    return redirect(url_for('csv_settings'))


@app.route('/csv/import_detail')
@permission_required('csv')
def csv_import_detail():
    """インポートファイル別の明細確認・選択削除画面"""
    db = get_db()
    # クエリパラメータ
    source_file = request.args.get('source_file', '').strip()
    q           = request.args.get('q', '').strip()
    page        = max(1, _to_int(request.args.get('page', 1), 1))
    per_page    = 50

    # ファイル一覧（ページネーション・検索付き）
    file_q     = request.args.get('file_q', '').strip()
    file_page  = max(1, _to_int(request.args.get('file_page', 1), 1))
    file_per   = 30

    # ── ファイル一覧（import_logsのfilenameでグループ集計）──
    file_where = "WHERE 1=1"
    file_params = []
    if file_q:
        file_where += " AND l.filename ILIKE %s"
        file_params.append(f'%{file_q}%')

    file_total_row = db.execute(f"""
        SELECT COUNT(DISTINCT l.filename) AS cnt
        FROM import_logs l
        {file_where}
    """, file_params).fetchone()
    file_total = int(file_total_row['cnt'] or 0)
    file_pages = max(1, (file_total + file_per - 1) // file_per)
    file_page  = min(file_page, file_pages)

    file_list = db.execute(f"""
        SELECT l.filename,
               MAX(l.imported_at) AS last_imported_at,
               SUM(l.rows_ok)     AS total_rows_ok,
               COUNT(l.id)        AS import_count,
               MAX(s.name)        AS setting_name,
               (SELECT COUNT(*) FROM sales_history sh WHERE sh.source_file = l.filename) AS sales_count
        FROM import_logs l
        LEFT JOIN csv_import_settings s ON s.id = l.setting_id
        {file_where}
        GROUP BY l.filename
        ORDER BY last_imported_at DESC
        LIMIT %s OFFSET %s
    """, file_params + [file_per, (file_page - 1) * file_per]).fetchall()

    # ── 選択ファイルの明細 ──
    detail_rows = []
    detail_total = 0
    detail_pages = 1
    if source_file:
        detail_where = "WHERE sh.source_file = %s"
        detail_params = [source_file]
        if q:
            detail_where += " AND (sh.jan ILIKE %s OR sh.product_name ILIKE %s OR sh.store_name ILIKE %s OR sh.client_name ILIKE %s)"
            detail_params += [f'%{q}%', f'%{q}%', f'%{q}%', f'%{q}%']

        detail_total_row = db.execute(f"""
            SELECT COUNT(*) AS cnt FROM sales_history sh {detail_where}
        """, detail_params).fetchone()
        detail_total = int(detail_total_row['cnt'] or 0)
        detail_pages = max(1, (detail_total + per_page - 1) // per_page)
        page = min(page, detail_pages)

        detail_rows = db.execute(f"""
            SELECT sh.id, sh.jan, sh.product_name, sh.quantity, sh.sale_date,
                   sh.chain_cd, sh.client_name, sh.store_cd, sh.store_name,
                   sh.created_at
            FROM sales_history sh
            {detail_where}
            ORDER BY sh.sale_date DESC, sh.jan
            LIMIT %s OFFSET %s
        """, detail_params + [per_page, (page - 1) * per_page]).fetchall()

    return render_template('csv_import_detail.html',
        file_list=file_list,
        file_total=file_total,
        file_pages=file_pages,
        file_page=file_page,
        file_per=file_per,
        file_q=file_q,
        source_file=source_file,
        detail_rows=detail_rows,
        detail_total=detail_total,
        detail_pages=detail_pages,
        page=page,
        per_page=per_page,
        q=q,
    )


@app.route('/csv/import_detail/delete', methods=['POST'])
@permission_required('csv')
def csv_import_detail_delete():
    """インポート明細の選択削除"""
    db = get_db()
    source_file = request.form.get('source_file', '').strip()
    mode        = request.form.get('mode', 'selected')  # selected / file_all
    ids         = [int(x) for x in request.form.getlist('detail_id') if str(x).isdigit()]
    q           = request.form.get('q', '').strip()
    page        = request.form.get('page', '1')
    file_page   = request.form.get('file_page', '1')
    file_q      = request.form.get('file_q', '').strip()

    try:
        if mode == 'file_all' and source_file:
            # ファイル単位で全削除
            db.execute("DELETE FROM sales_history WHERE source_file=%s", [source_file])
            db.commit()
            flash(f'「{source_file}」のインポートデータを全件削除しました。', 'warning')
            return redirect(url_for('csv_import_detail',
                file_q=file_q, file_page=file_page))
        elif mode == 'selected' and ids:
            placeholders = ','.join(['%s'] * len(ids))
            db.execute(f"DELETE FROM sales_history WHERE id IN ({placeholders})", ids)
            db.commit()
            flash(f'{len(ids)}件のインポートデータを削除しました。', 'warning')
        else:
            flash('削除対象が選択されていません。', 'warning')
    except Exception as e:
        db.rollback()
        flash(f'削除エラー: {e}', 'danger')

    return redirect(url_for('csv_import_detail',
        source_file=source_file, q=q, page=page,
        file_q=file_q, file_page=file_page))


@app.route('/csv/new', methods=['GET','POST'])
@admin_required
def csv_new():
    if request.method == 'POST':
        f = request.form
        db = get_db()
        db.execute("""
            INSERT INTO csv_import_settings
            (name,import_type,folder_path,filename_pattern,encoding,
             col_jan,col_qty,col_date,col_expiry,col_slip_no,col_chain_cd,col_row_no,
             col_filter_cd,filter_cd_values,net_user,net_pass,
             run_times,run_hour,run_minute,is_active,
             month_end_enabled,month_end_folder,month_end_pattern,month_end_date_col)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
        """, [f['name'],f.get('import_type','sales'),f['folder_path'],
              f.get('filename_pattern','*{yyyymmdd}.csv'),
              f.get('encoding','utf-8-sig'),
              f.get('col_jan','JANコード'),f.get('col_qty','数量'),
              f.get('col_date','納品日'),f.get('col_expiry','賞味期限'),
              f.get('col_slip_no','伝票番号'),
              f.get('col_chain_cd','チェーンCD'),f.get('col_row_no','行番号'),
              f.get('col_filter_cd','担当CD'),f.get('filter_cd_values',''),
              f.get('net_user',''),f.get('net_pass',''),
              f.get('run_times','06:00'),
              int(f.get('run_hour',6)),int(f.get('run_minute',0)),
              1 if f.get('is_active') else 0,
              1 if f.get('month_end_enabled') else 0,
              f.get('month_end_folder',''),
              f.get('month_end_pattern','{yyyymm}_売上実績.csv'),
              f.get('month_end_date_col','納品日')])
        db.commit()
        flash('CSV取込設定を保存しました。', 'success')
        return redirect(url_for('csv_settings'))
    return render_template('csv_form.html', setting=None, env=dict(os.environ))

@app.route('/csv/<int:sid>/edit', methods=['GET','POST'])
def csv_edit(sid):
    db = get_db()
    setting = db.execute("SELECT * FROM csv_import_settings WHERE id=%s",[sid]).fetchone()
    if not setting: return redirect(url_for('csv_settings'))
    if request.method == 'POST':
        f = request.form
        new_type = f.get('import_type', 'sales')
        old_type = setting['import_type'] if setting else 'sales'
        # record_only → sales/receipt に変更した場合は自動でis_activeを有効化
        is_active_val = 1 if f.get('is_active') else 0
        if old_type == 'record_only' and new_type in ('sales', 'receipt'):
            is_active_val = 1
            flash('取込種別を在庫処理モードに変更しました。自動取込を有効にしました。', 'success')
        db.execute("""
            UPDATE csv_import_settings SET
            name=%s,import_type=%s,folder_path=%s,filename_pattern=%s,encoding=%s,
            col_jan=%s,col_qty=%s,col_date=%s,col_expiry=%s,col_slip_no=%s,col_chain_cd=%s,col_row_no=%s,
            col_filter_cd=%s,filter_cd_values=%s,net_user=%s,net_pass=%s,
            run_times=%s,run_hour=%s,run_minute=%s,is_active=%s,
            month_end_enabled=%s,month_end_folder=%s,
            month_end_pattern=%s,month_end_date_col=%s
            WHERE id=%s
        """, [f['name'],f.get('import_type','sales'),f['folder_path'],
              f.get('filename_pattern','*{yyyymmdd}.csv'),
              f.get('encoding','utf-8-sig'),
              f.get('col_jan','JANコード'),f.get('col_qty','数量'),
              f.get('col_date','納品日'),f.get('col_expiry','賞味期限'),
              f.get('col_slip_no','伝票番号'),
              f.get('col_chain_cd','チェーンCD'),f.get('col_row_no','行番号'),
              f.get('col_filter_cd','担当CD'),f.get('filter_cd_values',''),
              f.get('net_user',''),f.get('net_pass',''),
              f.get('run_times','06:00'),
              int(f.get('run_hour',6)),int(f.get('run_minute',0)),
              is_active_val,
              1 if f.get('month_end_enabled') else 0,
              f.get('month_end_folder',''),
              f.get('month_end_pattern','{yyyymm}_売上実績.csv'),
              f.get('month_end_date_col','納品日'),
              sid])
        db.commit()
        # 月末月次取込時刻も更新
        me_h = f.get('MONTH_END_IMPORT_HOUR', '').strip()
        me_m = f.get('MONTH_END_IMPORT_MINUTE', '').strip()
        if me_h or me_m:
            env_path = Path(os.path.join(os.path.dirname(__file__), '.env'))
            if env_path.exists():
                env_lines = env_path.read_text(encoding='utf-8').splitlines()
                new_lines = []
                updated_h = updated_m = False
                for line in env_lines:
                    if line.startswith('MONTH_END_IMPORT_HOUR=') and me_h:
                        new_lines.append(f'MONTH_END_IMPORT_HOUR={me_h}')
                        updated_h = True
                    elif line.startswith('MONTH_END_IMPORT_MINUTE=') and me_m:
                        new_lines.append(f'MONTH_END_IMPORT_MINUTE={me_m}')
                        updated_m = True
                    else:
                        new_lines.append(line)
                if me_h and not updated_h:
                    new_lines.append(f'MONTH_END_IMPORT_HOUR={me_h}')
                if me_m and not updated_m:
                    new_lines.append(f'MONTH_END_IMPORT_MINUTE={me_m}')
                env_path.write_text('\n'.join(new_lines), encoding='utf-8')
                os.environ['MONTH_END_IMPORT_HOUR'] = me_h or os.getenv('MONTH_END_IMPORT_HOUR','5')
                os.environ['MONTH_END_IMPORT_MINUTE'] = me_m or os.getenv('MONTH_END_IMPORT_MINUTE','0')
        if old_type != 'record_only' or new_type == 'record_only':
            flash('設定を更新しました。', 'success')
        return redirect(url_for('csv_settings'))
    return render_template('csv_form.html', setting=setting, env=dict(os.environ))

@app.route('/csv/<int:sid>/run', methods=['POST'])
def csv_run(sid):
    import uuid
    job_id = str(uuid.uuid4())
    target_ym       = request.form.get('target_ym', '').strip()
    target_date_str = request.form.get('target_date', '').strip()

    run_all = request.form.get('run_all', '') == '1'

    def _run():
        with _csv_lock:
            _csv_progress[job_id] = []
        def cb(ev):
            _csv_progress_push(job_id, ev)
        if run_all:
            results = run_csv_import(setting_id=sid, progress_cb=cb, trigger_type='manual', all_files=True)
        elif target_ym and len(target_ym) == 6 and target_ym.isdigit():
            results = run_csv_import(setting_id=sid, target_ym=target_ym, progress_cb=cb, trigger_type='manual')
        elif target_date_str:
            try:
                td = date.fromisoformat(target_date_str)
            except ValueError:
                td = date.today()
            results = run_csv_import(setting_id=sid, target_date=td, progress_cb=cb, trigger_type='manual')
        else:
            results = run_csv_import(setting_id=sid, progress_cb=cb, trigger_type='manual')
        _csv_progress_push(job_id, {'phase': 'finished', 'results': results})

    t = threading.Thread(target=_run, daemon=True)
    t.start()
    return redirect(url_for('csv_progress_page', job_id=job_id))


@app.route('/csv/progress/<job_id>')
def csv_progress_page(job_id):
    return render_template('csv_progress.html', job_id=job_id)


@app.route('/csv/progress/<job_id>/stream')
def csv_progress_stream(job_id):
    def generate():
        sent = 0
        for _ in range(3600):  # 最大1時間
            with _csv_lock:
                events = _csv_progress.get(job_id, [])
            while sent < len(events):
                ev = events[sent]
                yield f"data: {json.dumps(ev, ensure_ascii=False)}\n\n"
                sent += 1
                if ev.get('phase') == 'finished':
                    return
            _time.sleep(0.3)
        yield f"data: {json.dumps({'phase': 'timeout'})}\n\n"
    return Response(generate(), mimetype='text/event-stream',
                    headers={'Cache-Control': 'no-cache', 'X-Accel-Buffering': 'no'})

@app.route('/csv/run_all_redirect', methods=['GET','POST'])
@admin_required
def csv_run_all_redirect():
    """CSV設定一覧から最初の有効な設定の進捗モニターへリダイレクト"""
    db = get_db()
    settings = db.execute("SELECT id FROM csv_import_settings WHERE is_active=1 ORDER BY id").fetchall()
    if not settings:
        flash('有効なCSV設定がありません', 'error')
        return redirect(url_for('dashboard.dashboard'))
    sid = settings[0]['id']
    # job_idを生成してバックグラウンド実行し進捗モニターへ
    import uuid, threading
    job_id = str(uuid.uuid4())
    _csv_progress[job_id] = []
    def cb(ev): _csv_progress[job_id].append(ev)
    def run():
        run_csv_import(setting_id=sid, progress_cb=cb, trigger_type='manual')
    threading.Thread(target=run, daemon=True).start()
    return redirect(f'/csv/progress/{job_id}')


@app.route('/csv/run_all', methods=['POST'])
@admin_required
def csv_run_all():
    results = run_csv_import()
    ok = sum(1 for r in results if r['status']=='ok')
    flash(f'CSVインポート完了: {len(results)}設定処理、{ok}件成功。', 'success')
    return redirect(url_for('csv_settings'))

@app.route('/csv/<int:sid>/run_month_end', methods=['POST'])
def csv_run_month_end(sid):
    """月末月次CSV手動取込（進捗モニター対応）"""
    import uuid
    job_id = str(uuid.uuid4())
    target_ym = request.form.get('target_ym', '').strip()
    all_dates = request.form.get('all_dates', '') == '1'
    if not (len(target_ym) == 6 and target_ym.isdigit()):
        target_ym = None

    def _run():
        with _csv_lock:
            _csv_progress[job_id] = []
        def cb(ev):
            _csv_progress_push(job_id, ev)
        results = run_month_end_import(
            setting_id=sid, target_ym=target_ym,
            all_dates=all_dates, progress_cb=cb,
            trigger_type='manual'
        )
        _csv_progress_push(job_id, {'phase': 'finished', 'results': results})

    t = threading.Thread(target=_run, daemon=True)
    t.start()
    return redirect(url_for('csv_progress_page', job_id=job_id))

@app.route('/csv/<int:sid>/reimport_month_end', methods=['POST'])
@login_required
def csv_reimport_month_end(sid):
    """未登録商品スキップ分の月末月次CSV再取込（商品登録後に使用）"""
    import uuid
    job_id = str(uuid.uuid4())
    target_ym = request.form.get('target_ym', '').strip()
    if not (len(target_ym) == 6 and target_ym.isdigit()):
        flash('対象年月(YYYYMM)が不正です。', 'error')
        return redirect(url_for('csv_settings'))

    def _run():
        with _csv_lock:
            _csv_progress[job_id] = []
        def cb(ev):
            _csv_progress_push(job_id, ev)
        # all_dates=True で重複チェックをスキップ（row_hashで二重取込防止）
        results = run_month_end_import(
            setting_id=sid, target_ym=target_ym,
            all_dates=False, progress_cb=cb,
            trigger_type='manual'
        )
        # partial_skip ログを削除して再登録（再取込できるようにリセット）
        _run.results = results
        _csv_progress_push(job_id, {'phase': 'finished', 'results': results})

    # 既存の partial_skip ログを削除して再取込可能にする
    db = get_db()
    ym_str = target_ym
    try:
        yr, mo = int(ym_str[:4]), int(ym_str[4:6])
        import calendar as _cal2
        last_day2 = _cal2.monthrange(yr, mo)[1]
        from datetime import date as _date2
        me_date2 = _date2(yr, mo, last_day2)
        log_key_pat = f"month_end_{ym_str}_%"
        db.execute(
            "DELETE FROM import_logs WHERE setting_id=%s AND filename LIKE %s AND status='partial_skip'",
            [sid, log_key_pat]
        )
        db.commit()
    except Exception:
        pass

    t = threading.Thread(target=_run, daemon=True)
    t.start()
    return redirect(url_for('csv_progress_page', job_id=job_id))

@app.route('/csv/<int:sid>/reimport_daily', methods=['POST'])
@login_required
def csv_reimport_daily(sid):
    """未登録商品スキップ分の日次CSV再取込（商品登録後に使用）"""
    import uuid
    filename = request.form.get('filename', '').strip()
    if not filename:
        flash('ファイル名が指定されていません。', 'error')
        return redirect(url_for('csv_settings'))

    job_id = str(uuid.uuid4())

    def _run():
        with _csv_lock:
            _csv_progress[job_id] = []
        def cb(ev):
            _csv_progress_push(job_id, ev)
        # partial_skip ログを削除して再取込可能にする
        db2 = get_db()
        try:
            db2.execute(
                "DELETE FROM import_logs WHERE setting_id=%s AND filename=%s AND status='partial_skip'",
                [sid, filename]
            )
            db2.commit()
        except Exception:
            pass
        # ファイル名から日付を推測して run_csv_import を呼ぶ
        # ファイル名中の8桁数字を日付として使用
        import re as _re
        m = _re.search(r'(\d{4})(\d{2})(\d{2})', filename)
        if m:
            from datetime import date as _date3
            try:
                tdate = _date3(int(m.group(1)), int(m.group(2)), int(m.group(3)))
            except Exception:
                tdate = None
        else:
            tdate = None
        results = run_csv_import(
            setting_id=sid, target_date=tdate,
            progress_cb=cb, trigger_type='manual', all_files=True
        )
        _csv_progress_push(job_id, {'phase': 'finished', 'results': results})

    t = threading.Thread(target=_run, daemon=True)
    t.start()
    return redirect(url_for('csv_progress_page', job_id=job_id))

@app.route('/csv/run_all_month_end', methods=['POST'])
def csv_run_all_month_end():
    """全設定の月末月次CSV手動取込"""
    target_ym = request.form.get('target_ym', '').strip()
    if not (len(target_ym) == 6 and target_ym.isdigit()):
        target_ym = None
    results = run_month_end_import(target_ym=target_ym)
    ok = sum(1 for r in results if r['status'] == 'ok')
    flash(f'月末月次インポート完了: {len(results)}設定処理、{ok}件成功。', 'success')
    return redirect(url_for('csv_settings'))



# ─── メール受信先管理 ─────────────────────────────────────────────
@app.route('/recipients')
@permission_required('recipients')
def recipients():
    db = get_db()
    rows = db.execute("SELECT * FROM mail_recipients ORDER BY id").fetchall()
    return render_template('recipients.html', recipients=rows)

@app.route('/recipients/new', methods=['POST'])
@admin_required
def recipient_new():
    f = request.form
    db = get_db()
    try:
        send_type = f.get('send_type', 'both')
        db.execute(
            "INSERT INTO mail_recipients (name,email,send_type,is_active) VALUES (%s,%s,%s,1)",
            [f['name'], f['email'], send_type])
        db.commit()
        flash(f"{f['name']} を追加しました。", 'success')
    except Exception as e:
        db.rollback()
        flash(f"追加エラー: {e}", 'danger')
    return redirect(url_for('recipients'))

@app.route('/recipients/<int:rid>/send_type', methods=['POST'])
@admin_required
def recipient_send_type(rid):
    db = get_db()
    send_type = request.form.get('send_type', 'both')
    db.execute("UPDATE mail_recipients SET send_type=%s WHERE id=%s", [send_type, rid])
    db.commit()
    return redirect(url_for('recipients'))

@app.route('/recipients/<int:rid>/toggle', methods=['POST'])
def recipient_toggle(rid):
    db = get_db()
    r = db.execute("SELECT * FROM mail_recipients WHERE id=%s",[rid]).fetchone()
    if r:
        db.execute("UPDATE mail_recipients SET is_active=%s WHERE id=%s",
                   [0 if r['is_active'] else 1, rid])
        db.commit()
    return redirect(url_for('recipients'))

@app.route('/recipients/<int:rid>/delete', methods=['POST'])
def recipient_delete(rid):
    db = get_db()
    db.execute("DELETE FROM mail_recipients WHERE id=%s",[rid])
    db.commit()
    flash('削除しました。', 'success')
    return redirect(url_for('recipients'))

@app.route('/recipients/<int:rid>/supplier_cd', methods=['POST'])
@admin_required
def recipient_supplier_cd(rid):
    db = get_db()
    supplier_cd = request.form.get('supplier_cd', '').strip()
    db.execute("UPDATE mail_recipients SET supplier_cd=%s WHERE id=%s", [supplier_cd, rid])
    db.commit()
    return redirect(url_for('recipients'))

@app.route('/recipients/template')
@admin_required
def recipients_template():
    """Excelテンプレートダウンロード"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from io import BytesIO
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'メール受信者'
    headers = ['名前', 'メールアドレス', '送信区分', '仕入先CD', '有効']
    notes   = ['必須', '必須', 'both/order/expiry', 'カンマ区切りで複数可（例:244,360）', '1=有効 0=無効']
    header_fill = PatternFill('solid', fgColor='1d4ed8')
    note_fill   = PatternFill('solid', fgColor='eff6ff')
    for col, (h, n) in enumerate(zip(headers, notes), 1):
        hc = ws.cell(row=1, column=col, value=h)
        hc.font = Font(bold=True, color='FFFFFF')
        hc.fill = header_fill
        hc.alignment = Alignment(horizontal='center')
        nc = ws.cell(row=2, column=col, value=n)
        nc.fill = note_fill
        nc.font = Font(italic=True, color='6b7280')
    # サンプル行
    ws.append(['山田商事', 'yamada@example.com', 'both', '244,360', '1'])
    ws.append(['鈴木物産', 'suzuki@example.com', 'order', '584', '1'])
    col_widths = [20, 30, 14, 30, 8]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    from flask import send_file
    _db = get_db()
    _tn = _db.execute("SELECT value FROM settings WHERE key='recipients_template_name'").fetchone()
    _dl_name = ((_tn['value'] if _tn else None) or '受信者一覧_テンプレート') + '.xlsx'
    return send_file(buf, as_attachment=True,
                     download_name=_dl_name,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/recipients/import', methods=['POST'])
@admin_required
def recipients_import():
    """Excel一括インポート"""
    import openpyxl
    f = request.files.get('excel_file')
    if not f or not f.filename.endswith(('.xlsx', '.xls')):
        flash('Excelファイル（.xlsx）を選択してください。', 'danger')
        return redirect(url_for('recipients'))
    db = get_db()
    try:
        wb = openpyxl.load_workbook(f, data_only=True)
        ws = wb.active
        added = updated = skipped = 0
        for row in ws.iter_rows(min_row=3, values_only=True):  # 1行目=ヘッダ, 2行目=説明
            name        = str(row[0] or '').strip() if row[0] is not None else ''
            email       = str(row[1] or '').strip() if row[1] is not None else ''
            send_type   = str(row[2] or 'both').strip() if row[2] is not None else 'both'
            supplier_cd = str(row[3] or '').strip() if row[3] is not None else ''
            is_active   = int(row[4]) if row[4] is not None else 1
            if not name or not email:
                skipped += 1
                continue
            if send_type not in ('both', 'order', 'expiry'):
                send_type = 'both'
            existing = db.execute("SELECT id FROM mail_recipients WHERE email=%s", [email]).fetchone()
            if existing:
                db.execute(
                    "UPDATE mail_recipients SET name=%s, send_type=%s, supplier_cd=%s, is_active=%s WHERE email=%s",
                    [name, send_type, supplier_cd, is_active, email])
                updated += 1
            else:
                db.execute(
                    "INSERT INTO mail_recipients (name,email,send_type,supplier_cd,is_active) VALUES (%s,%s,%s,%s,%s)",
                    [name, email, send_type, supplier_cd, is_active])
                added += 1
        db.commit()
        flash(f'インポート完了：追加 {added}件 / 更新 {updated}件 / スキップ {skipped}件', 'success')
    except Exception as e:
        db.rollback()
        flash(f'インポートエラー: {e}', 'danger')
    return redirect(url_for('recipients'))


# ─── 棚卸（月末棚卸リスト・実棚修正）────────────────────────────
@app.route('/stocktake')
@permission_required('stocktake')
def stocktake():
    db = get_db()
    today = date.today()
    # 月末日を計算
    import calendar
    last_day = calendar.monthrange(today.year, today.month)[1]
    month_end = f"{today.year}-{today.month:02d}-{last_day:02d}"

    count_date = request.args.get('date', str(today))
    q = request.args.get('q', '').strip()
    rows = db.execute("""
        SELECT ic.*, p.product_cd, p.supplier_cd, p.supplier_name, p.unit_qty, p.cost_price
        FROM inventory_count ic
        LEFT JOIN products p ON ic.jan = p.jan
        WHERE ic.count_date=%s
        ORDER BY CAST(NULLIF(regexp_replace(p.supplier_cd,'[^0-9]','','g'),'') AS BIGINT) NULLS LAST, CAST(NULLIF(regexp_replace(p.product_cd,'[^0-9]','','g'),'') AS BIGINT) NULLS LAST
    """, [count_date]).fetchall()
    if q:
        rows = [r for r in rows if q.lower() in (r['jan'] or '').lower()
                or q.lower() in (r['product_cd'] or '').lower()
                or q.lower() in (r['product_name'] or '').lower()
                or q.lower() in (r['supplier_cd'] or '').lower()
                or q.lower() in (r['supplier_name'] or '').lower()]

    # 賞味期限別在庫内訳（JAN→リスト）
    expiry_detail = {}
    if rows:
        jans = [r['jan'] for r in rows]
        placeholders = ','.join(['%s'] * len(jans))
        stocks = db.execute(f"""
            SELECT id, jan, expiry_date, SUM(quantity) as qty
            FROM stocks
            WHERE jan IN ({placeholders}) AND quantity > 0
            GROUP BY id, jan, expiry_date
            ORDER BY jan, CASE WHEN expiry_date='' THEN '9999-99-99' ELSE expiry_date END ASC
        """, jans).fetchall()
        for s in stocks:
            if s['jan'] not in expiry_detail:
                expiry_detail[s['jan']] = []
            expiry_detail[s['jan']].append({
                'stock_id': s['id'],
                'expiry_date': s['expiry_date'] or '期限なし',
                'qty': s['qty'],
                'actual': None
            })

    # 棚卸日一覧
    dates = db.execute(
        "SELECT DISTINCT count_date FROM inventory_count ORDER BY count_date DESC LIMIT 12"
    ).fetchall()

    return render_template('stocktake.html', rows=rows, count_date=count_date,
                           month_end=month_end, dates=dates, today=today,
                           expiry_detail=expiry_detail, q=q)

@app.route('/stocktake/create', methods=['POST'])
def stocktake_create():
    count_date = request.form.get('count_date') or str(date.today())
    n = create_inventory_count(count_date)
    flash(f'棚卸リストを作成しました（{count_date}、{n}件）。', 'success')
    return redirect(url_for('stocktake', date=count_date))

@app.route('/stocktake/save', methods=['POST'])
def stocktake_save():
    db = get_db()
    count_date = request.form.get('count_date')
    jans = request.form.getlist('jan')
    for jan in jans:
        note = request.form.get(f'note_{jan}', '')
        diff_reason_category = request.form.get(f'diff_reason_category_{jan}', '')
        diff_reason_detail = request.form.get(f'diff_reason_detail_{jan}', '')
        # 賞味期限別実棚数を集計
        stocks = db.execute(
            "SELECT * FROM stocks WHERE jan=%s AND quantity>0 ORDER BY CASE WHEN expiry_date=\'\' THEN \'9999-99-99\' ELSE expiry_date END ASC",
            [jan]).fetchall()
        total_actual = 0
        has_input = False
        expiry_actuals = {}
        for s in stocks:
            key = f'actual_{jan}_{s["id"]}'
            val = request.form.get(key, '').strip()
            if val != '':
                expiry_actuals[s['id']] = int(val)
                total_actual += int(val)
                has_input = True
        # 賞味期限なしの合計入力もチェック
        total_val = request.form.get(f'actual_{jan}', '').strip()
        if total_val != '' and not has_input:
            total_actual = int(total_val)
            has_input = True
        if not has_input:
            continue
        row = db.execute(
            "SELECT * FROM inventory_count WHERE count_date=%s AND jan=%s",
            [count_date, jan]).fetchone()
        if row:
            diff = total_actual - row['system_qty']
            db.execute("""
                UPDATE inventory_count
                SET actual_qty=%s, diff_qty=%s, note=%s, adjusted=1,
                    expiry_detail=%s, diff_reason_category=%s, diff_reason_detail=%s
                WHERE count_date=%s AND jan=%s
            """, [total_actual, diff, note,
                  str(expiry_actuals) if expiry_actuals else None,
                  diff_reason_category, diff_reason_detail,
                  count_date, jan])
    db.commit()
    flash('実棚数を保存しました。', 'success')
    return redirect(url_for('stocktake', date=count_date))

@app.route('/stocktake/apply', methods=['POST'])
def stocktake_apply():
    """棚卸差異を在庫に反映（賞味期限別実棚数でシステム在庫を上書き）"""
    import ast
    db = get_db()
    count_date = request.form.get('count_date')
    rows = db.execute("""
        SELECT ic.*, p.product_cd, p.supplier_cd, p.supplier_name, p.unit_qty
        FROM inventory_count ic
        LEFT JOIN products p ON ic.jan=p.jan
        WHERE ic.count_date=%s AND ic.adjusted=1 AND ic.diff_qty!=0
    """, [count_date]).fetchall()
    applied = 0
    for row in rows:
        stocks = db.execute(
            "SELECT * FROM stocks WHERE jan=%s AND quantity>0 ORDER BY expiry_date ASC",
            [row['jan']]).fetchall()
        current = sum(s['quantity'] for s in stocks)

        # 賞味期限別実棚数がある場合は個別反映
        expiry_detail = None
        if row.get('expiry_detail'):
            try:
                expiry_detail = ast.literal_eval(row['expiry_detail'])
            except Exception:
                expiry_detail = None

        if expiry_detail:
            # 賞味期限別に個別反映
            for stock_id_str, actual_qty in expiry_detail.items():
                stock_id = int(stock_id_str)
                s = db.execute("SELECT * FROM stocks WHERE id=%s", [stock_id]).fetchone()
                if not s:
                    continue
                diff_each = actual_qty - s['quantity']
                if diff_each == 0:
                    continue
                if actual_qty <= 0:
                    db.execute("UPDATE stocks SET quantity=0 WHERE id=%s", [stock_id])
                else:
                    db.execute("UPDATE stocks SET quantity=%s WHERE id=%s", [actual_qty, stock_id])
            # 0以下の在庫レコードを削除
            db.execute("DELETE FROM stocks WHERE jan=%s AND quantity<=0", [row['jan']])
        else:
            # 合計のみの場合はFIFO調整
            diff = row['actual_qty'] - current
            if diff == 0:
                continue
            if diff > 0:
                s = stocks[0] if stocks else None
                if s:
                    db.execute("UPDATE stocks SET quantity=quantity+%s WHERE id=%s", [diff, s['id']])
                else:
                    product = db.execute("SELECT * FROM products WHERE jan=%s",[row['jan']]).fetchone()
                    if product:
                        db.execute("""
                            INSERT INTO stocks (product_id,jan,product_name,supplier_cd,
                            supplier_name,product_cd,unit_qty,quantity)
                            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)
                        """, [product['id'],row['jan'],row['product_name'],
                              product['supplier_cd'],product['supplier_name'],
                              product['product_cd'],product['unit_qty'],diff,product.get('location_code','')])
            else:
                remaining = abs(diff)
                for s in stocks:
                    if remaining <= 0: break
                    d = min(s['quantity'], remaining)
                    db.execute("UPDATE stocks SET quantity=quantity-%s WHERE id=%s", [d, s['id']])
                    remaining -= d

        db.execute("""
            INSERT INTO stock_movements
            (jan,product_name,move_type,quantity,before_qty,after_qty,note,source_file,move_date)
            VALUES (%s,%s,'adjust',%s,%s,%s,%s,%s,%s)
        """, [row['jan'],row['product_name'],abs(diff),current,row['actual_qty'],
              f"棚卸調整({count_date}) [{row.get('diff_reason_category') or '未分類'}]: {row['note'] or row.get('diff_reason_detail') or '差異修正'}",
              'stocktake', count_date])
        applied += 1

    db.commit()
    invalidate_forecast_cache()   # 在庫変更でカバー日数等が変わるためキャッシュ破棄
    flash(f'棚卸差異を在庫に反映しました（{applied}件）。', 'success')
    return redirect(url_for('stocktake', date=count_date))

@app.route('/stocktake/export_csv')
def stocktake_export_csv():
    """棚卸リストをCSVダウンロード"""
    db = get_db()
    count_date = request.args.get('date', str(date.today()))
    rows = db.execute("""SELECT ic.*, p.product_cd, p.supplier_cd, p.supplier_name, p.unit_qty, p.cost_price
        FROM inventory_count ic LEFT JOIN products p ON ic.jan=p.jan
        WHERE ic.count_date=%s ORDER BY CAST(NULLIF(regexp_replace(p.supplier_cd,'[^0-9]','','g'),'') AS BIGINT) NULLS LAST, CAST(NULLIF(regexp_replace(p.product_cd,'[^0-9]','','g'),'') AS BIGINT) NULLS LAST""",
        [count_date]).fetchall()

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(['棚卸日','仕入先CD','仕入先名','商品CD','JAN','商品名','入数','理論在庫','実棚数','差異','原価(円)','差異金額(円)','調整済','備考'])
    for r in rows:
        cost = float(r.get('cost_price') or 0)
        diff = r['diff_qty'] or 0
        loss = round(diff * cost, 0) if r['adjusted'] and diff else ''
        writer.writerow([r['count_date'],
                         r.get('supplier_cd',''),r.get('supplier_name',''),
                         r.get('product_cd',''),r['jan'],r['product_name'],
                         r.get('unit_qty',''),r['system_qty'],r['actual_qty'],diff,
                         cost if cost else '',loss,
                         '済' if r['adjusted'] else '',r['note']])
    output.seek(0)
    bom = '\ufeff'
    csv_data = bom + output.getvalue()
    filename = f"stocktake_{count_date}.csv"
    return Response(
        csv_data.encode('utf-8'),
        mimetype='text/csv; charset=utf-8',
        headers={'Content-Disposition': f'attachment; filename={filename}'}
    )

@app.route('/stocktake/export_excel')
def stocktake_export_excel():
    """棚卸リストをExcelダウンロード"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        flash('openpyxl が必要です。pip install openpyxl でインストールしてください。', 'danger')
        return redirect(url_for('stocktake'))

    db = get_db()
    count_date = request.args.get('date', str(date.today()))
    rows = db.execute(
        """SELECT ic.*, p.product_cd, p.supplier_cd, p.supplier_name, p.unit_qty, p.cost_price
        FROM inventory_count ic LEFT JOIN products p ON ic.jan=p.jan
        WHERE ic.count_date=%s ORDER BY CAST(NULLIF(regexp_replace(p.supplier_cd,'[^0-9]','','g'),'') AS BIGINT) NULLS LAST, CAST(NULLIF(regexp_replace(p.product_cd,'[^0-9]','','g'),'') AS BIGINT) NULLS LAST""",
        [count_date]).fetchall()

    # 賞味期限別在庫内訳
    expiry_detail_xl = {}
    if rows:
        jans = [r['jan'] for r in rows]
        placeholders = ','.join(['%s'] * len(jans))
        stocks = db.execute(f"""
            SELECT jan, expiry_date, SUM(quantity) as qty FROM stocks
            WHERE jan IN ({placeholders}) AND quantity > 0
            GROUP BY jan, expiry_date
            ORDER BY jan, CASE WHEN expiry_date='' THEN '9999-99-99' ELSE expiry_date END ASC
        """, jans).fetchall()
        for s in stocks:
            if s['jan'] not in expiry_detail_xl:
                expiry_detail_xl[s['jan']] = []
            expiry_detail_xl[s['jan']].append(
                f"{s['expiry_date'] or '期限なし'}:{s['qty']}個"
            )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"棚卸_{count_date}"

    # ── タイトル行 ──────────────────────────────
    ws.merge_cells('A1:O1')
    ws['A1'] = f"棚 卸 リ ス ト　　{count_date}"
    ws['A1'].font = Font(bold=True, size=16)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 32

    # ── ヘッダー行 ──────────────────────────────
    headers = ['No.', '仕入先CD', '仕入先名', '商品CD', 'JAN', '商品名', '入数', '理論在庫', '賞味期限別内訳', '実棚数', '差異', '原価(円)', '差異金額(円)', '調整済', '備考']
    hdr_fill = PatternFill(fgColor='1A2744', fill_type='solid')
    thin = Side(style='thin', color='CCCCCC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=col, value=h)
        c.font = Font(bold=True, color='FFFFFF', size=11)
        c.fill = hdr_fill
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = border
    ws.row_dimensions[2].height = 22

    # ── データ行 ────────────────────────────────
    red_fill    = PatternFill(fgColor='FEE2E2', fill_type='solid')
    green_fill  = PatternFill(fgColor='DCFCE7', fill_type='solid')
    normal_fill = PatternFill(fgColor='FFFFFF', fill_type='solid')

    for i, r in enumerate(rows, 1):
        row_num = i + 2
        diff = r['diff_qty'] or 0
        row_fill = red_fill if diff < 0 else (green_fill if diff > 0 else normal_fill)

        cost = float(r['cost_price'] or 0)
        loss = round(diff * cost, 0) if r['adjusted'] and diff else ''
        expiry_str = ' / '.join(expiry_detail_xl.get(r['jan'], [])) or '-'
        values = [i,
                  r['supplier_cd'] or '', r['supplier_name'] or '',
                  r['product_cd'] or '', r['jan'], r['product_name'],
                  r['unit_qty'] or '',
                  r['system_qty'],
                  expiry_str,
                  r['actual_qty'] if r['adjusted'] else '',
                  diff if r['adjusted'] else '',
                  cost if cost else '',
                  loss,
                  '済' if r['adjusted'] else '',
                  r['note'] or '']
        align_right = {1, 7, 9, 10, 11, 12, 13}
        for col, val in enumerate(values, 1):
            c = ws.cell(row=row_num, column=col, value=val)
            c.fill = row_fill
            c.border = border
            c.font = Font(size=11)
            c.alignment = Alignment(
                horizontal='right' if col in align_right else 'left',
                vertical='center',
                wrap_text=(col == 9)
            )
        ws.row_dimensions[row_num].height = 20

    # ── 列幅 ──────────────────────────────────
    col_widths = [5, 10, 18, 12, 16, 28, 7, 10, 30, 10, 8, 10, 12, 8, 22]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=2, column=col).column_letter].width = width

    # ── フッター（担当者サイン欄）─────────────
    footer_row = len(rows) + 4
    ws.cell(row=footer_row, column=1, value="担当者：").font = Font(size=11)
    ws.cell(row=footer_row, column=4, value="確認者：").font = Font(size=11)
    ws.cell(row=footer_row, column=7, value="承認者：").font = Font(size=11)
    ws.row_dimensions[footer_row].height = 30

    # ── ページ設定（印刷用）──────────────────
    ws.page_setup.orientation = 'portrait'
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.print_title_rows = '1:2'

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    _name_row = db.execute("SELECT value FROM settings WHERE key='stocktake_export_name'").fetchone()
    _base_name = _name_row['value'] if _name_row else '棚卸リスト'
    filename = f"{_base_name}_{count_date}.xlsx"
    from urllib.parse import quote
    return Response(
        output.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f"attachment; filename*=UTF-8''{quote(filename)}"}
    )


# ─── レポート ────────────────────────────────────────────────────
@app.route('/reports')
@permission_required('reports')
def reports():
    db = get_db()
    today = date.today()
    q = request.args.get('q', '').strip()

    def match(r):
        if not q:
            return True
        ql = q.lower()
        return (ql in (r.get('jan') or '').lower()
                or ql in (r.get('product_cd') or '').lower()
                or ql in (r.get('product_name') or '').lower()
                or ql in (r.get('supplier_cd') or '').lower()
                or ql in (r.get('supplier_name') or '').lower())

    # ① 賞味期限アラート
    expiry_soon = db.execute("""
        SELECT s.*, p.expiry_alert_days, p.product_cd, p.supplier_cd, p.supplier_name,
               (s.expiry_date::date - CURRENT_DATE) as days_left
        FROM stocks s JOIN products p ON s.jan=p.jan
        WHERE s.quantity>0 AND s.expiry_date!=''
        AND s.expiry_date::date <= CURRENT_DATE + (p.expiry_alert_days || ' days')::INTERVAL
        ORDER BY CAST(NULLIF(regexp_replace(p.supplier_cd,'[^0-9]','','g'),'') AS BIGINT) NULLS LAST, CAST(NULLIF(regexp_replace(p.product_cd,'[^0-9]','','g'),'') AS BIGINT) NULLS LAST, s.expiry_date ASC
    """).fetchall()
    if q: expiry_soon = [r for r in expiry_soon if match(r)]

    # ② 在庫一覧（全商品・在庫あり）
    stock_summary = db.execute("""
        SELECT p.supplier_cd, p.supplier_name, p.product_cd, p.jan, p.product_name,
               p.unit_qty, p.reorder_point,
               COALESCE(p.cost_price, 0) as cost_price,
               COALESCE(SUM(s.quantity),0) as total_qty,
               MIN(CASE WHEN s.expiry_date IS NOT NULL AND s.expiry_date != '' THEN s.expiry_date END) as earliest_expiry,
               MAX(CASE WHEN s.expiry_date IS NOT NULL AND s.expiry_date != '' THEN s.expiry_date END) as latest_expiry
        FROM products p
        LEFT JOIN stocks s ON p.jan=s.jan AND s.quantity>0
        WHERE p.is_active=1
        GROUP BY p.id, p.supplier_cd, p.supplier_name, p.product_cd, p.jan,
                 p.product_name, p.unit_qty, p.reorder_point, p.cost_price
        HAVING COALESCE(SUM(s.quantity),0) > 0
        ORDER BY CAST(NULLIF(regexp_replace(p.supplier_cd,'[^0-9]','','g'),'') AS BIGINT) NULLS LAST, CAST(NULLIF(regexp_replace(p.product_cd,'[^0-9]','','g'),'') AS BIGINT) NULLS LAST
    """).fetchall()
    if q: stock_summary = [r for r in stock_summary if match(r)]

    # ③ 在庫なし商品（発注点以下含む）
    zero_stock = db.execute("""
        SELECT p.supplier_cd, p.supplier_name, p.product_cd, p.jan, p.product_name,
               p.reorder_point, p.ordered_at,
               COALESCE(SUM(s.quantity),0) as total_qty
        FROM products p
        LEFT JOIN stocks s ON p.jan=s.jan AND s.quantity>0
        WHERE p.is_active=1
        GROUP BY p.id, p.supplier_cd, p.supplier_name, p.product_cd, p.jan,
                 p.product_name, p.reorder_point, p.ordered_at
        HAVING COALESCE(SUM(s.quantity),0) = 0
        ORDER BY CAST(NULLIF(regexp_replace(p.supplier_cd,'[^0-9]','','g'),'') AS BIGINT) NULLS LAST, CAST(NULLIF(regexp_replace(p.product_cd,'[^0-9]','','g'),'') AS BIGINT) NULLS LAST
    """).fetchall()
    if q: zero_stock = [r for r in zero_stock if match(r)]

    # ④ 発注点以下（在庫あり）
    low_stock = db.execute("""
        SELECT p.supplier_cd, p.supplier_name, p.product_cd, p.jan, p.product_name,
               p.reorder_point, p.order_qty, p.ordered_at,
               COALESCE(SUM(s.quantity),0) as total_qty
        FROM products p
        LEFT JOIN stocks s ON p.jan=s.jan AND s.quantity>0
        WHERE p.is_active=1
        GROUP BY p.id, p.supplier_cd, p.supplier_name, p.product_cd, p.jan,
                 p.product_name, p.reorder_point, p.order_qty, p.ordered_at
        HAVING COALESCE(SUM(s.quantity),0) > 0
           AND COALESCE(SUM(s.quantity),0) <= p.reorder_point
        ORDER BY CAST(NULLIF(regexp_replace(p.supplier_cd,'[^0-9]','','g'),'') AS BIGINT) NULLS LAST, CAST(NULLIF(regexp_replace(p.product_cd,'[^0-9]','','g'),'') AS BIGINT) NULLS LAST
    """).fetchall()
    if q: low_stock = [r for r in low_stock if match(r)]

    # ⑤ 月別発注集計
    monthly_order = db.execute("""
        SELECT to_char(order_date::date,'YYYY-MM') as month,
               COUNT(*) as cnt, SUM(order_qty) as total_qty,
               COUNT(DISTINCT jan) as product_cnt
        FROM order_history GROUP BY month ORDER BY month DESC LIMIT 12
    """).fetchall()

    # ⑥ 仕入先別在庫金額
    supplier_stock = db.execute("""
        SELECT p.supplier_cd, p.supplier_name,
               COUNT(DISTINCT p.jan) as product_cnt,
               SUM(s.quantity) as total_qty,
               SUM(s.quantity * COALESCE(p.cost_price,0)) as stock_cost
        FROM products p
        JOIN stocks s ON p.jan=s.jan
        WHERE p.is_active=1 AND s.quantity>0
        GROUP BY p.supplier_cd, p.supplier_name
        ORDER BY stock_cost DESC
    """).fetchall()

    # ⑦ 廃棄・退避ロス集計（月別）
    loss_monthly = db.execute("""
        SELECT to_char(disposed_at::date,'YYYY-MM') as month,
               COUNT(*) as cnt,
               SUM(quantity) as total_qty,
               SUM(loss_amount) as total_loss
        FROM disposed_stocks
        GROUP BY month ORDER BY month DESC LIMIT 12
    """).fetchall()

    # ⑧ 廃棄・退避ロス集計（仕入先別）
    loss_supplier = db.execute("""
        SELECT supplier_cd, supplier_name,
               COUNT(*) as cnt,
               SUM(quantity) as total_qty,
               SUM(loss_amount) as total_loss
        FROM disposed_stocks
        GROUP BY supplier_cd, supplier_name
        ORDER BY total_loss DESC LIMIT 20
    """).fetchall()

    # ⑨ 在庫総額サマリー
    stock_total = db.execute("""
        SELECT SUM(s.quantity * COALESCE(p.cost_price,0)) as total_cost,
               SUM(s.quantity * COALESCE(p.sell_price,0)) as total_sell,
               COUNT(DISTINCT p.jan) as product_cnt,
               SUM(s.quantity) as total_qty
        FROM stocks s JOIN products p ON s.jan=p.jan
        WHERE s.quantity>0
    """).fetchone()

    # ⑩ 発注設定一覧
    order_settings = db.execute("""
        SELECT p.id, p.supplier_cd, p.supplier_name, p.product_cd, p.jan, p.product_name,
               p.unit_qty, p.order_qty, p.reorder_point, p.reorder_auto,
               p.lead_time_days, p.safety_factor,
               p.mixed_group, p.mixed_lot_cases, p.mixed_force_days,
               p.expiry_alert_days,
               COALESCE(SUM(s.quantity),0) as current_stock
        FROM products p
        LEFT JOIN stocks s ON p.jan=s.jan AND s.quantity>0
        WHERE p.is_active=1
        GROUP BY p.id, p.supplier_cd, p.supplier_name, p.product_cd, p.jan,
                 p.product_name, p.unit_qty, p.order_qty, p.reorder_point,
                 p.reorder_auto, p.lead_time_days, p.safety_factor,
                 p.mixed_group, p.mixed_lot_cases, p.mixed_force_days,
                 p.expiry_alert_days
        ORDER BY CAST(NULLIF(regexp_replace(p.supplier_cd,'[^0-9]','','g'),'') AS BIGINT) NULLS LAST, CAST(NULLIF(regexp_replace(p.product_cd,'[^0-9]','','g'),'') AS BIGINT) NULLS LAST
    """).fetchall()

    # ⑪ 前年実績サマリーは AJAX で遅延ロード（/reports/sales_data）
    # → ページ初期表示時に重いクエリを実行せず、即座にレンダリング
    last_year  = today.year - 1
    sales_mode = request.args.get('sales_mode', 'master')
    sales_q    = request.args.get('sales_q', '').strip()

    return render_template('reports.html',
        expiry_soon=expiry_soon,
        stock_summary=stock_summary,
        zero_stock=zero_stock,
        low_stock=low_stock,
        monthly_order=monthly_order,
        supplier_stock=supplier_stock,
        loss_monthly=loss_monthly,
        loss_supplier=loss_supplier,
        stock_total=stock_total,
        order_settings=order_settings,
        last_year=last_year, today=today,
        sales_mode=sales_mode, q=q, sales_q=sales_q)


@app.route('/reports/sales_data')
@permission_required('reports')
def reports_sales_data():
    """前年実績サマリーを JSON で返す AJAX エンドポイント。
    sale_date は TEXT 型 'YYYY-MM-DD' → BETWEEN で文字列比較しインデックスを活用。"""
    db = get_db()
    today      = date.today()
    last_year  = today.year - 1
    sales_mode = request.args.get('mode', 'master')
    sales_q    = request.args.get('q', '').strip()

    yr_start = f'{last_year}-01-01'
    yr_end   = f'{last_year}-12-31'

    if sales_mode == 'all':
        rows = db.execute("""
            SELECT sh.jan,
                   COALESCE(p.product_name, sh.product_name, sh.jan) AS product_name,
                   COALESCE(p.supplier_cd,   '') AS supplier_cd,
                   COALESCE(p.supplier_name, '') AS supplier_name,
                   COALESCE(p.product_cd,    '') AS product_cd,
                   SUBSTRING(sh.sale_date, 6, 2) AS month,
                   SUM(sh.quantity) AS total
            FROM sales_history sh
            LEFT JOIN products p ON sh.jan = p.jan AND p.is_active = 1
            WHERE sh.sale_date BETWEEN %s AND %s
            GROUP BY sh.jan, p.product_name, sh.product_name,
                     p.supplier_cd, p.supplier_name, p.product_cd, month
            ORDER BY supplier_cd, product_cd, sh.jan, month
        """, [yr_start, yr_end]).fetchall()
    else:
        rows = db.execute("""
            SELECT sh.jan,
                   p.product_name,
                   p.supplier_cd,
                   p.supplier_name,
                   p.product_cd,
                   SUBSTRING(sh.sale_date, 6, 2) AS month,
                   SUM(sh.quantity) AS total
            FROM sales_history sh
            INNER JOIN products p ON sh.jan = p.jan AND p.is_active = 1
            WHERE sh.sale_date BETWEEN %s AND %s
            GROUP BY sh.jan, p.product_name, p.supplier_cd, p.supplier_name, p.product_cd, month
            ORDER BY p.supplier_cd, p.product_cd, sh.jan, month
        """, [yr_start, yr_end]).fetchall()

    if sales_q:
        ql = sales_q.lower()
        rows = [r for r in rows if
                ql in (r['jan']          or '').lower()
                or ql in (r['product_cd']   or '').lower()
                or ql in (r['product_name'] or '').lower()
                or ql in (r['supplier_cd']  or '').lower()
                or ql in (r['supplier_name'] or '').lower()]

    # JAN ごとにグループ化（順序維持）
    from collections import OrderedDict
    grouped = OrderedDict()
    for r in rows:
        jan = r['jan']
        if jan not in grouped:
            grouped[jan] = {
                'jan': jan,
                'product_name': r['product_name'],
                'supplier_cd':  r['supplier_cd'],
                'supplier_name': r['supplier_name'],
                'product_cd':   r['product_cd'],
                'months': {}
            }
        grouped[jan]['months'][r['month']] = int(r['total'] or 0)

    result = []
    for d in grouped.values():
        d['total'] = sum(d['months'].values())
        result.append(d)

    return jsonify({'rows': result, 'year': last_year, 'count': len(result)})


@app.route('/expiry_check', methods=['POST'])
@login_required
def expiry_check():
    items = run_expiry_check()
    if items:
        flash(f'賞味期限アラート: {len(items)}件検出・メール送信。', 'warning')
    else:
        flash('期限アラートなし。', 'success')
    return redirect(url_for('reports'))

# ─── ユーザー管理（管理者専用）────────────────────────────────────
@app.route('/users')
@permission_required('users')
def users():
    db = get_db()
    rows = db.execute("SELECT id,username,role,is_active,permissions,created_at FROM users ORDER BY id").fetchall()
    return render_template('users.html', users=rows)

@app.route('/users/new', methods=['POST'])
@admin_required
def user_new():
    f = request.form
    username = f.get('username','').strip()
    password = f.get('password','').strip()
    role     = f.get('role','user')
    # 権限チェックボックス（一般ユーザーのみ）
    if role == 'admin':
        permissions = ','.join([p[0] for p in PAGE_PERMISSIONS])
    else:
        permissions = ','.join([p[0] for p in PAGE_PERMISSIONS if f.get(f'perm_{p[0]}') == '1'])
    if not username or not password:
        flash('IDとパスワードを入力してください。', 'danger')
        return redirect(url_for('users'))
    db = get_db()
    try:
        db.execute("INSERT INTO users (username,password,role,permissions) VALUES (%s,%s,%s,%s)",
                   [username, _hash(password), role, permissions])
        db.commit()
        flash(f'ユーザー「{username}」を追加しました。', 'success')
    except Exception as e:
        db.rollback()
        flash(f'追加エラー: {e}', 'danger')
    return redirect(url_for('users'))

@app.route('/users/<int:uid>/permissions', methods=['POST'])
@admin_required
def user_permissions(uid):
    db = get_db()
    u = db.execute("SELECT * FROM users WHERE id=%s", [uid]).fetchone()
    if not u:
        flash('ユーザーが見つかりません', 'error')
        return redirect(url_for('users'))
    if u['role'] == 'admin':
        permissions = ','.join([p[0] for p in PAGE_PERMISSIONS])
    else:
        permissions = ','.join([p[0] for p in PAGE_PERMISSIONS if request.form.get(f'perm_{p[0]}') == '1'])
    db.execute("UPDATE users SET permissions=%s WHERE id=%s", [permissions, uid])
    db.commit()
    flash(f'「{u["username"]}」の権限を更新しました', 'success')
    return redirect(url_for('users'))

@app.route('/users/<int:uid>/toggle', methods=['POST'])
@admin_required
def user_toggle(uid):
    db = get_db()
    u = db.execute("SELECT * FROM users WHERE id=%s", [uid]).fetchone()
    if u:
        if u['username'] == session.get('user'):
            flash('自分自身を無効化することはできません。', 'danger')
        else:
            db.execute("UPDATE users SET is_active=%s WHERE id=%s",
                       [0 if u['is_active'] else 1, uid])
            db.commit()
    return redirect(url_for('users'))

@app.route('/users/<int:uid>/delete', methods=['POST'])
@admin_required
def user_delete(uid):
    db = get_db()
    u = db.execute("SELECT * FROM users WHERE id=%s", [uid]).fetchone()
    if u and u['username'] == session.get('user'):
        flash('自分自身は削除できません。', 'danger')
    elif u:
        db.execute("DELETE FROM users WHERE id=%s", [uid])
        db.commit()
        flash('削除しました。', 'success')
    return redirect(url_for('users'))

@app.route('/users/<int:uid>/password', methods=['POST'])
@admin_required
def user_password(uid):
    pw = request.form.get('password','').strip()
    if not pw:
        flash('パスワードを入力してください。', 'danger')
        return redirect(url_for('users'))
    db = get_db()
    db.execute("UPDATE users SET password=%s WHERE id=%s", [_hash(pw), uid])
    db.commit()
    flash('パスワードを変更しました。', 'success')
    return redirect(url_for('users'))

# ─── 設定 ────────────────────────────────────────────────────────
@app.route('/settings/mail_templates', methods=['GET','POST'])
@admin_required
def mail_templates():
    db = get_db()
    defaults = {
        'order': {
            'subject':     '【発注一覧】{date} ({count}件)',
            'body_header': '発注日: {date}\n件数: {count}件\n',
            'body_item':   '{supplier_cd}  {supplier_name}  {jan}  {product_cd}  {product_name}  {order_qty}個  {trigger}',
            'body_footer': '--\n{from_name}',
        },
        'expiry': {
            'subject':     '【賞味期限アラート】期限切れ間近の在庫があります - {date}',
            'body_header': '以下の商品の賞味期限が近づいています。早めの対応をお願いします。\n\n確認日: {date}\n',
            'body_item':   '  ・{product_name}　LOT:{lot_no}　残り{days_left}日 (期限: {expiry_date})　在庫{quantity}個',
            'body_footer': '',
        },
    }
    if request.method == 'POST':
        for mail_type in ['order', 'expiry']:
            subject     = request.form.get(f'{mail_type}_subject', '').strip()
            body_header = request.form.get(f'{mail_type}_body_header', '')
            body_item   = request.form.get(f'{mail_type}_body_item', '').strip()
            body_footer = request.form.get(f'{mail_type}_body_footer', '')
            if not subject or not body_item:
                continue
            existing = db.execute("SELECT id FROM mail_templates WHERE mail_type=%s", [mail_type]).fetchone()
            if existing:
                db.execute("""UPDATE mail_templates SET subject=%s,body_header=%s,body_item=%s,body_footer=%s WHERE mail_type=%s""",
                           [subject, body_header, body_item, body_footer, mail_type])
            else:
                db.execute("""INSERT INTO mail_templates (mail_type,subject,body_header,body_item,body_footer) VALUES (%s,%s,%s,%s,%s)""",
                           [mail_type, subject, body_header, body_item, body_footer])
        db.commit()
        flash('メールテンプレートを保存しました', 'success')
        return redirect(url_for('mail_templates'))
    templates = {}
    for mail_type in ['order', 'expiry']:
        row = db.execute("SELECT * FROM mail_templates WHERE mail_type=%s", [mail_type]).fetchone()
        templates[mail_type] = dict(row) if row else defaults[mail_type]
    return render_template('mail_templates.html', templates=templates, defaults=defaults)


@app.route('/settings/test_order_mail', methods=['POST'])
@admin_required
def test_order_mail():
    from mail_service import queue_order, flush_order_mail
    today = str(date.today())
    # サンプル発注データ
    sample_product = {
        'supplier_cd': 'SUP001', 'supplier_name': '山田食品',
        'supplier_email': '', 'jan': '4901234567890',
        'product_cd': 'P001', 'product_name': 'りんごジュース 1L（サンプル）'
    }
    queue_order(sample_product, 24, 'reorder')
    queue_order({**sample_product, 'product_name': 'オレンジジュース 1L（サンプル）', 'jan': '4901234567891'}, 12, 'lot')
    ok, msg = flush_order_mail()
    if ok:
        flash(f'発注サンプルメールを送信しました: {msg}', 'success')
    else:
        flash(f'送信失敗: {msg}', 'danger')
    return redirect(url_for('settings'))

@app.route('/settings/test_expiry_mail', methods=['POST'])
@admin_required
def test_expiry_mail():
    from mail_service import send_expiry_alert
    db = get_db()
    sample_alerts = [
        {'product_name': 'りんごジュース 1L（サンプル）', 'lot_no': 'LOT001',
         'days_left': 3, 'expiry_date': str(date.today()), 'quantity': 24},
        {'product_name': 'オレンジジュース 1L（サンプル）', 'lot_no': 'LOT002',
         'days_left': 7, 'expiry_date': str(date.today()), 'quantity': 12},
    ]
    ok, msg = send_expiry_alert(db, sample_alerts)
    if ok:
        flash(f'賞味期限アラートサンプルメールを送信しました: {msg}', 'success')
    else:
        flash(f'送信失敗: {msg}', 'danger')
    return redirect(url_for('settings'))

@app.route('/clear_import_errors', methods=['POST'])
@login_required
def clear_import_errors():
    session.pop('import_errors', None)
    return '', 204

@app.route('/admin/backup')
@admin_required
def backup():
    """全データをExcelにエクスポートしてダウンロード"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from urllib.parse import quote
    import io

    db = get_db()
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # デフォルトシート削除

    hfill = PatternFill('solid', fgColor='1E3A8A')
    hfont = Font(bold=True, color='FFFFFF', name='Meiryo UI', size=10)
    bdr = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    dfont = Font(name='Meiryo UI', size=10)

    def write_sheet(ws, rows, headers):
        for ci, h in enumerate(headers, 1):
            c = ws.cell(1, ci, h)
            c.font = hfont; c.fill = hfill
            c.alignment = Alignment(horizontal='center', vertical='center')
            c.border = bdr
        if rows:
            for ri, row in enumerate(rows, 2):
                for ci, key in enumerate(headers, 1):
                    val = row[ci-1] if isinstance(row, (list, tuple)) else row.get(headers[ci-1]) if hasattr(row, 'get') else None
                    try:
                        val = row[key]
                    except Exception:
                        val = ''
                    c = ws.cell(ri, ci, val)
                    c.font = dfont; c.border = bdr
        ws.freeze_panes = 'A2'

    # ── 商品マスタ ──
    ws = wb.create_sheet('商品マスタ')
    rows = db.execute("SELECT * FROM products ORDER BY CAST(NULLIF(regexp_replace(supplier_cd,'[^0-9]','','g'),'') AS BIGINT) NULLS LAST, CAST(NULLIF(regexp_replace(product_cd,'[^0-9]','','g'),'') AS BIGINT) NULLS LAST").fetchall()
    if rows:
        keys = list(rows[0].keys())
        for ci, h in enumerate(keys, 1):
            c = ws.cell(1, ci, h); c.font = hfont; c.fill = hfill
            c.alignment = Alignment(horizontal='center'); c.border = bdr
        for ri, row in enumerate(rows, 2):
            for ci, key in enumerate(keys, 1):
                c = ws.cell(ri, ci, row[key]); c.font = dfont; c.border = bdr
    ws.freeze_panes = 'A2'

    # ── 在庫一覧 ──
    ws = wb.create_sheet('在庫一覧')
    rows = db.execute("""
        SELECT s.*, p.product_cd, p.supplier_cd, p.supplier_name
        FROM stocks s LEFT JOIN products p ON s.jan=p.jan
        WHERE s.quantity>0 ORDER BY CAST(NULLIF(regexp_replace(p.supplier_cd,'[^0-9]','','g'),'') AS BIGINT) NULLS LAST, CAST(NULLIF(regexp_replace(p.product_cd,'[^0-9]','','g'),'') AS BIGINT) NULLS LAST
    """).fetchall()
    if rows:
        keys = list(rows[0].keys())
        for ci, h in enumerate(keys, 1):
            c = ws.cell(1, ci, h); c.font = hfont; c.fill = hfill
            c.alignment = Alignment(horizontal='center'); c.border = bdr
        for ri, row in enumerate(rows, 2):
            for ci, key in enumerate(keys, 1):
                c = ws.cell(ri, ci, row[key]); c.font = dfont; c.border = bdr
    ws.freeze_panes = 'A2'

    # ── 発注履歴 ──
    ws = wb.create_sheet('発注履歴')
    rows = db.execute("SELECT * FROM order_history ORDER BY created_at DESC").fetchall()
    if rows:
        keys = list(rows[0].keys())
        for ci, h in enumerate(keys, 1):
            c = ws.cell(1, ci, h); c.font = hfont; c.fill = hfill
            c.alignment = Alignment(horizontal='center'); c.border = bdr
        for ri, row in enumerate(rows, 2):
            for ci, key in enumerate(keys, 1):
                c = ws.cell(ri, ci, row[key]); c.font = dfont; c.border = bdr
    ws.freeze_panes = 'A2'

    # ── 廃棄退避在庫 ──
    ws = wb.create_sheet('廃棄退避在庫')
    rows = db.execute("SELECT * FROM disposed_stocks ORDER BY disposed_at DESC").fetchall()
    if rows:
        keys = list(rows[0].keys())
        for ci, h in enumerate(keys, 1):
            c = ws.cell(1, ci, h); c.font = hfont; c.fill = hfill
            c.alignment = Alignment(horizontal='center'); c.border = bdr
        for ri, row in enumerate(rows, 2):
            for ci, key in enumerate(keys, 1):
                c = ws.cell(ri, ci, row[key]); c.font = dfont; c.border = bdr
    ws.freeze_panes = 'A2'

    # ── 入庫履歴 ──
    ws = wb.create_sheet('入庫履歴')
    rows = db.execute("SELECT * FROM stock_movements ORDER BY created_at DESC LIMIT 10000").fetchall()
    if rows:
        keys = list(rows[0].keys())
        for ci, h in enumerate(keys, 1):
            c = ws.cell(1, ci, h); c.font = hfont; c.fill = hfill
            c.alignment = Alignment(horizontal='center'); c.border = bdr
        for ri, row in enumerate(rows, 2):
            for ci, key in enumerate(keys, 1):
                c = ws.cell(ri, ci, row[key]); c.font = dfont; c.border = bdr
    ws.freeze_panes = 'A2'

    # ── 棚卸データ ──
    ws = wb.create_sheet('棚卸データ')
    rows = db.execute("SELECT * FROM inventory_count ORDER BY count_date DESC, jan").fetchall()
    if rows:
        keys = list(rows[0].keys())
        for ci, h in enumerate(keys, 1):
            c = ws.cell(1, ci, h); c.font = hfont; c.fill = hfill
            c.alignment = Alignment(horizontal='center'); c.border = bdr
        for ri, row in enumerate(rows, 2):
            for ci, key in enumerate(keys, 1):
                c = ws.cell(ri, ci, row[key]); c.font = dfont; c.border = bdr
    ws.freeze_panes = 'A2'

    # ── CSV取込ログ ──
    ws = wb.create_sheet('CSV取込ログ')
    rows = db.execute("""
        SELECT l.*, s.name as setting_name FROM import_logs l
        LEFT JOIN csv_import_settings s ON l.setting_id=s.id
        ORDER BY l.imported_at DESC
    """).fetchall()
    if rows:
        keys = list(rows[0].keys())
        for ci, h in enumerate(keys, 1):
            c = ws.cell(1, ci, h); c.font = hfont; c.fill = hfill
            c.alignment = Alignment(horizontal='center'); c.border = bdr
        for ri, row in enumerate(rows, 2):
            for ci, key in enumerate(keys, 1):
                c = ws.cell(ri, ci, row[key]); c.font = dfont; c.border = bdr
    ws.freeze_panes = 'A2'

    # ── CSVインポートデータ（sales_history）──
    ws = wb.create_sheet('CSVインポートデータ')
    rows = db.execute("""
        SELECT jan, product_name, quantity, sale_date, source_file,
               chain_cd, client_name, store_cd, store_name, row_hash, created_at
        FROM sales_history
        ORDER BY sale_date DESC, created_at DESC
    """).fetchall()
    if rows:
        keys = list(rows[0].keys())
        for ci, h in enumerate(keys, 1):
            c = ws.cell(1, ci, h); c.font = hfont; c.fill = hfill
            c.alignment = Alignment(horizontal='center'); c.border = bdr
        for ri, row in enumerate(rows, 2):
            for ci, key in enumerate(keys, 1):
                c = ws.cell(ri, ci, row[key]); c.font = dfont; c.border = bdr
    ws.freeze_panes = 'A2'

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    filename = f"在庫管理システム_バックアップ_{date.today()}.xlsx"
    return Response(
        buf.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f"attachment; filename*=UTF-8''{quote(filename)}"}
    )


@app.route('/admin/restore', methods=['GET', 'POST'])
@admin_required
def restore():
    if request.method == 'GET':
        return render_template('restore.html')

    f = request.files.get('file')
    if not f or not f.filename.endswith('.xlsx'):
        flash('Excelファイル(.xlsx)を選択してください', 'error')
        return redirect(url_for('restore'))

    targets = request.form.getlist('targets')
    if not targets:
        flash('復元対象を選択してください', 'error')
        return redirect(url_for('restore'))

    import openpyxl, io as _io
    try:
        wb = openpyxl.load_workbook(_io.BytesIO(f.read()), data_only=True)
    except Exception as e:
        flash(f'Excelファイルの読み込みエラー: {e}', 'error')
        return redirect(url_for('restore'))

    db = get_db()
    results = []
    errors = []

    def get_sheet_rows(sheet_name):
        if sheet_name not in wb.sheetnames:
            return None, None
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            return None, None
        headers = [str(h) if h is not None else '' for h in rows[0]]
        return headers, rows[1:]

    def sv(rd, key, default=''):
        v = rd.get(key)
        return str(v).strip() if v is not None else default

    def iv(rd, key, default=0):
        v = rd.get(key)
        try: return int(float(str(v))) if v is not None else default
        except: return default

    def fv(rd, key, default=0.0):
        v = rd.get(key)
        try: return float(str(v)) if v is not None else default
        except: return default

    try:
        # ── 商品マスタ ──
        if '商品マスタ' in targets:
            headers, data = get_sheet_rows('商品マスタ')
            if headers and data:
                cnt = 0
                for row in data:
                    if not any(row): continue
                    rd = dict(zip(headers, row))
                    jan = sv(rd, 'jan')
                    product_cd = sv(rd, 'product_cd')
                    if not jan and not product_cd: continue
                    # JANまたは商品CDで既存チェック
                    existing = None
                    if jan:
                        existing = db.execute("SELECT id FROM products WHERE jan=%s", [jan]).fetchone()
                    if not existing and product_cd:
                        existing = db.execute("SELECT id,jan FROM products WHERE product_cd=%s AND is_active=1", [product_cd]).fetchone()
                        if existing and not jan:
                            jan = existing['jan']
                    vals = [
                        sv(rd,'supplier_cd'), sv(rd,'supplier_name'),
                        product_cd, sv(rd,'product_name'),
                        iv(rd,'unit_qty',1), iv(rd,'order_unit',1),
                        iv(rd,'order_qty',1), iv(rd,'reorder_point',0),
                        iv(rd,'reorder_auto',1), iv(rd,'lead_time_days',3),
                        fv(rd,'safety_factor',1.3), iv(rd,'lot_size',0),
                        iv(rd,'shelf_life_days',365), iv(rd,'expiry_alert_days',30),
                        sv(rd,'mixed_group'), sv(rd,'mixed_lot_mode','gte'),
                        iv(rd,'mixed_lot_cases',3), iv(rd,'mixed_force_days',3),
                        fv(rd,'cost_price',0), fv(rd,'sell_price',0),
                        iv(rd,'is_active',1)
                    ]
                    if existing:
                        db.execute("""
                            UPDATE products SET
                            supplier_cd=%s,supplier_name=%s,product_cd=%s,product_name=%s,
                            unit_qty=%s,order_unit=%s,order_qty=%s,reorder_point=%s,
                            reorder_auto=%s,lead_time_days=%s,safety_factor=%s,lot_size=%s,
                            shelf_life_days=%s,expiry_alert_days=%s,
                            mixed_group=%s,mixed_lot_mode=%s,mixed_lot_cases=%s,mixed_force_days=%s,
                            cost_price=%s,sell_price=%s,is_active=%s
                            WHERE jan=%s
                        """, vals + [jan])
                    else:
                        if not jan: continue
                        db.execute("""
                            INSERT INTO products
                            (supplier_cd,supplier_name,product_cd,product_name,
                             unit_qty,order_unit,order_qty,reorder_point,
                             reorder_auto,lead_time_days,safety_factor,lot_size,
                             shelf_life_days,expiry_alert_days,
                             mixed_group,mixed_lot_mode,mixed_lot_cases,mixed_force_days,
                             cost_price,sell_price,is_active,jan)
                            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                        """, vals + [jan])
                    cnt += 1
                results.append(f'商品マスタ: {cnt}件復元')
            else:
                errors.append('商品マスタシートなし')

        # ── 在庫一覧 ──
        if '在庫一覧' in targets:
            headers, data = get_sheet_rows('在庫一覧')
            if headers and data:
                db.execute("DELETE FROM stocks")
                cnt = 0
                for row in data:
                    if not any(row): continue
                    rd = dict(zip(headers, row))
                    jan = sv(rd, 'jan')
                    if not jan or not rd.get('quantity'): continue
                    p = db.execute("SELECT id FROM products WHERE jan=%s", [jan]).fetchone()
                    if not p: continue
                    db.execute("""
                        INSERT INTO stocks
                        (product_id,jan,product_name,supplier_cd,supplier_name,
                         product_cd,unit_qty,quantity,expiry_date,lot_no)
                        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                    """, [p['id'],jan,sv(rd,'product_name'),sv(rd,'supplier_cd'),
                          sv(rd,'supplier_name'),sv(rd,'product_cd'),
                          iv(rd,'unit_qty',1),iv(rd,'quantity',0),
                          sv(rd,'expiry_date'),sv(rd,'lot_no')])
                    cnt += 1
                results.append(f'在庫一覧: {cnt}件復元')
            else:
                errors.append('在庫一覧シートなし')

        # ── 発注履歴 ──
        if '発注履歴' in targets:
            headers, data = get_sheet_rows('発注履歴')
            if headers and data:
                db.execute("DELETE FROM order_history")
                cnt = 0
                for row in data:
                    if not any(row): continue
                    rd = dict(zip(headers, row))
                    jan = sv(rd, 'jan')
                    if not jan: continue
                    db.execute("""
                        INSERT INTO order_history
                        (jan,product_name,supplier_cd,supplier_name,product_cd,
                         order_qty,order_date,trigger_type,mail_sent,mail_result)
                        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                    """, [jan,sv(rd,'product_name'),sv(rd,'supplier_cd'),
                          sv(rd,'supplier_name'),sv(rd,'product_cd'),
                          iv(rd,'order_qty',0),sv(rd,'order_date',str(date.today())),
                          sv(rd,'trigger_type','manual'),
                          iv(rd,'mail_sent',0),sv(rd,'mail_result')])
                    cnt += 1
                results.append(f'発注履歴: {cnt}件復元')
            else:
                errors.append('発注履歴シートなし')

        # ── 廃棄退避在庫 ──
        if '廃棄退避在庫' in targets:
            headers, data = get_sheet_rows('廃棄退避在庫')
            if headers and data:
                db.execute("DELETE FROM disposed_stocks")
                cnt = 0
                for row in data:
                    if not any(row): continue
                    rd = dict(zip(headers, row))
                    jan = sv(rd, 'jan')
                    if not jan: continue
                    db.execute("""
                        INSERT INTO disposed_stocks
                        (jan,product_name,supplier_cd,supplier_name,product_cd,
                         quantity,expiry_date,lot_no,reason_type,reason_note,
                         disposed_at,cost_price,loss_amount)
                        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                    """, [jan,sv(rd,'product_name'),sv(rd,'supplier_cd'),sv(rd,'supplier_name'),
                          sv(rd,'product_cd'),iv(rd,'quantity',0),
                          sv(rd,'expiry_date'),sv(rd,'lot_no'),
                          sv(rd,'reason_type'),sv(rd,'reason_note'),
                          sv(rd,'disposed_at',str(date.today())),
                          fv(rd,'cost_price',0),fv(rd,'loss_amount',0)])
                    cnt += 1
                results.append(f'廃棄退避在庫: {cnt}件復元')
            else:
                errors.append('廃棄退避在庫シートなし')

        # ── 入庫履歴 ──
        if '入庫履歴' in targets:
            headers, data = get_sheet_rows('入庫履歴')
            if headers and data:
                db.execute("DELETE FROM stock_movements")
                cnt = 0
                for row in data:
                    if not any(row): continue
                    rd = dict(zip(headers, row))
                    jan = sv(rd, 'jan')
                    if not jan: continue
                    db.execute("""
                        INSERT INTO stock_movements
                        (jan,product_name,move_type,quantity,before_qty,after_qty,
                         note,source_file,move_date)
                        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)
                    """, [jan,sv(rd,'product_name'),sv(rd,'move_type','receipt'),
                          iv(rd,'quantity',0),iv(rd,'before_qty',0),iv(rd,'after_qty',0),
                          sv(rd,'note'),sv(rd,'source_file'),
                          sv(rd,'move_date',str(date.today()))])
                    cnt += 1
                results.append(f'入庫履歴: {cnt}件復元')
            else:
                errors.append('入庫履歴シートなし')

        # ── 棚卸データ ──
        if '棚卸データ' in targets:
            headers, data = get_sheet_rows('棚卸データ')
            if headers and data:
                db.execute("DELETE FROM inventory_count")
                cnt = 0
                for row in data:
                    if not any(row): continue
                    rd = dict(zip(headers, row))
                    jan = sv(rd, 'jan')
                    count_date = sv(rd, 'count_date')
                    if not jan or not count_date: continue
                    db.execute("""
                        INSERT INTO inventory_count
                        (count_date,jan,product_name,system_qty,actual_qty,diff_qty,adjusted,note)
                        VALUES (%s,%s,%s,%s,%s,%s,%s,%s)
                    """, [count_date,jan,sv(rd,'product_name'),
                          iv(rd,'system_qty',0),iv(rd,'actual_qty',0),
                          iv(rd,'diff_qty',0),iv(rd,'adjusted',0),sv(rd,'note')])
                    cnt += 1
                results.append(f'棚卸データ: {cnt}件復元')
            else:
                errors.append('棚卸データシートなし')

        # ── CSV取込ログ ──
        if 'CSV取込ログ' in targets:
            headers, data = get_sheet_rows('CSV取込ログ')
            if headers and data:
                db.execute("DELETE FROM import_logs")
                cnt = 0
                for row in data:
                    if not any(row): continue
                    rd = dict(zip(headers, row))
                    if not rd.get('imported_at'): continue
                    # setting_id FK制約回避: 存在しなければNULL
                    raw_sid = iv(rd, 'setting_id', 0)
                    if raw_sid:
                        _ex = db.execute("SELECT 1 FROM csv_import_settings WHERE id=%s", [raw_sid]).fetchone()
                        sid = raw_sid if _ex else None
                    else:
                        sid = None
                    db.execute("""
                        INSERT INTO import_logs
                        (setting_id, filename, rows_ok, rows_err, status, detail,
                         imported_at, trigger_type)
                        VALUES (%s,%s,%s,%s,%s,%s,%s,%s)
                    """, [sid, sv(rd,'filename'),
                          iv(rd,'rows_ok',0), iv(rd,'rows_err',0),
                          sv(rd,'status','ok'), sv(rd,'detail'),
                          sv(rd,'imported_at'), sv(rd,'trigger_type','auto')])
                    cnt += 1
                results.append(f'CSV取込ログ: {cnt}件復元')
            else:
                errors.append('CSV取込ログシートなし')

        # ── CSVインポートデータ ──
        if 'CSVインポートデータ' in targets:
            headers, data = get_sheet_rows('CSVインポートデータ')
            if headers and data:
                db.execute("DELETE FROM sales_history")
                cnt = 0
                for row in data:
                    if not any(row): continue
                    rd = dict(zip(headers, row))
                    if not sv(rd,'jan'): continue
                    db.execute("""
                        INSERT INTO sales_history
                        (jan, product_name, quantity, sale_date, source_file,
                         chain_cd, client_name, store_cd, store_name, row_hash)
                        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                        ON CONFLICT (row_hash) DO NOTHING
                    """, [sv(rd,'jan'), sv(rd,'product_name'),
                          iv(rd,'quantity',0), sv(rd,'sale_date'),
                          sv(rd,'source_file'), sv(rd,'chain_cd'),
                          sv(rd,'client_name'), sv(rd,'store_cd'),
                          sv(rd,'store_name'), sv(rd,'row_hash')])
                    cnt += 1
                results.append(f'CSVインポートデータ: {cnt}件復元')
            else:
                errors.append('CSVインポートデータシートなし')

        db.commit()
        flash('復元完了: ' + '、'.join(results), 'success')
        for e in errors:
            flash(e, 'warning')

    except Exception as e:
        db.rollback()
        flash(f'復元エラー: {e}', 'error')

    return redirect(url_for('restore'))


@app.route('/admin/bulk_delete', methods=['GET','POST'])
@admin_required
def bulk_delete():
    if request.method == 'GET':
        return render_template('bulk_delete.html')
    db = get_db()
    targets = request.form.getlist('targets')
    if not targets:
        flash('削除対象を選択してください', 'error')
        return redirect(url_for('bulk_delete'))
    results = []
    try:
        if 'alert_logs' in targets:
            db.execute("DELETE FROM alert_logs")
            results.append('アラートログ')
        if 'inventory_count' in targets:
            db.execute("DELETE FROM inventory_count")
            results.append('棚卸データ')
        if 'import_logs' in targets:
            db.execute("DELETE FROM import_logs")
            results.append('CSV取込ログ')
        if 'products' in targets:
            db.execute("DELETE FROM stocks")
            db.execute("DELETE FROM order_history")
            db.execute("DELETE FROM order_pending")
            db.execute("DELETE FROM disposed_stocks")
            db.execute("DELETE FROM stock_movements")
            db.execute("DELETE FROM inventory_count")
            db.execute("DELETE FROM products")
            results.append('商品マスタ（全関連データ含む）')
        else:
            if 'stocks' in targets:
                db.execute("DELETE FROM stocks")
                db.execute("UPDATE products SET ordered_at=''")
                results.append('在庫一覧')
            if 'stock_movements' in targets:
                db.execute("DELETE FROM stock_movements")
                results.append('入庫履歴')
            if 'disposed_stocks' in targets:
                db.execute("DELETE FROM disposed_stocks")
                results.append('廃棄・退避在庫')
            if 'order_history' in targets:
                db.execute("DELETE FROM order_history")
                db.execute("UPDATE order_pending SET status='pending' WHERE status='sent'")
                db.execute("UPDATE products SET ordered_at=''")
                results.append('発注履歴')
            if 'order_pending' in targets and 'order_history' not in targets:
                db.execute("DELETE FROM order_pending")
                db.execute("UPDATE products SET ordered_at=''")
                results.append('発注データ（保留中）')
        db.commit()
        flash('削除完了: ' + '、'.join(results), 'success')
    except Exception as e:
        db.rollback()
        flash(f'エラー: {e}', 'error')
    return redirect(url_for('bulk_delete'))


@app.route('/settings')
@permission_required('settings')
def settings():
    env_path = BASE_DIR / '.env'
    env_content = {}
    if env_path.exists():
        _text = ''
        for _enc in ('utf-8-sig', 'utf-8', 'shift_jis', 'cp932'):
            try:
                _text = env_path.read_text(encoding=_enc)
                break
            except UnicodeDecodeError:
                continue
        for line in _text.splitlines():
            if '=' in line and not line.startswith('#'):
                k,_,v = line.partition('=')
                env_content[k.strip()] = v.strip()
    db = get_db()
    # 保持月数設定を取得
    def get_setting(key, default):
        row = db.execute("SELECT value FROM settings WHERE key=%s", [key]).fetchone()
        return row['value'] if row else str(default)
    retention = {
        'order_history_months': get_setting('order_history_months', 12),
        'disposed_months': get_setting('disposed_months', 12),
        'sales_history_months': get_setting('sales_history_months', 12),
        'csv_log_months': get_setting('csv_log_months', 6),
        'product_template_name': get_setting('product_template_name', '商品マスタ_テンプレート'),
        'product_export_name': get_setting('product_export_name', '商品マスタ'),
        'receipt_template_name': get_setting('receipt_template_name', '入庫一括インポート_テンプレート'),
        'stocktake_export_name': get_setting('stocktake_export_name', '棚卸リスト'),
        'recipients_template_name': get_setting('recipients_template_name', '受信者一覧_テンプレート'),
        'chain_template_name': get_setting('chain_template_name', 'チェーンマスタ_テンプレート'),
        'store_template_name': get_setting('store_template_name', '店舗マスタ_テンプレート'),
        'supplier_setting_template_name': get_setting('supplier_setting_template_name', '仕入先CD設定_テンプレート'),
        'product_setting_template_name':  get_setting('product_setting_template_name', '商品CD設定_テンプレート'),
    }
    forecast_flags = {
        'forecast_ai_mode':      get_setting('forecast_ai_mode', 1),
        'forecast_reorder_mode': get_setting('forecast_reorder_mode', 'sf'),
        'reorder_auto_mode':     get_setting('reorder_auto_mode', 'ai'),
        'safety_level_z':        get_setting('safety_level_z', '1.65'),
        'abc_a_threshold':       get_setting('abc_a_threshold', '0.70'),
        'abc_b_threshold':       get_setting('abc_b_threshold', '0.90'),
    }
    weather_settings = {
        'weather_auto_fetch_enabled':    get_setting('weather_auto_fetch_enabled', '1'),
        'weather_auto_fetch_locations':  get_setting('weather_auto_fetch_locations', ''),
        'weather_auto_fetch_days':       get_setting('weather_auto_fetch_days', '3'),
        'weather_auto_fetch_hour':       get_setting('weather_auto_fetch_hour', '3'),
        'weather_data_retention_days':   get_setting('weather_data_retention_days', '365'),
    }
    return render_template('settings.html', env=env_content, retention=retention,
                           forecast_flags=forecast_flags, weather_settings=weather_settings)

@app.route('/settings/save', methods=['POST'])
@admin_required
def settings_save():
    env_path = BASE_DIR / '.env'
    f = request.form
    lines = [
        f"SECRET_KEY={f.get('SECRET_KEY', app.secret_key)}",
        f"PG_HOST={f.get('PG_HOST','localhost')}",
        f"PG_PORT={f.get('PG_PORT','5432')}",
        f"PG_DBNAME={f.get('PG_DBNAME','inventory')}",
        f"PG_USER={f.get('PG_USER','inventory_user')}",
        f"PG_PASSWORD={f.get('PG_PASSWORD','')}",
        f"MAIL_SERVER={f.get('MAIL_SERVER','')}",
        f"MAIL_PORT={f.get('MAIL_PORT','25')}",
        f"MAIL_USE_TLS={f.get('MAIL_USE_TLS','False')}",
        f"MAIL_USE_SSL={f.get('MAIL_USE_SSL','False')}",
        f"MAIL_AUTH={f.get('MAIL_AUTH','False')}",
        f"MAIL_USERNAME={f.get('MAIL_USERNAME','')}",
        f"MAIL_PASSWORD={f.get('MAIL_PASSWORD','')}",
        f"MAIL_FROM={f.get('MAIL_FROM','')}",
        f"MAIL_FROM_NAME={f.get('MAIL_FROM_NAME','在庫管理システム')}",
        f"DAILY_MAIL_HOUR={f.get('DAILY_MAIL_HOUR','8')}",
        f"DAILY_MAIL_MINUTE={f.get('DAILY_MAIL_MINUTE','0')}",
        f"MONTH_END_IMPORT_HOUR={f.get('MONTH_END_IMPORT_HOUR','5')}",
        f"MONTH_END_IMPORT_MINUTE={f.get('MONTH_END_IMPORT_MINUTE','0')}",
        f"WEATHER_LOCATION={f.get('WEATHER_LOCATION', os.getenv('WEATHER_LOCATION','東京'))}",
        f"WEATHER_LAT={f.get('WEATHER_LAT', os.getenv('WEATHER_LAT','35.6897'))}",
        f"WEATHER_LON={f.get('WEATHER_LON', os.getenv('WEATHER_LON','139.6922'))}",
        f"WEATHER_LOCATIONS_JSON={f.get('WEATHER_LOCATIONS_JSON', os.getenv('WEATHER_LOCATIONS_JSON',''))}",
        f"USE_WAITRESS={os.getenv('USE_WAITRESS','1')}",
    ]
    env_path.write_text('\n'.join(lines), encoding='utf-8')
    # 保持月数をDBに保存
    db = get_db()
    int_keys = {'order_history_months', 'disposed_months', 'sales_history_months', 'csv_log_months',
                'forecast_ai_mode', 'weather_auto_fetch_enabled', 'weather_auto_fetch_days',
                'weather_auto_fetch_hour', 'weather_data_retention_days'}
    float_keys = {'safety_level_z', 'abc_a_threshold', 'abc_b_threshold'}
    for key, default in [
        ('order_history_months', 12), ('disposed_months', 12), ('sales_history_months', 12),
        ('csv_log_months', 6),
        ('product_template_name', '商品マスタ_テンプレート'),
        ('product_export_name', '商品マスタ'),
        ('receipt_template_name', '入庫一括インポート_テンプレート'),
        ('stocktake_export_name', '棚卸リスト'),
        ('recipients_template_name', '受信者一覧_テンプレート'),
        ('chain_template_name', 'チェーンマスタ_テンプレート'),
        ('store_template_name', '店舗マスタ_テンプレート'),
        ('supplier_setting_template_name', '仕入先CD設定_テンプレート'),
        ('product_setting_template_name', '商品CD設定_テンプレート'),
        ('forecast_ai_mode', 1),
        ('forecast_reorder_mode', 'sf'),
        ('reorder_auto_mode', 'ai'),
        ('safety_level_z', '1.65'),
        ('abc_a_threshold', '0.70'),
        ('abc_b_threshold', '0.90'),
        ('weather_auto_fetch_enabled', '1'),
        ('weather_auto_fetch_locations', ''),
        ('weather_auto_fetch_days', '3'),
        ('weather_auto_fetch_hour', '3'),
        ('weather_data_retention_days', '365'),
    ]:
        if key == 'forecast_ai_mode':
            raw = '1' if f.get(key) == '1' else '0'
        elif key == 'weather_auto_fetch_enabled':
            raw = '1' if f.get(key) else '0'
        elif key == 'forecast_reorder_mode':
            raw = f.get(key, 'sf') or 'sf'
        elif key == 'reorder_auto_mode':
            raw = f.get(key, 'ai') or 'ai'
            if raw not in ('ai', 'ly', 'manual'):
                raw = 'ai'
        else:
            raw = f.get(key, default) or default
        if key in int_keys:
            try:
                val = str(int(raw))
            except Exception:
                val = str(default)
        elif key in float_keys:
            try:
                val = str(float(raw))
            except Exception:
                val = str(default)
        else:
            val = str(raw).strip()
        existing = db.execute("SELECT id FROM settings WHERE key=%s", [key]).fetchone()
        if existing:
            db.execute("UPDATE settings SET value=%s WHERE key=%s", [val, key])
        else:
            db.execute("INSERT INTO settings (key, value) VALUES (%s,%s)", [key, val])
    db.commit()

    # reorder_auto_mode に応じて全商品を一括更新
    # ※ 設定保存用の接続とは別の新規接続で実行
    new_ram = f.get('reorder_auto_mode', 'ai') or 'ai'
    _mode_map = {'ai': (1, '🤖 AIモード'), 'ly': (2, '📅 前年実績モード'), 'manual': (0, '🚫 手動')}
    bulk_cnt = 0
    if new_ram in _mode_map:
        mode_val, mode_label = _mode_map[new_ram]
        try:
            bulk_cnt = _bulk_set_reorder_auto(mode_val, mode_label)
            logger.info(f'[settings] reorder_auto_mode={new_ram}: {bulk_cnt}商品を{mode_label}に一括更新')
        except Exception as _e:
            logger.error(f'[settings] reorder_auto_mode={new_ram} 一括更新エラー: {_e}')
            flash(f'設定は保存しましたが、商品の一括更新でエラーが発生しました: {_e}', 'danger')
            invalidate_forecast_cache()
            return redirect(url_for('settings'))

    invalidate_forecast_cache()
    mode_label = _mode_map.get(new_ram, (None, new_ram))[1]
    flash(f'設定を保存しました。全 {bulk_cnt} 商品の発注点自動更新を「{mode_label}」に切り替えました。', 'success')
    return redirect(url_for('settings'))


def _bulk_set_reorder_auto(mode_val, mode_label):
    """全商品の reorder_auto を一括更新する共通処理"""
    import psycopg2 as _pg2
    from database import get_dsn as _get_dsn
    _conn = _pg2.connect(**_get_dsn())
    _conn.autocommit = False
    _cur = _conn.cursor()
    _cur.execute("UPDATE products SET reorder_auto=%s WHERE is_active=1", [mode_val])
    cnt = _cur.rowcount
    _conn.commit()
    _cur.close()
    _conn.close()
    return cnt


@app.route('/settings/apply_manual_all', methods=['POST'])
@admin_required
def settings_apply_manual_all():
    """全商品の reorder_auto を 0（手動）に一括更新"""
    try:
        cnt = _bulk_set_reorder_auto(0, '手動')
        logger.info(f'[apply_manual_all] {cnt}商品を手動モードに一括更新')
        flash(f'✅ 全 {cnt} 商品の発注点自動更新を「🚫 手動」に切り替えました。', 'success')
    except Exception as e:
        logger.error(f'[apply_manual_all] エラー: {e}')
        flash(f'❌ 一括更新エラー: {e}', 'danger')
    return redirect(url_for('settings'))


@app.route('/settings/apply_ai_all', methods=['POST'])
@admin_required
def settings_apply_ai_all():
    """全商品の reorder_auto を 1（AIモード）に一括更新"""
    try:
        cnt = _bulk_set_reorder_auto(1, 'AIモード')
        logger.info(f'[apply_ai_all] {cnt}商品をAIモードに一括更新')
        flash(f'✅ 全 {cnt} 商品の発注点自動更新を「🤖 AIモード」に切り替えました。', 'success')
    except Exception as e:
        logger.error(f'[apply_ai_all] エラー: {e}')
        flash(f'❌ 一括更新エラー: {e}', 'danger')
    return redirect(url_for('settings'))


@app.route('/settings/apply_ly_all', methods=['POST'])
@admin_required
def settings_apply_ly_all():
    """全商品の reorder_auto を 2（前年実績モード）に一括更新"""
    try:
        cnt = _bulk_set_reorder_auto(2, '前年実績')
        logger.info(f'[apply_ly_all] {cnt}商品を前年実績モードに一括更新')
        flash(f'✅ 全 {cnt} 商品の発注点自動更新を「📅 前年実績モード」に切り替えました。', 'success')
    except Exception as e:
        logger.error(f'[apply_ly_all] エラー: {e}')
        flash(f'❌ 一括更新エラー: {e}', 'danger')
    return redirect(url_for('settings'))


@app.route('/settings/test_pg', methods=['POST'])
@admin_required
def test_pg():
    from database import get_dsn
    import psycopg2
    try:
        conn = psycopg2.connect(**get_dsn())
        cur = conn.cursor()
        cur.execute("SELECT version()")
        ver = cur.fetchone()[0]
        conn.close()
        return {'ok': True, 'msg': ver}
    except Exception as e:
        return {'ok': False, 'msg': str(e)}

@app.route('/settings/test_mail', methods=['POST'])
@admin_required
def test_mail():
    from mail_service import _get_cfg, _send_smtp
    from email.mime.text import MIMEText
    from email.header import Header
    cfg = _get_cfg()
    to_addr = request.form.get('test_to', '') or cfg['from_addr'] or cfg['username']
    try:
        if not cfg['server']:
            raise ValueError('SMTPサーバーが未設定です')
        from_addr = cfg['from_addr'] or cfg['username']
        if not from_addr:
            raise ValueError('送信元メールアドレスが未設定です')
        msg = MIMEText('在庫管理システムからのテスト送信です。', 'plain', 'utf-8')
        msg['Subject'] = Header('【テスト】在庫管理システム メール接続確認', 'utf-8')
        msg['From'] = from_addr
        msg['To']   = to_addr
        _send_smtp(cfg, from_addr, [to_addr], msg.as_string())
        flash(f'✅ テストメール送信成功！（宛先: {to_addr}）', 'success')
    except Exception as e:
        flash(f'❌ 接続エラー: {e}', 'danger')
    return redirect(url_for('settings'))





# ─── 起動 ────────────────────────────────────────────────────────





@app.route('/api/products/lookup')
@login_required
def api_products_lookup():
    q = (request.args.get('q') or '').strip()
    if not q:
        return jsonify({'ok': False})
    db = get_db()
    p = _resolve_product_by_code(db, q)
    if not p:
        return jsonify({'ok': False})
    return jsonify({'ok': True, 'jan': p['jan'], 'product_cd': p['product_cd'], 'product_name': p['product_name'], 'supplier_name': p['supplier_name']})

@app.route('/reports/forecast')
@permission_required('reports')
def reports_forecast():
    db = get_db()
    q = request.args.get('q','').strip().lower()
    rows = _build_forecast_rows(db, q)
    promo_rows = db.execute("SELECT pp.*, p.product_name, p.product_cd, p.supplier_name FROM promotion_plans pp JOIN products p ON p.jan=pp.jan WHERE pp.promo_date >= CURRENT_DATE - INTERVAL '7 days' ORDER BY pp.promo_date, p.supplier_name, p.product_cd").fetchall()
    demand_rows = db.execute("SELECT dp.*, p.product_name, p.product_cd, p.supplier_name FROM demand_plans dp JOIN products p ON p.jan=dp.jan WHERE dp.demand_date >= CURRENT_DATE - INTERVAL '7 days' ORDER BY dp.demand_date, p.supplier_name, p.product_cd").fetchall()
    flags_for_tmpl = _get_forecast_feature_flags(db)
    forecast_mode = 'ai' if flags_for_tmpl.get('forecast_ai_mode', True) else 'ly'
    return render_template('forecast.html', rows=rows, q=q, promo_rows=promo_rows, demand_rows=demand_rows,
                           forecast_mode=forecast_mode)

@app.route('/reports/forecast/apply', methods=['POST'])
@permission_required('reports')
def reports_forecast_apply():
    import uuid as _uuid
    import calendar as _calendar
    q         = request.form.get('q', '').strip().lower()
    mode      = (request.form.get('mode') or 'reorder_point').strip()
    calc_mode = (request.form.get('calc_mode') or 'ai').strip()   # 'ai' or 'ly'
    job_id = str(_uuid.uuid4())
    _csv_progress_push(job_id, {'phase': 'start', 'file': '需要予測 一括反映', 'total': 0})

    def _run():
        import psycopg2
        from database import get_dsn, DBConn
        from datetime import date as _date
        updated = 0
        try:
            conn = psycopg2.connect(**get_dsn())
            conn.autocommit = False
            db = DBConn(conn)
        except Exception as e:
            _csv_progress_push(job_id, {'phase': 'finished', 'results': [{'name': 'エラー', 'status': 'err', 'detail': str(e)}]})
            return
        try:
            label = 'AI' if calc_mode == 'ai' else '前年実績'
            if calc_mode == 'ly':
                # 前年実績モード: 前年同月の出荷実績から発注点を計算
                today     = _date.today()
                last_year = today.year - 1
                month     = today.month
                days_in_month = _calendar.monthrange(last_year, month)[1]
                ly_start  = f'{last_year}-{month:02d}-01'
                ly_end    = f'{last_year}-{month:02d}-{days_in_month:02d}'

                sql = "SELECT * FROM products WHERE is_active=1"
                params = []
                if q:
                    sql += " AND (LOWER(product_name) LIKE %s OR LOWER(jan) LIKE %s OR LOWER(product_cd) LIKE %s)"
                    params += [f'%{q}%', f'%{q}%', f'%{q}%']
                prods = db.execute(sql, params).fetchall()
                rows = []
                for p in prods:
                    total = db.execute(
                        "SELECT COALESCE(SUM(quantity),0) AS s FROM sales_history WHERE jan=%s AND sale_date BETWEEN %s AND %s",
                        [p['jan'], ly_start, ly_end]
                    ).fetchone()['s']
                    daily = total / days_in_month if days_in_month else 0
                    rp = daily * p['lead_time_days'] * p['safety_factor']
                    # ロット丸め
                    lot = p['lot_size'] or 0
                    if lot > 1:
                        import math as _math
                        rp = _math.ceil(rp / lot) * lot
                    rows.append({'product_id': p['id'], 'suggested_reorder_point': int(round(rp)), 'suggested_order_qty': int(round(rp))})
            else:
                rows = _build_forecast_rows(db, q)

            total = len(rows)
            _csv_progress_push(job_id, {'phase': 'start', 'file': f'需要予測 一括反映（{label}）', 'total': total})
            updated = 0
            for i, r in enumerate(rows, 1):
                try:
                    if mode == 'both':
                        db.execute("UPDATE products SET reorder_point=%s, order_qty=%s WHERE id=%s",
                                   [int(r['suggested_reorder_point'] or 0), int(r['suggested_order_qty'] or 0), r['product_id']])
                    elif mode == 'order_qty':
                        db.execute("UPDATE products SET order_qty=%s WHERE id=%s",
                                   [int(r['suggested_order_qty'] or 0), r['product_id']])
                    else:
                        db.execute("UPDATE products SET reorder_point=%s WHERE id=%s",
                                   [int(r['suggested_reorder_point'] or 0), r['product_id']])
                    updated += 1
                except Exception:
                    pass
                if i % 50 == 0 or i == total:
                    _csv_progress_push(job_id, {'phase': 'progress', 'current': i, 'total': total, 'ok': updated, 'skip': 0, 'err': i - updated})
            db.commit()
            invalidate_forecast_cache()   # 発注点/発注数更新後にキャッシュを破棄
            _csv_progress_push(job_id, {'phase': 'done', 'file': f'需要予測 一括反映（{label}）', 'ok': updated, 'skip': 0, 'err': 0,
                                        'detail': f'{updated}/{total}件 更新完了'})
        except Exception as e:
            try:
                db.rollback()
            except Exception:
                pass
            _csv_progress_push(job_id, {'phase': 'done', 'file': '需要予測 一括反映', 'ok': 0, 'skip': 0, 'err': 1,
                                        'detail': f'エラー: {e}'})
        finally:
            try:
                db.close()
            except Exception:
                pass
        _csv_progress_push(job_id, {'phase': 'finished', 'results': [{'name': f'需要予測 一括反映（{label}）', 'status': 'ok', 'detail': f'{updated}件更新'}]})

    t = threading.Thread(target=_run, daemon=True)
    t.start()
    return redirect(url_for('forecast_apply_progress', job_id=job_id, q=q, mode=mode))


@app.route('/reports/forecast/apply/progress/<job_id>')
@permission_required('reports')
def forecast_apply_progress(job_id):
    q    = request.args.get('q', '')
    mode = request.args.get('mode', 'reorder_point')
    return render_template('forecast_apply_progress.html', job_id=job_id, q=q, mode=mode)


@app.route('/reports/forecast/apply/progress/<job_id>/stream')
def forecast_apply_progress_stream(job_id):
    def generate():
        sent = 0
        for _ in range(3600):
            with _csv_lock:
                events = _csv_progress.get(job_id, [])
            while sent < len(events):
                ev = events[sent]
                yield f"data: {json.dumps(ev, ensure_ascii=False)}\n\n"
                sent += 1
                if ev.get('phase') == 'finished':
                    return
            _time.sleep(0.3)
        yield f"data: {json.dumps({'phase': 'timeout'})}\n\n"
    return Response(generate(), mimetype='text/event-stream',
                    headers={'Cache-Control': 'no-cache', 'X-Accel-Buffering': 'no'})


@app.route('/reports/forecast/promotions', methods=['POST'])
@permission_required('reports')
def reports_forecast_promotion_add():
    db = get_db()
    jan = (request.form.get('jan') or '').strip()
    promo_date = (request.form.get('promo_date') or '').strip()
    promo_name = (request.form.get('promo_name') or '').strip()
    uplift_factor = float(request.form.get('uplift_factor') or 1.0)
    product = _resolve_product_by_code(db, jan)
    if not product:
        flash('対象商品が見つかりません。', 'danger')
        return redirect(url_for('reports_forecast'))
    jan = product['jan']
    db.execute("INSERT INTO promotion_plans (jan, promo_date, promo_name, uplift_factor) VALUES (%s,%s,%s,%s) ON CONFLICT (jan, promo_date, promo_name) DO UPDATE SET uplift_factor=EXCLUDED.uplift_factor", [jan, promo_date, promo_name, uplift_factor])
    db.commit()
    flash(f"{product['product_name']} の販促予定を登録しました。", 'success')
    return redirect(url_for('reports_forecast', q=jan))

@app.route('/reports/forecast/promotions/<int:promo_id>/delete', methods=['POST'])
@permission_required('reports')
def reports_forecast_promotion_delete(promo_id):
    db = get_db()
    db.execute("DELETE FROM promotion_plans WHERE id=%s", [promo_id])
    db.commit()
    flash('販促予定を削除しました。', 'warning')
    return redirect(url_for('reports_forecast'))


@app.route('/reports/forecast/promotions/template')
@permission_required('reports')
def reports_forecast_promotion_template():
    import io, csv
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(['JANコード', '商品CD', '販促日', '販促名', '上振れ係数'])
    writer.writerow(['4901234567890', 'P001', '2026-04-01', '特売', '1.3'])
    writer.writerow(['4901234567891', 'P002', '2026-04-05', 'チラシ', '1.5'])
    output.seek(0)
    from urllib.parse import quote
    fname = '販促予定一括インポート_テンプレート.csv'
    return Response(
        output.getvalue().encode('utf-8-sig'),
        mimetype='text/csv',
        headers={'Content-Disposition': f"attachment; filename*=UTF-8''{quote(fname)}"}
    )

@app.route('/reports/forecast/promotions/import', methods=['POST'])
@permission_required('reports')
def reports_forecast_promotion_import():
    db = get_db()
    f = request.files.get('file')
    if not f or not f.filename:
        flash('インポートファイルを選択してください。', 'danger')
        return redirect(url_for('reports_forecast'))
    rows = []
    name = (f.filename or '').lower()
    try:
        if name.endswith('.csv'):
            content = f.read().decode('utf-8-sig', errors='ignore')
            rows = list(csv.DictReader(io.StringIO(content)))
        elif name.endswith(('.xlsx', '.xls')):
            import openpyxl
            wb = openpyxl.load_workbook(f, data_only=True)
            ws = wb.active
            headers = [str(ws.cell(1, c).value or '').strip() for c in range(1, ws.max_column + 1)]
            for r in range(2, ws.max_row + 1):
                row = {headers[c-1]: ws.cell(r, c).value for c in range(1, ws.max_column + 1)}
                if any(v not in (None, '') for v in row.values()):
                    rows.append(row)
        else:
            flash('CSV または Excel ファイルを選択してください。', 'danger')
            return redirect(url_for('reports_forecast'))
    except Exception as e:
        flash(f'販促予定インポートに失敗しました: {e}', 'danger')
        return redirect(url_for('reports_forecast'))
    created = 0
    errors = []
    for i, row in enumerate(rows, 2):
        jan_code = _normalize_jan(row.get('JAN') or row.get('JANコード') or row.get('jan') or '')
        product_cd_code = str(row.get('商品CD') or row.get('product_cd') or '').strip()
        promo_date = _normalize_date(row.get('販促日') or row.get('promo_date') or '')
        promo_name = str(row.get('販促名') or row.get('promo_name') or '').strip()
        try:
            uplift_factor = float(row.get('上振れ係数') or row.get('uplift_factor') or 1.3)
        except Exception:
            uplift_factor = 1.3
        # JANコードで検索→なければ商品CDでフォールバック
        product = (_resolve_product_by_code(db, jan_code) if jan_code else None) or \
                  (_resolve_product_by_code(db, product_cd_code) if product_cd_code else None)
        used_code = jan_code or product_cd_code
        if not product or not promo_date:
            errors.append(f'{i}行目: 商品/JAN={repr(used_code)} 日付={repr(promo_date)} → 確認してください')
            continue
        if uplift_factor < 1.0:
            uplift_factor = 1.0
        db.execute(
            "INSERT INTO promotion_plans (jan, promo_date, promo_name, uplift_factor) VALUES (%s,%s,%s,%s) "
            "ON CONFLICT (jan, promo_date, promo_name) DO UPDATE SET uplift_factor=EXCLUDED.uplift_factor",
            [product['jan'], promo_date, promo_name, uplift_factor]
        )
        created += 1
    db.commit()
    if errors:
        flash(f'販促予定を {created} 件取り込みました。エラー {len(errors)} 件', 'warning')
    else:
        flash(f'販促予定を {created} 件取り込みました。', 'success')
    return redirect(url_for('reports_forecast'))


@app.route('/reports/forecast/promotions/clear', methods=['POST'])
@permission_required('reports')
def reports_forecast_promotion_clear():
    db = get_db()
    db.execute("DELETE FROM promotion_plans")
    db.commit()
    flash('販促予定を全件削除しました。', 'warning')
    return redirect(url_for('reports_forecast'))

@app.route('/reports/forecast/promotions/bulk_delete', methods=['POST'])
@permission_required('reports')
def reports_forecast_promotion_bulk_delete():
    db = get_db()
    ids = [int(x) for x in request.form.getlist('promo_id') if str(x).isdigit()]
    if not ids:
        flash('削除する項目を選択してください。', 'warning')
        return redirect(url_for('reports_forecast'))
    placeholders = ','.join(['%s'] * len(ids))
    db.execute(f"DELETE FROM promotion_plans WHERE id IN ({placeholders})", ids)
    db.commit()
    flash(f'販促予定を {len(ids)} 件削除しました。', 'warning')
    return redirect(url_for('reports_forecast'))

@app.route('/reports/forecast/demands', methods=['POST'])
@permission_required('reports')
def reports_forecast_demand_add():
    db = get_db()
    jan = (request.form.get('jan') or '').strip()
    demand_date = (request.form.get('demand_date') or '').strip()
    demand_qty = _to_int(request.form.get('demand_qty'))
    customer_name = (request.form.get('customer_name') or '').strip()
    note = (request.form.get('note') or '').strip()
    product = _resolve_product_by_code(db, jan)
    if not product or not demand_date or demand_qty <= 0:
        flash('受注予定の登録内容が不正です。', 'danger')
        return redirect(url_for('reports_forecast', q=jan))
    jan = product['jan']
    db.execute("INSERT INTO demand_plans (jan, demand_date, demand_qty, demand_type, customer_name, note, created_by) VALUES (%s,%s,%s,'order',%s,%s,%s)", [jan, demand_date, demand_qty, customer_name, note, session.get('user','')])
    db.commit()
    flash(f"{product['product_name']} の受注予定を登録しました。", 'success')
    return redirect(url_for('reports_forecast', q=jan))



@app.route('/reports/forecast/demands/selected_delete', methods=['POST'])
@permission_required('reports')
def reports_forecast_demand_selected_delete():
    db = get_db()
    ids = [int(x) for x in request.form.getlist('demand_id') if str(x).isdigit()]
    if not ids:
        flash('削除する項目を選択してください。', 'warning')
        return redirect(url_for('reports_forecast'))
    placeholders = ','.join(['%s'] * len(ids))
    db.execute(f"DELETE FROM demand_plans WHERE id IN ({placeholders})", ids)
    db.commit()
    flash(f'受注予定を {len(ids)} 件削除しました。', 'warning')
    return redirect(url_for('reports_forecast'))

@app.route('/reports/forecast/demands/template')
@permission_required('reports')
def reports_forecast_demand_template():
    import io, csv
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(['JANコード', '商品CD', '数量', '得意先', '受注日', '備考'])
    output.seek(0)
    from urllib.parse import quote
    fname = '受注予定一括インポート_テンプレート.csv'
    return Response(
        output.getvalue().encode('utf-8'),
        mimetype='text/csv',
        headers={'Content-Disposition': f"attachment; filename*=UTF-8''{quote(fname)}"}
    )

@app.route('/reports/forecast/demands/import', methods=['POST'])
@permission_required('reports')
def reports_forecast_demand_import():
    db = get_db()
    f = request.files.get('file')
    if not f or not f.filename:
        flash('インポートファイルを選択してください。', 'danger')
        return redirect(url_for('reports_forecast'))

    rows = []
    name = (f.filename or '').lower()
    try:
        if name.endswith('.csv'):
            content = f.read().decode('utf-8-sig', errors='ignore')
            rows = list(csv.DictReader(io.StringIO(content)))
        elif name.endswith(('.xlsx', '.xls')):
            import openpyxl
            wb = openpyxl.load_workbook(f, data_only=True)
            ws = wb.active
            headers = [str(ws.cell(1, c).value or '').strip() for c in range(1, ws.max_column + 1)]
            for r in range(2, ws.max_row + 1):
                row = {headers[c-1]: ws.cell(r, c).value for c in range(1, ws.max_column + 1)}
                if any(v not in (None, '') for v in row.values()):
                    rows.append(row)
        else:
            flash('CSV または Excel ファイルを選択してください。', 'danger')
            return redirect(url_for('reports_forecast'))
    except Exception as e:
        flash(f'受注予定インポートに失敗しました: {e}', 'danger')
        return redirect(url_for('reports_forecast'))

    created = 0
    errors = []
    for i, row in enumerate(rows, 2):
        jan_code = _normalize_jan(row.get('JAN') or row.get('JANコード') or row.get('jan') or '')
        product_cd_code = str(row.get('商品CD') or row.get('product_cd') or '').strip()
        demand_date = _normalize_date(row.get('受注日') or row.get('需要日') or row.get('demand_date') or '')
        demand_qty = _to_int(row.get('数量') or row.get('受注数量') or row.get('demand_qty'))
        customer_name = str(row.get('得意先') or row.get('customer_name') or '').strip()
        note = str(row.get('備考') or row.get('note') or '').strip()
        # JANコードで検索→なければ商品CDでフォールバック
        product = (_resolve_product_by_code(db, jan_code) if jan_code else None) or \
                  (_resolve_product_by_code(db, product_cd_code) if product_cd_code else None)
        used_code = jan_code or product_cd_code
        if not product or not demand_date or demand_qty <= 0:
            errors.append(f'{i}行目: 商品/JAN={repr(used_code)} 日付={repr(demand_date)} 数量={demand_qty} → 確認してください')
            continue
        db.execute("INSERT INTO demand_plans (jan, demand_date, demand_qty, demand_type, customer_name, note, created_by) VALUES (%s,%s,%s,'order',%s,%s,%s)", [product['jan'], demand_date, demand_qty, customer_name, note, session.get('user','')])
        created += 1
    db.commit()
    if errors:
        flash(f'受注予定を {created} 件取り込みました。エラー {len(errors)} 件', 'warning')
    else:
        flash(f'受注予定を {created} 件取り込みました。', 'success')
    return redirect(url_for('reports_forecast'))

@app.route('/reports/forecast/demands/bulk_delete', methods=['POST'])
@permission_required('reports')
def reports_forecast_demand_bulk_delete():
    db = get_db()
    ids = [int(x) for x in request.form.getlist('demand_id') if str(x).isdigit()]
    if ids:
        placeholders = ','.join(['%s'] * len(ids))
        db.execute(f"DELETE FROM demand_plans WHERE id IN ({placeholders})", ids)
        deleted = len(ids)
    else:
        db.execute("DELETE FROM demand_plans WHERE demand_date >= CURRENT_DATE - INTERVAL '7 days'")
        deleted = db.rowcount if getattr(db, 'rowcount', None) is not None else 0
    db.commit()
    flash(f'受注予定を削除しました。', 'warning')
    return redirect(url_for('reports_forecast'))

@app.route('/reports/forecast/demands/<int:demand_id>/delete', methods=['POST'])
@permission_required('reports')
def reports_forecast_demand_delete(demand_id):
    db = get_db()
    db.execute("DELETE FROM demand_plans WHERE id=%s", [demand_id])
    db.commit()
    flash('受注予定を削除しました。', 'warning')
    return redirect(url_for('reports_forecast'))


@app.route('/reports/forecast/demands/clear', methods=['POST'])
@permission_required('reports')
def reports_forecast_demand_clear():
    db = get_db()
    db.execute("DELETE FROM demand_plans")
    db.commit()
    flash('受注予定を全件削除しました。', 'warning')
    return redirect(url_for('reports_forecast'))

@app.route('/reports/shortage')
@permission_required('reports')
def reports_shortage():
    db = get_db()
    q = request.args.get('q','').strip().lower()
    rows = _build_shortage_rows(db, q)
    return render_template('shortage.html', rows=rows, q=q, today=str(date.today()))

@app.route('/reports/shortage/create_orders', methods=['POST'])
@permission_required('reports')
def reports_shortage_create_orders():
    db = get_db()
    q = request.form.get('q','').strip().lower()
    selected = set(request.form.getlist('selected_jan'))
    rows = _build_shortage_rows(db, q)
    created = 0
    for r in rows:
        if selected and r['jan'] not in selected:
            continue
        if r['risk_level'] not in ('入荷前欠品','予測欠品','要注意'):
            continue
        open_pending = db.execute("SELECT id FROM order_pending WHERE jan=%s AND status='pending' ORDER BY id DESC LIMIT 1", [r['jan']]).fetchone()
        if open_pending:
            continue
        order_qty = max(int(r.get('order_qty') or 0), int((r.get('forecast_30d') or 0) - max(r.get('projected_30d_stock') or 0, 0)))
        if order_qty <= 0:
            order_qty = max(int(r.get('reorder_point') or 0), 1)
        prod = db.execute("SELECT * FROM products WHERE jan=%s", [r['jan']]).fetchone()
        db.execute("""
            INSERT INTO order_pending
            (supplier_cd,supplier_name,supplier_email,mixed_group,mixed_lot_mode,mixed_lot_cases,jan,product_cd,product_name,order_qty,order_cases,trigger_type,pending_since,force_send_date,status)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,'shortage_forecast',%s,%s,'pending')
        """, [prod['supplier_cd'], prod['supplier_name'], prod.get('supplier_email',''), prod.get('mixed_group',''), prod.get('mixed_lot_mode','gte'), int(prod.get('mixed_lot_cases') or 0), prod['jan'], prod['product_cd'], prod['product_name'], order_qty, 0, str(date.today()), str(date.today() + timedelta(days=1))])
        created += 1
    db.commit()
    flash(f'欠品予測から {created} 件の発注候補を作成しました。', 'success')
    return redirect(url_for('reports_shortage', q=q))













# ─── sales_history 移行ツール ─────────────────────────────────────
@app.route('/sales_history/import', methods=['GET', 'POST'])
@admin_required
def sales_history_import_page():
    if request.method == 'GET':
        return render_template('sales_history_import.html')
    # POST: バックグラウンドインポート
    import uuid as _uuid
    f = request.files.get('file')
    if not f or not f.filename:
        flash('ファイルを選択してください。', 'danger')
        return render_template('sales_history_import.html')
    content = f.read()
    job_id = str(_uuid.uuid4())
    _csv_progress_push(job_id, {'phase': 'start', 'file': f.filename, 'total': 0})

    def _run():
        import csv as _csv, io as _io, hashlib as _hl, psycopg2
        from database import get_dsn, DBConn
        ok = skip = err = 0
        try:
            conn = psycopg2.connect(**get_dsn())
            conn.autocommit = False
            db = DBConn(conn)
        except Exception as e:
            _csv_progress_push(job_id, {'phase': 'finished', 'results': [{'name': 'エラー', 'status': 'err', 'detail': str(e)}]})
            return
        try:
            text = None
            for enc in ('utf-8-sig', 'utf-8', 'cp932'):
                try:
                    text = content.decode(enc)
                    break
                except Exception:
                    pass
            if text is None:
                text = content.decode('utf-8', errors='ignore')
            rows_all = list(_csv.DictReader(_io.StringIO(text)))
            total = len(rows_all)
            _csv_progress_push(job_id, {'phase': 'start', 'file': f.filename, 'total': total})
            ok = skip = err = 0
            for i, row in enumerate(rows_all, 1):
                try:
                    jan         = (row.get('JANコード') or row.get('jan') or '').strip()
                    product_name= (row.get('商品名') or row.get('product_name') or '').strip()
                    qty_raw     = row.get('数量') or row.get('quantity') or ''
                    sale_date   = (row.get('販売日') or row.get('sale_date') or '').strip()
                    source_file = (row.get('ソースファイル名') or row.get('source_file') or f.filename).strip()
                    if not jan or not sale_date or qty_raw == '':
                        skip += 1
                        continue
                    quantity = int(float(str(qty_raw).replace(',', '')))
                    # 重複チェック用ハッシュ
                    row_hash = _hl.sha256(f'{jan}|{sale_date}|{quantity}|{source_file}'.encode()).hexdigest()
                    db.execute("""
                        INSERT INTO sales_history (jan, product_name, quantity, sale_date, source_file, row_hash)
                        VALUES (%s, %s, %s, %s, %s, %s)
                        ON CONFLICT DO NOTHING
                    """, [jan, product_name, quantity, sale_date, source_file, row_hash])
                    ok += 1
                except Exception as e2:
                    err += 1
                if i % 100 == 0 or i == total:
                    _csv_progress_push(job_id, {'phase': 'progress', 'current': i, 'total': total, 'ok': ok, 'skip': skip, 'err': err})
            db.commit()
            _csv_progress_push(job_id, {'phase': 'done', 'file': f.filename, 'ok': ok, 'skip': skip, 'err': err,
                                        'detail': f'{ok}件取込 スキップ{skip}件 エラー{err}件'})
        except Exception as e:
            try:
                db.rollback()
            except Exception:
                pass
            _csv_progress_push(job_id, {'phase': 'done', 'file': f.filename, 'ok': 0, 'skip': 0, 'err': 1,
                                        'detail': f'エラー: {e}'})
        finally:
            try:
                db.close()
            except Exception:
                pass
        _csv_progress_push(job_id, {'phase': 'finished', 'results': [{'name': f.filename, 'status': 'ok', 'detail': f'{ok}件取込完了'}]})

    threading.Thread(target=_run, daemon=True).start()
    return redirect(url_for('sales_history_import_progress', job_id=job_id))


@app.route('/sales_history/import/progress/<job_id>')
@admin_required
def sales_history_import_progress(job_id):
    return render_template('sales_history_import.html', job_id=job_id)


@app.route('/sales_history/import/progress/<job_id>/stream')
def sales_history_import_stream(job_id):
    def generate():
        sent = 0
        for _ in range(3600):
            with _csv_lock:
                events = _csv_progress.get(job_id, [])
            while sent < len(events):
                ev = events[sent]
                yield f"data: {json.dumps(ev, ensure_ascii=False)}\n\n"
                sent += 1
                if ev.get('phase') == 'finished':
                    return
            _time.sleep(0.3)
        yield f"data: {json.dumps({'phase': 'timeout'})}\n\n"
    return Response(generate(), mimetype='text/event-stream',
                    headers={'Cache-Control': 'no-cache', 'X-Accel-Buffering': 'no'})


@app.route('/sales_history/import/template')
@admin_required
def sales_history_import_template():
    import io as _io, csv as _csv
    from urllib.parse import quote
    sio = _io.StringIO()
    w = _csv.writer(sio)
    w.writerow(['JANコード', '商品名', '数量', '販売日', 'ソースファイル名'])
    w.writerow(['4901234567890', 'サンプル商品A', '10', '2024-01-15', '2024年1月売上.csv'])
    w.writerow(['4901234567891', 'サンプル商品B', '5', '2024-01-15', '2024年1月売上.csv'])
    return Response(sio.getvalue().encode('utf-8-sig'), mimetype='text/csv',
                    headers={'Content-Disposition': "attachment; filename*=UTF-8''" + quote('sales_history_テンプレート.csv')})


if __name__ == '__main__':
    port = int(os.getenv('PORT', 5000))
    use_waitress = os.getenv('USE_WAITRESS', '1') == '1'

    logger.info("\n  ========================================")
    logger.info("  Inventory System v3.0")
    logger.info(f"  URL: http://localhost:{port}")
    logger.info("  ========================================")
    # DB接続確認
    try:
        logger.info("  Connecting to PostgreSQL...")
        init_db()
        logger.info("  OK: Database ready.")
    except Exception as e:
        logger.warning("  [ERROR] PostgreSQL connection failed: %s", e)
        logger.warning("  Possible causes:")
        logger.warning("    1. PostgreSQL server is not running")
        logger.warning("    2. Wrong connection info in .env file")
        logger.warning("  The web server will still start. Fix the .env file and restart.")

    try:
        start_scheduler()
    except Exception as e:
        logger.warning(f"  [WARN] Scheduler failed to start: {e}")
    # ── 起動時バックグラウンドウォームアップ ──────────────────────────────
    # DB接続成功後のみ実行。ユーザーの最初のアクセスまでにキャッシュを温める
    try:
        threading.Thread(target=_bg_refresh_sales_daily_agg,  daemon=True).start()
        threading.Thread(target=_bg_rebuild_forecast_cache,    daemon=True).start()
        logger.info("  [PerfOpt] バックグラウンドで予測キャッシュ・集計テーブルをウォームアップ中...")
    except Exception as _we:
        logger.warning(f"  [WARN] Warmup start failed: {_we}")
    logger.info(f"  Starting web server on port {port}...")
    logger.info(f"  Open browser: http://localhost:{port}")
    logger.info("  Press Ctrl+C to stop\n")
    # waitress が使えればwaitress、なければFlask開発サーバー
    if use_waitress:
        try:
            from waitress import serve
            logger.info("  Server: waitress (production)")
            serve(app, host='0.0.0.0', port=port, threads=8,
                  channel_timeout=120, cleanup_interval=30)
        except ImportError:
            logger.info("  waitress not found. Using Flask dev server.")
            logger.info("  To install: pip install waitress")
            app.run(host='0.0.0.0', port=port, debug=False,
                    use_reloader=False, threaded=True)
    else:
        app.run(host='0.0.0.0', port=port, debug=False,
                use_reloader=False, threaded=True)
