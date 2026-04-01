"""チェーン・店舗管理 Blueprint"""
from flask import Blueprint, render_template, request, redirect, url_for, flash, send_file
import logging
from db import get_db
from auth_helpers import permission_required, admin_required

logger = logging.getLogger('inventory.chains')
bp = Blueprint('chains', __name__)


@bp.route('/chains')
@permission_required('chains')
def chains():
    db = get_db()
    chains = db.execute("SELECT * FROM chain_masters ORDER BY chain_cd").fetchall()
    stores = db.execute("SELECT * FROM store_masters ORDER BY store_cd").fetchall()
    supplier_settings = db.execute(
        "SELECT * FROM supplier_cd_settings ORDER BY chain_cd NULLS LAST, store_cd NULLS LAST, supplier_cd"
    ).fetchall()
    product_settings = db.execute(
        "SELECT * FROM product_cd_settings ORDER BY chain_cd NULLS LAST, store_cd NULLS LAST, product_cd NULLS LAST, jan NULLS LAST"
    ).fetchall()
    chain_list = [c['chain_cd'] for c in chains]
    store_list = [s['store_cd'] for s in stores]
    return render_template('chains.html', chains=chains, stores=stores,
                           supplier_settings=supplier_settings,
                           product_settings=product_settings,
                           chain_list=chain_list, store_list=store_list)

@bp.route('/chains/chain/add', methods=['POST'])
@admin_required
def chain_add():
    db = get_db()
    chain_cd       = request.form.get('chain_cd', '').strip()
    chain_name     = request.form.get('chain_name', '').strip()
    exclude_deduct = 1 if request.form.get('exclude_deduct') == '1' else 0
    if not chain_cd:
        flash('チェーンCDを入力してください', 'error')
        return redirect('/chains')
    db.execute(
        "INSERT INTO chain_masters (chain_cd,chain_name,exclude_deduct) VALUES (%s,%s,%s) ON CONFLICT (chain_cd) DO UPDATE SET chain_name=%s, exclude_deduct=%s",
        [chain_cd, chain_name, exclude_deduct, chain_name, exclude_deduct]
    )
    db.commit()
    flash(f'チェーンCD {chain_cd} を追加しました', 'success')
    return redirect('/chains')

@bp.route('/chains/chain/<int:chain_id>/update', methods=['POST'])
@admin_required
def chain_update(chain_id):
    db = get_db()
    chain_name     = request.form.get('chain_name', '').strip()
    exclude_deduct = 1 if request.form.get('exclude_deduct') == '1' else 0
    db.execute(
        "UPDATE chain_masters SET chain_name=%s, exclude_deduct=%s WHERE id=%s",
        [chain_name, exclude_deduct, chain_id]
    )
    db.commit()
    flash('更新しました', 'success')
    return redirect('/chains')

@bp.route('/chains/store/add', methods=['POST'])
@admin_required
def store_add():
    db = get_db()
    store_cd       = request.form.get('store_cd', '').strip()
    store_name     = request.form.get('store_name', '').strip()
    chain_cd       = request.form.get('chain_cd', '').strip()
    client_name    = request.form.get('client_name', '').strip()
    exclude_deduct = 1 if request.form.get('exclude_deduct') == '1' else 0
    if not store_cd:
        flash('店舗CDを入力してください', 'error')
        return redirect('/chains#store')
    db.execute(
        "INSERT INTO store_masters (store_cd,store_name,chain_cd,client_name,exclude_deduct) VALUES (%s,%s,%s,%s,%s) ON CONFLICT (store_cd) DO UPDATE SET store_name=%s, chain_cd=%s, client_name=%s, exclude_deduct=%s",
        [store_cd, store_name, chain_cd, client_name, exclude_deduct,
         store_name, chain_cd, client_name, exclude_deduct]
    )
    db.commit()
    flash(f'店舗CD {store_cd} を追加しました', 'success')
    return redirect('/chains')

@bp.route('/chains/store/<int:store_id>/update', methods=['POST'])
@admin_required
def store_update(store_id):
    db = get_db()
    store_name     = request.form.get('store_name', '').strip()
    client_name    = request.form.get('client_name', '').strip()
    exclude_deduct = 1 if request.form.get('exclude_deduct') == '1' else 0
    db.execute(
        "UPDATE store_masters SET store_name=%s, client_name=%s, exclude_deduct=%s WHERE id=%s",
        [store_name, client_name, exclude_deduct, store_id]
    )
    db.commit()
    flash('更新しました', 'success')
    return redirect('/chains')


# ── 仕入先CD設定 ─────────────────────────────────────────────────────

@bp.route('/chains/supplier-setting/add', methods=['POST'])
@admin_required
def supplier_setting_add():
    db = get_db()
    supplier_cd    = request.form.get('supplier_cd', '').strip()
    chain_cd       = request.form.get('chain_cd', '').strip() or None
    store_cd       = request.form.get('store_cd', '').strip() or None
    exclude_deduct = 1 if request.form.get('exclude_deduct') == '1' else 0
    notes          = request.form.get('notes', '').strip()
    if not supplier_cd:
        flash('仕入先CDを入力してください', 'error')
        return redirect('/chains#supplier')
    db.execute(
        """INSERT INTO supplier_cd_settings (chain_cd, store_cd, supplier_cd, exclude_deduct, notes)
           VALUES (%s, %s, %s, %s, %s)
           ON CONFLICT (chain_cd, store_cd, supplier_cd)
           DO UPDATE SET exclude_deduct=%s, notes=%s""",
        [chain_cd, store_cd, supplier_cd, exclude_deduct, notes,
         exclude_deduct, notes]
    )
    db.commit()
    flash(f'仕入先CD {supplier_cd} の設定を追加しました', 'success')
    return redirect('/chains#supplier')

@bp.route('/chains/supplier-setting/<int:setting_id>/update', methods=['POST'])
@admin_required
def supplier_setting_update(setting_id):
    db = get_db()
    exclude_deduct = 1 if request.form.get('exclude_deduct') == '1' else 0
    notes          = request.form.get('notes', '').strip()
    db.execute(
        "UPDATE supplier_cd_settings SET exclude_deduct=%s, notes=%s WHERE id=%s",
        [exclude_deduct, notes, setting_id]
    )
    db.commit()
    flash('仕入先CD設定を更新しました', 'success')
    return redirect('/chains#supplier')

@bp.route('/chains/supplier-setting/<int:setting_id>/delete', methods=['POST'])
@admin_required
def supplier_setting_delete(setting_id):
    db = get_db()
    db.execute("DELETE FROM supplier_cd_settings WHERE id=%s", [setting_id])
    db.commit()
    flash('削除しました', 'success')
    return redirect('/chains#supplier')


# ── 商品CD設定 ───────────────────────────────────────────────────────

@bp.route('/chains/product-setting/add', methods=['POST'])
@admin_required
def product_setting_add():
    db = get_db()
    product_cd     = request.form.get('product_cd', '').strip() or None
    jan            = request.form.get('jan', '').strip() or None
    chain_cd       = request.form.get('chain_cd', '').strip() or None
    store_cd       = request.form.get('store_cd', '').strip() or None
    exclude_deduct = 1 if request.form.get('exclude_deduct') == '1' else 0
    notes          = request.form.get('notes', '').strip()
    if not product_cd and not jan:
        flash('商品CDまたはJANコードを入力してください', 'error')
        return redirect('/chains#product')
    db.execute(
        """INSERT INTO product_cd_settings (chain_cd, store_cd, product_cd, jan, exclude_deduct, notes)
           VALUES (%s, %s, %s, %s, %s, %s)""",
        [chain_cd, store_cd, product_cd, jan, exclude_deduct, notes]
    )
    db.commit()
    flash(f'商品CD {product_cd or jan} の設定を追加しました', 'success')
    return redirect('/chains#product')

@bp.route('/chains/product-setting/<int:setting_id>/update', methods=['POST'])
@admin_required
def product_setting_update(setting_id):
    db = get_db()
    exclude_deduct = 1 if request.form.get('exclude_deduct') == '1' else 0
    notes          = request.form.get('notes', '').strip()
    db.execute(
        "UPDATE product_cd_settings SET exclude_deduct=%s, notes=%s WHERE id=%s",
        [exclude_deduct, notes, setting_id]
    )
    db.commit()
    flash('商品CD設定を更新しました', 'success')
    return redirect('/chains#product')

@bp.route('/chains/product-setting/<int:setting_id>/delete', methods=['POST'])
@admin_required
def product_setting_delete(setting_id):
    db = get_db()
    db.execute("DELETE FROM product_cd_settings WHERE id=%s", [setting_id])
    db.commit()
    flash('削除しました', 'success')
    return redirect('/chains#product')


# ── チェーンマスタ Excel テンプレートDL ───────────────────────────────────

@bp.route('/chains/chain/template')
@admin_required
def chain_template():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from io import BytesIO
    from urllib.parse import quote
    db = get_db()
    _tn = db.execute("SELECT value FROM settings WHERE key='chain_template_name'").fetchone()
    dl_name = ((_tn['value'] if _tn else None) or 'チェーンマスタ_テンプレート') + '.xlsx'

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'チェーンマスタ'
    headers = ['チェーンCD', '取引先名', '在庫引き当て除外(0/1)']
    notes   = ['必須・一意', '任意', '1=除外 0=引き当て対象']
    hfill = PatternFill('solid', fgColor='1d4ed8')
    nfill = PatternFill('solid', fgColor='eff6ff')
    for col, (h, n) in enumerate(zip(headers, notes), 1):
        hc = ws.cell(row=1, column=col, value=h)
        hc.font = Font(bold=True, color='FFFFFF')
        hc.fill = hfill
        hc.alignment = Alignment(horizontal='center')
        nc = ws.cell(row=2, column=col, value=n)
        nc.fill = nfill
        nc.font = Font(italic=True, color='6b7280')
    ws.append(['CHAIN001', '山田食品', 0])
    ws.append(['CHAIN002', '佐藤物産', 1])
    for i, w in enumerate([18, 28, 22], 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name=dl_name,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ── チェーンマスタ Excel 一括インポート ──────────────────────────────────

@bp.route('/chains/chain/import', methods=['POST'])
@admin_required
def chain_import():
    import openpyxl
    from io import BytesIO
    f = request.files.get('excel_file')
    if not f or not f.filename.lower().endswith(('.xlsx', '.xls')):
        flash('Excelファイル（.xlsx）を選択してください。', 'danger')
        return redirect('/chains')
    db = get_db()
    try:
        wb = openpyxl.load_workbook(BytesIO(f.read()), data_only=True)
        ws = wb.active
    except Exception as e:
        flash(f'ファイルの読み込みに失敗しました: {e}', 'danger')
        return redirect('/chains')

    added = updated = skipped = 0
    for row in ws.iter_rows(min_row=3, values_only=True):
        chain_cd = str(row[0]).strip() if row[0] is not None else ''
        if not chain_cd or chain_cd == 'None':
            skipped += 1
            continue
        chain_name     = str(row[1]).strip() if row[1] is not None else ''
        try:
            exclude_deduct = int(row[2]) if row[2] is not None else 0
        except (ValueError, TypeError):
            exclude_deduct = 0
        exclude_deduct = 1 if exclude_deduct else 0
        existing = db.execute("SELECT id FROM chain_masters WHERE chain_cd=%s", [chain_cd]).fetchone()
        if existing:
            db.execute(
                "UPDATE chain_masters SET chain_name=%s, exclude_deduct=%s WHERE chain_cd=%s",
                [chain_name, exclude_deduct, chain_cd]
            )
            updated += 1
        else:
            db.execute(
                "INSERT INTO chain_masters (chain_cd, chain_name, exclude_deduct) VALUES (%s,%s,%s)",
                [chain_cd, chain_name, exclude_deduct]
            )
            added += 1
    db.commit()
    flash(f'チェーンマスタをインポートしました。追加:{added}件 更新:{updated}件 スキップ:{skipped}件', 'success')
    return redirect('/chains')


# ── 店舗マスタ Excel テンプレートDL ────────────────────────────────────

@bp.route('/chains/store/template')
@admin_required
def store_template():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from io import BytesIO
    from urllib.parse import quote
    db = get_db()
    _tn = db.execute("SELECT value FROM settings WHERE key='store_template_name'").fetchone()
    dl_name = ((_tn['value'] if _tn else None) or '店舗マスタ_テンプレート') + '.xlsx'

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '店舗マスタ'
    headers = ['店舗CD', '店舗名', 'チェーンCD', '取引先名', '在庫引き当て除外(0/1)']
    notes   = ['必須・一意', '任意', '任意', '任意', '1=除外 0=引き当て対象']
    hfill = PatternFill('solid', fgColor='1d4ed8')
    nfill = PatternFill('solid', fgColor='eff6ff')
    for col, (h, n) in enumerate(zip(headers, notes), 1):
        hc = ws.cell(row=1, column=col, value=h)
        hc.font = Font(bold=True, color='FFFFFF')
        hc.fill = hfill
        hc.alignment = Alignment(horizontal='center')
        nc = ws.cell(row=2, column=col, value=n)
        nc.fill = nfill
        nc.font = Font(italic=True, color='6b7280')
    ws.append(['STORE001', '東京本店', 'CHAIN001', '山田食品', 0])
    ws.append(['STORE002', '大阪支店', 'CHAIN001', '山田食品', 0])
    for i, w in enumerate([14, 24, 14, 24, 22], 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name=dl_name,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ── 店舗マスタ Excel 一括インポート ────────────────────────────────────

@bp.route('/chains/store/import', methods=['POST'])
@admin_required
def store_import():
    import openpyxl
    from io import BytesIO
    f = request.files.get('excel_file')
    if not f or not f.filename.lower().endswith(('.xlsx', '.xls')):
        flash('Excelファイル（.xlsx）を選択してください。', 'danger')
        return redirect('/chains')
    db = get_db()
    try:
        wb = openpyxl.load_workbook(BytesIO(f.read()), data_only=True)
        ws = wb.active
    except Exception as e:
        flash(f'ファイルの読み込みに失敗しました: {e}', 'danger')
        return redirect('/chains')

    added = updated = skipped = 0
    for row in ws.iter_rows(min_row=3, values_only=True):
        store_cd = str(row[0]).strip() if row[0] is not None else ''
        if not store_cd or store_cd == 'None':
            skipped += 1
            continue
        store_name     = str(row[1]).strip() if row[1] is not None else ''
        chain_cd       = str(row[2]).strip() if row[2] is not None else ''
        client_name    = str(row[3]).strip() if row[3] is not None else ''
        try:
            exclude_deduct = int(row[4]) if row[4] is not None else 0
        except (ValueError, TypeError):
            exclude_deduct = 0
        exclude_deduct = 1 if exclude_deduct else 0
        existing = db.execute("SELECT id FROM store_masters WHERE store_cd=%s", [store_cd]).fetchone()
        if existing:
            db.execute(
                "UPDATE store_masters SET store_name=%s, chain_cd=%s, client_name=%s, exclude_deduct=%s WHERE store_cd=%s",
                [store_name, chain_cd or None, client_name, exclude_deduct, store_cd]
            )
            updated += 1
        else:
            db.execute(
                "INSERT INTO store_masters (store_cd, store_name, chain_cd, client_name, exclude_deduct) VALUES (%s,%s,%s,%s,%s)",
                [store_cd, store_name, chain_cd or None, client_name, exclude_deduct]
            )
            added += 1
    db.commit()
    flash(f'店舗マスタをインポートしました。追加:{added}件 更新:{updated}件 スキップ:{skipped}件', 'success')
    return redirect('/chains')


# ── 仕入先CD設定 Excel テンプレートDL ─────────────────────────────────

@bp.route('/chains/supplier/template')
@admin_required
def supplier_setting_template():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from io import BytesIO
    db = get_db()
    _tn = db.execute("SELECT value FROM settings WHERE key='supplier_setting_template_name'").fetchone()
    dl_name = ((_tn['value'] if _tn else None) or '仕入先CD設定_テンプレート') + '.xlsx'

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '仕入先CD設定'
    headers = ['仕入先CD', 'チェーンCD', '店舗CD', '引き当て除外(0/1)', '備考']
    notes   = ['必須', '任意（空白=全チェーン）', '任意（空白=全店舗）', '1=除外 0=対象', '任意']
    hfill = PatternFill('solid', fgColor='7c3aed')
    nfill = PatternFill('solid', fgColor='f5f3ff')
    for col, (h, n) in enumerate(zip(headers, notes), 1):
        hc = ws.cell(row=1, column=col, value=h)
        hc.font = Font(bold=True, color='FFFFFF')
        hc.fill = hfill
        hc.alignment = Alignment(horizontal='center')
        nc = ws.cell(row=2, column=col, value=n)
        nc.fill = nfill
        nc.font = Font(italic=True, color='6b7280')
    ws.append(['S001', 'CHAIN001', '', 0, ''])
    ws.append(['S002', '', '', 1, '全チェーン除外'])
    for i, w in enumerate([16, 16, 14, 22, 30], 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name=dl_name,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ── 仕入先CD設定 Excel 一括インポート ───────────────────────────────────

@bp.route('/chains/supplier/import', methods=['POST'])
@admin_required
def supplier_setting_import():
    import openpyxl
    from io import BytesIO
    f = request.files.get('excel_file')
    if not f or not f.filename.lower().endswith(('.xlsx', '.xls')):
        flash('Excelファイル（.xlsx）を選択してください。', 'danger')
        return redirect('/chains#supplier')
    db = get_db()
    try:
        wb = openpyxl.load_workbook(BytesIO(f.read()), data_only=True)
        ws = wb.active
    except Exception as e:
        flash(f'ファイルの読み込みに失敗しました: {e}', 'danger')
        return redirect('/chains#supplier')

    added = updated = skipped = 0
    for row in ws.iter_rows(min_row=3, values_only=True):
        supplier_cd = str(row[0]).strip() if row[0] is not None else ''
        if not supplier_cd or supplier_cd == 'None':
            skipped += 1
            continue
        chain_cd  = str(row[1]).strip() if row[1] is not None else ''
        store_cd  = str(row[2]).strip() if row[2] is not None else ''
        chain_cd  = chain_cd or None
        store_cd  = store_cd or None
        try:
            exclude_deduct = 1 if int(row[3]) else 0
        except (ValueError, TypeError):
            exclude_deduct = 0
        notes = str(row[4]).strip() if row[4] is not None else ''

        # NULL安全な存在確認
        existing = db.execute(
            """SELECT id FROM supplier_cd_settings
               WHERE supplier_cd=%s
                 AND COALESCE(chain_cd,'') = COALESCE(%s,'')
                 AND COALESCE(store_cd,'') = COALESCE(%s,'')""",
            [supplier_cd, chain_cd or '', store_cd or '']
        ).fetchone()
        if existing:
            db.execute(
                "UPDATE supplier_cd_settings SET exclude_deduct=%s, notes=%s WHERE id=%s",
                [exclude_deduct, notes, existing['id']]
            )
            updated += 1
        else:
            db.execute(
                """INSERT INTO supplier_cd_settings (supplier_cd, chain_cd, store_cd, exclude_deduct, notes)
                   VALUES (%s,%s,%s,%s,%s)""",
                [supplier_cd, chain_cd, store_cd, exclude_deduct, notes]
            )
            added += 1
    db.commit()
    flash(f'仕入先CD設定をインポートしました。追加:{added}件 更新:{updated}件 スキップ:{skipped}件', 'success')
    return redirect('/chains#supplier')


# ── 商品CD設定 Excel テンプレートDL ────────────────────────────────────

@bp.route('/chains/product/template')
@admin_required
def product_setting_template():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from io import BytesIO
    db = get_db()
    _tn = db.execute("SELECT value FROM settings WHERE key='product_setting_template_name'").fetchone()
    dl_name = ((_tn['value'] if _tn else None) or '商品CD設定_テンプレート') + '.xlsx'

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '商品CD設定'
    headers = ['商品CD', 'JANコード', 'チェーンCD', '店舗CD', '引き当て除外(0/1)', '備考']
    notes   = ['商品CDまたはJANを入力', '商品CDまたはJANを入力', '任意（空白=全チェーン）', '任意（空白=全店舗）', '1=除外 0=対象', '任意']
    hfill = PatternFill('solid', fgColor='0f766e')
    nfill = PatternFill('solid', fgColor='f0fdfa')
    for col, (h, n) in enumerate(zip(headers, notes), 1):
        hc = ws.cell(row=1, column=col, value=h)
        hc.font = Font(bold=True, color='FFFFFF')
        hc.fill = hfill
        hc.alignment = Alignment(horizontal='center')
        nc = ws.cell(row=2, column=col, value=n)
        nc.fill = nfill
        nc.font = Font(italic=True, color='6b7280')
    ws.append(['P001', '', 'CHAIN001', '', 0, ''])
    ws.append(['', '4901234567890', '', '', 1, '全除外'])
    for i, w in enumerate([14, 18, 14, 14, 22, 28], 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name=dl_name,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ── 商品CD設定 Excel 一括インポート ─────────────────────────────────────

@bp.route('/chains/product/import', methods=['POST'])
@admin_required
def product_setting_import():
    import openpyxl
    from io import BytesIO
    f = request.files.get('excel_file')
    if not f or not f.filename.lower().endswith(('.xlsx', '.xls')):
        flash('Excelファイル（.xlsx）を選択してください。', 'danger')
        return redirect('/chains#product')
    db = get_db()
    try:
        wb = openpyxl.load_workbook(BytesIO(f.read()), data_only=True)
        ws = wb.active
    except Exception as e:
        flash(f'ファイルの読み込みに失敗しました: {e}', 'danger')
        return redirect('/chains#product')

    added = updated = skipped = 0
    for row in ws.iter_rows(min_row=3, values_only=True):
        product_cd = str(row[0]).strip() if row[0] is not None else ''
        jan        = str(row[1]).strip() if row[1] is not None else ''
        product_cd = product_cd if product_cd and product_cd != 'None' else None
        jan        = jan if jan and jan != 'None' else None
        if not product_cd and not jan:
            skipped += 1
            continue
        chain_cd = str(row[2]).strip() if row[2] is not None else ''
        store_cd = str(row[3]).strip() if row[3] is not None else ''
        chain_cd = chain_cd or None
        store_cd = store_cd or None
        try:
            exclude_deduct = 1 if int(row[4]) else 0
        except (ValueError, TypeError):
            exclude_deduct = 0
        notes = str(row[5]).strip() if row[5] is not None else ''

        # 同じ組み合わせが既存なら更新、なければ追加
        existing = db.execute(
            """SELECT id FROM product_cd_settings
               WHERE COALESCE(product_cd,'') = COALESCE(%s,'')
                 AND COALESCE(jan,'') = COALESCE(%s,'')
                 AND COALESCE(chain_cd,'') = COALESCE(%s,'')
                 AND COALESCE(store_cd,'') = COALESCE(%s,'')""",
            [product_cd or '', jan or '', chain_cd or '', store_cd or '']
        ).fetchone()
        if existing:
            db.execute(
                "UPDATE product_cd_settings SET exclude_deduct=%s, notes=%s WHERE id=%s",
                [exclude_deduct, notes, existing['id']]
            )
            updated += 1
        else:
            db.execute(
                """INSERT INTO product_cd_settings (product_cd, jan, chain_cd, store_cd, exclude_deduct, notes)
                   VALUES (%s,%s,%s,%s,%s,%s)""",
                [product_cd, jan, chain_cd, store_cd, exclude_deduct, notes]
            )
            added += 1
    db.commit()
    flash(f'商品CD設定をインポートしました。追加:{added}件 更新:{updated}件 スキップ:{skipped}件', 'success')
    return redirect('/chains#product')
