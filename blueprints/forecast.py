"""問屋向け需要予測Blueprint: 予測一覧 / 52週MD / ABC分析 / 気温データ管理"""
import io, csv, logging
from datetime import date
from flask import Blueprint, render_template, request, redirect, url_for, flash, Response, jsonify
from db import get_db
from auth_helpers import permission_required
from wholesale_forecast import (
    build_wholesale_forecast_rows,
    generate_weekly_md_plan,
    recalc_temp_sensitivity,
)

logger = logging.getLogger('inventory.forecast')
bp = Blueprint('forecast', __name__)


def _save_sens_log(db, count: int, trigger: str):
    """気温感応度再計算の実行履歴をsettingsテーブルに保存"""
    from datetime import datetime as _dt
    labels = {'monthly': '月次自動', 'import': 'インポート後自動', 'manual': '手動実行'}
    val = f"{_dt.now().strftime('%Y-%m-%d %H:%M')} ／ {count}商品 ／ {labels.get(trigger, trigger)}"
    db.execute("""
        INSERT INTO settings (key, value) VALUES ('temp_sensitivity_last_run', %s)
        ON CONFLICT (key) DO UPDATE SET value = EXCLUDED.value
    """, [val])
    db.commit()


# ── 需要予測メイン ────────────────────────────────────────────────────────

@bp.route('/reports/forecast/wholesale')
@permission_required('reports')
def forecast_wholesale():
    db = get_db()
    q = request.args.get('q', '').strip().lower()
    rows, ai_mode = build_wholesale_forecast_rows(db, q)

    # ABC別件数集計
    abc_counts = {'A': 0, 'B': 0, 'C': 0}
    for r in rows:
        abc_counts[r.get('abc_rank', 'C')] += 1

    # 設定値取得
    def _setting(key, default):
        row = db.execute("SELECT value FROM settings WHERE key=%s", [key]).fetchone()
        return row['value'] if row else default

    z_score = float(_setting('safety_level_z', '1.65'))

    # ── 欠品日数（30日中ゼロ補完後のゼロ日数）──────────────────────────────
    zero_days_map = {}
    try:
        zd_rows = db.execute("""
            SELECT jan, 30 - COUNT(DISTINCT sale_dt) AS zero_days
            FROM sales_daily_agg
            WHERE sale_dt >= CURRENT_DATE - INTERVAL '29 days'
            GROUP BY jan
        """).fetchall()
        zero_days_map = {r['jan']: int(r['zero_days']) for r in zd_rows}
    except Exception:
        pass

    # ── IQR除外日数（過去30日で上限超えの日数）────────────────────────────
    iqr_excluded_map = {}
    try:
        iqr_rows = db.execute("""
            WITH bounds AS (
                SELECT jan,
                       PERCENTILE_CONT(0.75) WITHIN GROUP (ORDER BY qty)
                         + 1.5 * (PERCENTILE_CONT(0.75) WITHIN GROUP (ORDER BY qty)
                                - PERCENTILE_CONT(0.25) WITHIN GROUP (ORDER BY qty))
                         AS upper_limit
                FROM sales_daily_agg
                WHERE sale_dt >= CURRENT_DATE - INTERVAL '84 days'
                GROUP BY jan
            )
            SELECT d.jan, COUNT(*) AS excl
            FROM sales_daily_agg d
            JOIN bounds b ON b.jan = d.jan
            WHERE d.sale_dt >= CURRENT_DATE - INTERVAL '29 days'
              AND d.qty > b.upper_limit
            GROUP BY d.jan
        """).fetchall()
        iqr_excluded_map = {r['jan']: int(r['excl']) for r in iqr_rows}
    except Exception:
        pass

    # ── 予測精度MAPE（最新の計算値）────────────────────────────────────────
    mape_map = {}
    try:
        mape_rows = db.execute("""
            SELECT DISTINCT ON (jan) jan, mape, calc_date
            FROM forecast_accuracy
            ORDER BY jan, calc_date DESC
        """).fetchall()
        mape_map = {r['jan']: {'mape': float(r['mape']), 'dt': str(r['calc_date'])}
                    for r in mape_rows}
    except Exception:
        pass

    return render_template('forecast_wholesale.html',
                           rows=rows, q=q,
                           abc_counts=abc_counts,
                           z_score=z_score,
                           ai_mode=ai_mode,
                           zero_days_map=zero_days_map,
                           iqr_excluded_map=iqr_excluded_map,
                           mape_map=mape_map)


@bp.route('/reports/forecast/wholesale/apply', methods=['POST'])
@permission_required('reports')
def forecast_wholesale_apply():
    db = get_db()
    mode = request.form.get('mode', 'reorder_point')
    q    = request.form.get('q', '')
    selected_ids = request.form.getlist('selected_ids')
    selected_set = set(str(i) for i in selected_ids) if selected_ids else None
    rows, _ = build_wholesale_forecast_rows(db, q)
    updated = 0
    for r in rows:
        if selected_set and str(r['product_id']) not in selected_set:
            continue
        locked = int(r.get('lock_order_qty') or 0)
        if mode == 'both':
            if locked:
                db.execute("UPDATE products SET reorder_point=%s WHERE id=%s",
                           [r['suggested_reorder_point'], r['product_id']])
            else:
                db.execute("UPDATE products SET reorder_point=%s, order_qty=%s WHERE id=%s",
                           [r['suggested_reorder_point'], r['suggested_order_qty'], r['product_id']])
        elif mode == 'order_qty':
            if not locked:
                db.execute("UPDATE products SET order_qty=%s WHERE id=%s",
                           [r['suggested_order_qty'], r['product_id']])
        else:
            db.execute("UPDATE products SET reorder_point=%s WHERE id=%s",
                       [r['suggested_reorder_point'], r['product_id']])
        updated += 1
    db.commit()
    scope = '選択' if selected_set else '全件'
    flash(f'問屋向け予測をもとに {updated} 件（{scope}）の設定を更新しました。', 'success')
    return redirect(url_for('forecast.forecast_wholesale', q=q))


@bp.route('/reports/forecast/wholesale/export')
@permission_required('reports')
def forecast_wholesale_export():
    db  = get_db()
    q   = request.args.get('q', '').strip()
    rows, _ = build_wholesale_forecast_rows(db, q)
    sio  = io.StringIO()
    writer = csv.writer(sio)
    writer.writerow(['仕入先CD', '仕入先名', '商品CD', 'JAN', '商品名',
                     'ABCランク', 'アルゴリズム',
                     '現在庫', '消化日数',
                     '30日予測', '日次予測(Q50)',
                     'Q25日次', 'Q50日次', 'Q75日次',
                     'IQR標準偏差', '動的安全在庫',
                     '気温補正係数',
                     '推奨発注点', '推奨発注数',
                     '52週MD計画数', '達成率(%)'])
    for r in rows:
        writer.writerow([
            r['supplier_cd'], r['supplier_name'], r['product_cd'], r['jan'], r['product_name'],
            r['abc_rank'], r['algorithm'],
            r['stock_qty'], r.get('cover_days', ''),
            r['forecast_30d'], r['next_daily_forecast'],
            r['q25_daily'], r['q50_daily'], r['q75_daily'],
            r['iqr_std'], r['dynamic_safety_stock'],
            r['temp_adj_factor'],
            r['suggested_reorder_point'], r['suggested_order_qty'],
            r.get('md_plan_qty', ''), r.get('md_achievement', ''),
        ])
    data = sio.getvalue().encode('utf-8-sig')
    return Response(data, mimetype='text/csv; charset=utf-8',
                    headers={'Content-Disposition': f'attachment; filename=wholesale_forecast_{date.today()}.csv'})


# ── ABC分析 ───────────────────────────────────────────────────────────────

@bp.route('/reports/abc')
@permission_required('reports')
def abc_analysis():
    db = get_db()

    def _setting(key, default):
        row = db.execute("SELECT value FROM settings WHERE key=%s", [key]).fetchone()
        return row['value'] if row else default

    a_threshold = float(_setting('abc_a_threshold', '0.70'))
    b_threshold = float(_setting('abc_b_threshold', '0.90'))

    rows = db.execute("""
        WITH sales AS (
            SELECT a.jan,
                   p.product_cd, p.product_name, p.supplier_cd, p.supplier_name,
                   SUM(a.qty)::int AS total_qty,
                   SUM(a.qty * COALESCE(NULLIF(p.cost_price,0), 1)) AS sales_value
            FROM sales_daily_agg a
            JOIN products p ON p.jan = a.jan
            WHERE a.sale_dt >= CURRENT_DATE - INTERVAL '365 days'
            GROUP BY a.jan, p.product_cd, p.product_name, p.supplier_cd, p.supplier_name
        ), ranked AS (
            SELECT *,
                   SUM(sales_value) OVER () AS total_sales,
                   SUM(sales_value) OVER (ORDER BY sales_value DESC NULLS LAST, jan) AS running_sales
            FROM sales
        )
        SELECT *,
               CASE WHEN COALESCE(total_sales,0)=0 THEN 1.0
                    ELSE running_sales / NULLIF(total_sales,0) END AS running_ratio
        FROM ranked
        ORDER BY sales_value DESC
    """).fetchall()

    # XYZ分析用: 週次CV（変動係数）計算
    import statistics as _st
    xyz_rows = db.execute("""
        SELECT jan,
               EXTRACT(WEEK FROM sale_dt)::int AS wk,
               SUM(qty) AS wqty
        FROM sales_daily_agg
        WHERE sale_dt >= CURRENT_DATE - INTERVAL '365 days'
        GROUP BY jan, wk
    """).fetchall()
    xyz_map: dict = {}  # jan -> 'X'/'Y'/'Z'
    jan_weeks: dict = {}
    for r in xyz_rows:
        jan_weeks.setdefault(r['jan'], []).append(float(r['wqty'] or 0))
    for jan, wqtys in jan_weeks.items():
        if len(wqtys) < 4:
            xyz_map[jan] = 'Z'
            continue
        mu = _st.mean(wqtys)
        if mu <= 0:
            xyz_map[jan] = 'Z'
            continue
        cv = _st.stdev(wqtys) / mu
        if cv < 0.25:
            xyz_map[jan] = 'X'
        elif cv < 0.75:
            xyz_map[jan] = 'Y'
        else:
            xyz_map[jan] = 'Z'

    abc_rows = []
    for r in rows:
        ratio = float(r['running_ratio'] or 1.0)
        if ratio <= a_threshold:
            rank = 'A'
        elif ratio <= b_threshold:
            rank = 'B'
        else:
            rank = 'C'
        d = dict(r)
        d['abc_rank']          = rank
        d['xyz_rank']          = xyz_map.get(r['jan'], 'Z')
        d['running_ratio_pct'] = round(ratio * 100, 1)
        d['sales_value']       = round(float(r['sales_value'] or 0), 0)
        abc_rows.append(d)

    # ABC×XYZ 9マトリックス集計
    matrix = {a+x: 0 for a in 'ABC' for x in 'XYZ'}
    for r in abc_rows:
        matrix[r['abc_rank'] + r['xyz_rank']] += 1

    # 集計
    summary = {
        'A': {'count': 0, 'qty': 0, 'value': 0},
        'B': {'count': 0, 'qty': 0, 'value': 0},
        'C': {'count': 0, 'qty': 0, 'value': 0},
    }
    for r in abc_rows:
        rk = r['abc_rank']
        summary[rk]['count'] += 1
        summary[rk]['qty']   += int(r['total_qty'] or 0)
        summary[rk]['value'] += float(r['sales_value'] or 0)

    # ABCランクを products テーブルに保存（一括取得→差分のみ更新）
    existing_ranks = {}
    try:
        er = db.execute("SELECT jan, abc_rank FROM products WHERE is_active = 1").fetchall()
        existing_ranks = {r['jan']: r['abc_rank'] for r in er}
    except Exception:
        pass

    changed = 0
    today_str = str(date.today())
    for r in abc_rows:
        jan, new_rank = r['jan'], r['abc_rank']
        cur_rank = existing_ranks.get(jan)
        if cur_rank is not None and cur_rank != new_rank:
            old_rank = cur_rank or 'C'
            db.execute("""
                UPDATE products
                SET abc_rank_prev = %s, abc_rank = %s, abc_rank_updated = %s
                WHERE jan = %s
            """, [old_rank, new_rank, today_str, jan])
            db.execute("""
                INSERT INTO alert_logs (alert_type, jan, product_name, message, mail_sent)
                VALUES ('ABCランク変化', %s, %s, %s, 0)
            """, [jan, r['product_name'],
                  f"ABCランク変化: {old_rank} → {new_rank} ({today_str})"])
            changed += 1
        elif cur_rank is None:
            db.execute("""
                UPDATE products SET abc_rank = %s, abc_rank_updated = %s WHERE jan = %s
            """, [new_rank, today_str, jan])
    if changed:
        db.commit()
        logger.info(f'[ABC] ランク変化: {changed}商品')

    # 前回ランクマップ（ランク変化表示用）
    prev_rank_map = {}
    try:
        prev_rows = db.execute(
            "SELECT jan, abc_rank_prev, abc_rank_updated FROM products WHERE abc_rank_prev IS NOT NULL"
        ).fetchall()
        for pr in prev_rows:
            prev_rank_map[pr['jan']] = {
                'prev': pr['abc_rank_prev'],
                'updated': pr['abc_rank_updated'] or '',
            }
    except Exception:
        pass

    return render_template('abc_analysis.html',
                           rows=abc_rows, summary=summary,
                           a_threshold=a_threshold, b_threshold=b_threshold,
                           prev_rank_map=prev_rank_map,
                           matrix=matrix)


@bp.route('/reports/abc/export')
@permission_required('reports')
def abc_export():
    db = get_db()
    rows = db.execute("""
        WITH sales AS (
            SELECT a.jan, p.product_cd, p.product_name, p.supplier_cd, p.supplier_name,
                   SUM(a.qty)::int AS total_qty,
                   SUM(a.qty * COALESCE(NULLIF(p.cost_price,0), 1)) AS sales_value
            FROM sales_daily_agg a
            JOIN products p ON p.jan = a.jan
            WHERE a.sale_dt >= CURRENT_DATE - INTERVAL '365 days'
            GROUP BY a.jan, p.product_cd, p.product_name, p.supplier_cd, p.supplier_name
        ), ranked AS (
            SELECT *, SUM(sales_value) OVER () AS total,
                   SUM(sales_value) OVER (ORDER BY sales_value DESC NULLS LAST, jan) AS running
            FROM sales
        )
        SELECT *, CASE WHEN COALESCE(total,0)=0 THEN 1.0 ELSE running/NULLIF(total,0) END AS ratio
        FROM ranked ORDER BY sales_value DESC
    """).fetchall()

    def _setting(key, default):
        row = db.execute("SELECT value FROM settings WHERE key=%s", [key]).fetchone()
        return row['value'] if row else default
    a_thr = float(_setting('abc_a_threshold', '0.70'))
    b_thr = float(_setting('abc_b_threshold', '0.90'))

    sio = io.StringIO()
    w   = csv.writer(sio)
    w.writerow(['ABCランク', '仕入先CD', '仕入先名', '商品CD', 'JAN', '商品名', '販売数量', '売上金額', '累計構成比(%)'])
    for r in rows:
        ratio = float(r['ratio'] or 1.0)
        rank  = 'A' if ratio <= a_thr else ('B' if ratio <= b_thr else 'C')
        w.writerow([rank, r['supplier_cd'], r['supplier_name'], r['product_cd'], r['jan'], r['product_name'],
                    r['total_qty'], round(float(r['sales_value'] or 0), 0), round(ratio * 100, 1)])
    return Response(sio.getvalue().encode('utf-8-sig'), mimetype='text/csv; charset=utf-8',
                    headers={'Content-Disposition': f'attachment; filename=abc_analysis_{date.today()}.csv'})


# ── 52週MDプラン ─────────────────────────────────────────────────────────

@bp.route('/reports/weekly_md')
@permission_required('reports')
def weekly_md():
    db = get_db()
    today = date.today()
    # 年度（4月始まり）
    fiscal_year = today.year if today.month >= 4 else today.year - 1
    year_param  = int(request.args.get('year', fiscal_year))
    q           = request.args.get('q', '').strip().lower()

    rows = db.execute("""
        SELECT wm.*, p.product_name, p.product_cd, p.supplier_cd, p.supplier_name
        FROM weekly_md_plans wm
        JOIN products p ON p.jan = wm.jan
        WHERE wm.fiscal_year = %s
        ORDER BY p.supplier_cd, p.product_cd, wm.week_no
    """, [year_param]).fetchall()

    if q:
        rows = [r for r in rows if any(q in str(r.get(k) or '').lower()
                                       for k in ('jan', 'product_cd', 'product_name', 'supplier_cd', 'supplier_name'))]

    # 実績：sales_daily_agg から週番号別に集計（年度内の全週）
    # 年度開始月（4月）〜終了月（翌年3月）の範囲で絞り込む
    fiscal_start = date(year_param, 4, 1)
    fiscal_end   = date(year_param + 1, 3, 31)
    actual_rows = db.execute("""
        SELECT jan,
               dow,
               EXTRACT(WEEK FROM sale_dt)::int AS week_no,
               SUM(qty) AS qty
        FROM sales_daily_agg
        WHERE sale_dt BETWEEN %s AND %s
        GROUP BY jan, week_no, dow
    """, [fiscal_start, fiscal_end]).fetchall()

    # JAN × 週番号 → 実績数量マップ
    actual_map: dict = {}
    for ar in actual_rows:
        jan = ar['jan']
        wn  = int(ar['week_no'])
        actual_map.setdefault(jan, {})
        actual_map[jan][wn] = actual_map[jan].get(wn, 0) + int(ar['qty'] or 0)

    # 商品ごとに週次データを集約
    products_md: dict = {}
    for r in rows:
        jan = r['jan']
        wn  = int(r['week_no'])
        if jan not in products_md:
            products_md[jan] = {
                'jan': jan, 'product_cd': r['product_cd'],
                'product_name': r['product_name'],
                'supplier_cd': r['supplier_cd'],
                'supplier_name': r['supplier_name'],
                'weeks': {}
            }
        products_md[jan]['weeks'][wn] = {
            'plan':   int(r['plan_qty'] or 0),
            'actual': actual_map.get(jan, {}).get(wn, 0),
        }

    # 週次予測値マップ（当週のAI予測 = 直近30日日次平均 × 7）
    forecast_weekly_map: dict = {}
    try:
        fc_rows = db.execute("""
            SELECT jan,
                   ROUND(SUM(qty) / NULLIF(COUNT(DISTINCT sale_dt), 0) * 7, 0) AS weekly_fc
            FROM sales_daily_agg
            WHERE sale_dt >= CURRENT_DATE - INTERVAL '30 days'
            GROUP BY jan
        """).fetchall()
        forecast_weekly_map = {r['jan']: int(r['weekly_fc'] or 0) for r in fc_rows}
    except Exception:
        pass

    # 検索絞り込みに関わらず W1〜W52 を常に全列表示
    weeks     = list(range(1, 53))
    prod_list = sorted(products_md.values(), key=lambda x: (x['supplier_cd'], x['product_cd']))

    return render_template('weekly_md.html',
                           prod_list=prod_list, weeks=weeks,
                           fiscal_year=year_param,
                           current_week=today.isocalendar()[1],
                           forecast_weekly_map=forecast_weekly_map,
                           q=q)


@bp.route('/reports/weekly_md/generate', methods=['POST'])
@permission_required('reports')
def weekly_md_generate():
    db = get_db()
    jan_input = request.form.get('jan', '').strip()
    fiscal_year = int(request.form.get('fiscal_year', date.today().year))

    if not jan_input:
        # 全商品生成
        prods = db.execute("SELECT jan FROM products WHERE is_active=1").fetchall()
        total = 0
        for p in prods:
            total += generate_weekly_md_plan(db, p['jan'], fiscal_year)
        flash(f'{fiscal_year}年度の52週MDプランを全商品({len(prods)}品目)生成しました（{total}週分）。', 'success')
    else:
        n = generate_weekly_md_plan(db, jan_input, fiscal_year)
        flash(f'{jan_input} の{fiscal_year}年度52週MDプランを生成しました（{n}週分）。', 'success')

    return redirect(url_for('forecast.weekly_md', year=fiscal_year))


@bp.route('/reports/weekly_md/update', methods=['POST'])
@permission_required('reports')
def weekly_md_update():
    db  = get_db()
    jan = request.form.get('jan', '').strip()
    fiscal_year = int(request.form.get('fiscal_year', date.today().year))
    week_no     = int(request.form.get('week_no', 1))
    plan_qty    = int(request.form.get('plan_qty', 0) or 0)

    # fiscal_year と week_no から week_start（ISO週の月曜日）を算出
    # 4月始まり年度のため、week_no 1〜13 は翌年、14〜53 は当年の ISO 週に対応
    fiscal_start = date(fiscal_year, 4, 1)
    fiscal_end   = date(fiscal_year + 1, 3, 31)
    week_start_val = None
    for cal_year in [fiscal_year, fiscal_year + 1]:
        try:
            candidate = date.fromisocalendar(cal_year, week_no, 1)
            if fiscal_start <= candidate <= fiscal_end:
                week_start_val = candidate
                break
        except ValueError:
            continue
    if week_start_val is None:
        week_start_val = date.fromisocalendar(
            fiscal_year if week_no >= 14 else fiscal_year + 1, week_no, 1)

    db.execute("""
        INSERT INTO weekly_md_plans (jan, fiscal_year, week_no, week_start, plan_qty, actual_qty, updated_at)
        VALUES (%s, %s, %s, %s, %s, 0, NOW())
        ON CONFLICT (jan, fiscal_year, week_no)
        DO UPDATE SET plan_qty = EXCLUDED.plan_qty, updated_at = NOW()
    """, [jan, fiscal_year, week_no, week_start_val, plan_qty])
    db.commit()
    return jsonify({'ok': True})


@bp.route('/reports/weekly_md/batch_update', methods=['POST'])
@permission_required('reports')
def weekly_md_batch_update():
    db   = get_db()
    data = request.get_json(silent=True) or {}
    items = data.get('items', [])
    if not items:
        return jsonify({'ok': False, 'error': 'no items'})

    fiscal_start_cache: dict = {}
    for item in items:
        jan         = str(item.get('jan', '')).strip()
        fiscal_year = int(item.get('fiscal_year', date.today().year))
        week_no     = int(item.get('week_no', 1))
        plan_qty    = max(0, int(item.get('plan_qty', 0) or 0))

        if fiscal_year not in fiscal_start_cache:
            fiscal_start_cache[fiscal_year] = (
                date(fiscal_year, 4, 1), date(fiscal_year + 1, 3, 31))
        fiscal_start, fiscal_end = fiscal_start_cache[fiscal_year]

        week_start_val = None
        for cal_year in [fiscal_year, fiscal_year + 1]:
            try:
                candidate = date.fromisocalendar(cal_year, week_no, 1)
                if fiscal_start <= candidate <= fiscal_end:
                    week_start_val = candidate
                    break
            except ValueError:
                continue
        if week_start_val is None:
            week_start_val = date.fromisocalendar(
                fiscal_year if week_no >= 14 else fiscal_year + 1, week_no, 1)

        db.execute("""
            INSERT INTO weekly_md_plans
              (jan, fiscal_year, week_no, week_start, plan_qty, actual_qty, updated_at)
            VALUES (%s, %s, %s, %s, %s, 0, NOW())
            ON CONFLICT (jan, fiscal_year, week_no)
            DO UPDATE SET plan_qty = EXCLUDED.plan_qty, updated_at = NOW()
        """, [jan, fiscal_year, week_no, week_start_val, plan_qty])

    db.commit()
    return jsonify({'ok': True, 'saved': len(items)})


@bp.route('/reports/weekly_md/export')
@permission_required('reports')
def weekly_md_export():
    db = get_db()
    fiscal_year = int(request.args.get('year', date.today().year))
    rows = db.execute("""
        SELECT wm.*, p.product_name, p.product_cd, p.supplier_cd, p.supplier_name
        FROM weekly_md_plans wm JOIN products p ON p.jan=wm.jan
        WHERE wm.fiscal_year=%s
        ORDER BY p.supplier_cd, p.product_cd, wm.week_no
    """, [fiscal_year]).fetchall()

    sio = io.StringIO()
    w   = csv.writer(sio)
    w.writerow(['年度', '週番号', '週開始日', '仕入先CD', '仕入先名', '商品CD', 'JAN', '商品名',
                '計画数量', '実績数量', '達成率(%)'])
    for r in rows:
        plan = int(r['plan_qty'] or 0)
        actual = int(r['actual_qty'] or 0)
        ach = round(actual / plan * 100, 1) if plan > 0 else ''
        w.writerow([r['fiscal_year'], r['week_no'], r['week_start'],
                    r['supplier_cd'], r['supplier_name'], r['product_cd'], r['jan'], r['product_name'],
                    plan, actual, ach])
    return Response(sio.getvalue().encode('utf-8-sig'), mimetype='text/csv; charset=utf-8',
                    headers={'Content-Disposition': f'attachment; filename=weekly_md_{fiscal_year}.csv'})


@bp.route('/reports/weekly_md/apply_order', methods=['POST'])
@permission_required('reports')
def weekly_md_apply_order():
    """今後4週のMDプラン計画数からリードタイム分の発注数を自動算出しproductsに反映"""
    import math
    db = get_db()
    fiscal_year = int(request.form.get('fiscal_year', date.today().year))
    today_week  = int(date.today().strftime('%V'))  # ISO週番号

    # 今後4週の計画数を商品別に集計
    plan_rows = db.execute("""
        SELECT wm.jan, SUM(wm.plan_qty) AS plan4w,
               p.lead_time_days, p.order_unit, p.unit_qty, p.safety_factor
        FROM weekly_md_plans wm
        JOIN products p ON p.jan = wm.jan
        WHERE wm.fiscal_year = %s
          AND wm.week_no >= %s AND wm.week_no < %s
          AND wm.plan_qty > 0
          AND p.is_active = 1
        GROUP BY wm.jan, p.lead_time_days, p.order_unit, p.unit_qty, p.safety_factor
    """, [fiscal_year, today_week, today_week + 4]).fetchall()

    updated = 0
    for r in plan_rows:
        plan4w      = int(r['plan4w'] or 0)
        if plan4w <= 0:
            continue
        daily_avg   = plan4w / 28.0
        lt          = int(r['lead_time_days'] or 3)
        sf          = float(r['safety_factor'] or 1.3)
        order_unit  = max(1, int(r['order_unit'] or 1))
        # 発注量 = 日次平均 × (LT + 14日) × 安全係数、order_unit の倍数切り上げ
        raw_oq = daily_avg * (lt + 14) * sf
        new_oq = max(order_unit, math.ceil(raw_oq / order_unit) * order_unit)
        # 発注点 = 日次平均 × LT × 安全係数、order_unit 倍数切り上げ
        raw_rp = daily_avg * lt * sf
        new_rp = max(order_unit, math.ceil(raw_rp / order_unit) * order_unit)

        db.execute("""
            UPDATE products
            SET order_qty = %s, reorder_point = %s
            WHERE jan = %s AND lock_order_qty = 0
        """, [new_oq, new_rp, r['jan']])
        updated += 1

    db.commit()
    flash(f'MDプランから {updated} 商品の発注数・発注点を更新しました（{fiscal_year}年度 W{today_week}〜W{today_week+3}）。', 'success')
    return redirect(url_for('forecast.weekly_md', year=fiscal_year))


# ── アラート一覧 ──────────────────────────────────────────────────────────

@bp.route('/reports/alerts')
@permission_required('reports')
def alerts():
    db = get_db()
    alert_type = request.args.get('type', '').strip()
    page       = max(1, int(request.args.get('page', 1)))
    per_page   = 50

    sql = """
        SELECT al.*, p.product_cd, p.supplier_cd, p.supplier_name
        FROM alert_logs al
        LEFT JOIN products p ON p.jan = al.jan
        WHERE 1=1
    """
    params = []
    if alert_type:
        sql += " AND al.alert_type = %s"
        params.append(alert_type)
    sql += " ORDER BY al.created_at DESC LIMIT %s OFFSET %s"
    params += [per_page, (page - 1) * per_page]

    rows = db.execute(sql, params).fetchall()

    # 件数集計（タイプ別）
    counts = {}
    for row in db.execute("""
        SELECT alert_type, COUNT(*) AS cnt FROM alert_logs
        GROUP BY alert_type ORDER BY cnt DESC
    """).fetchall():
        counts[row['alert_type']] = int(row['cnt'])

    return render_template('alerts.html',
                           rows=rows,
                           counts=counts,
                           alert_type=alert_type,
                           page=page,
                           per_page=per_page)


# ── 気温データ管理 ───────────────────────────────────────────────────────

@bp.route('/reports/weather')
@permission_required('reports')
def weather_data():
    db = get_db()
    rows = db.execute("""
        SELECT * FROM weather_data
        ORDER BY obs_date DESC LIMIT 90
    """).fetchall()
    location_summary = db.execute("""
        SELECT location,
               COUNT(*)          AS cnt,
               MIN(obs_date)     AS oldest,
               MAX(obs_date)     AS newest,
               AVG(avg_temp)     AS avg_temp_avg
        FROM weather_data
        GROUP BY location
        ORDER BY location
    """).fetchall()
    _sens_row = db.execute(
        "SELECT value FROM settings WHERE key='temp_sensitivity_last_run'"
    ).fetchone()
    sens_last_run = _sens_row['value'] if _sens_row else None
    return render_template('weather_data.html', rows=rows,
                           location_summary=location_summary,
                           sens_last_run=sens_last_run)


@bp.route('/reports/weather/add', methods=['POST'])
@permission_required('reports')
def weather_add():
    db       = get_db()
    obs_date = request.form.get('obs_date', '').strip()
    location = request.form.get('location', '東京').strip()
    avg_temp = request.form.get('avg_temp', '')
    max_temp = request.form.get('max_temp', '')
    min_temp = request.form.get('min_temp', '')
    precip   = request.form.get('precipitation', '0')

    if not obs_date or avg_temp == '':
        flash('日付と平均気温は必須です。', 'danger')
        return redirect(url_for('forecast.weather_data'))

    db.execute("""
        INSERT INTO weather_data (obs_date, location, avg_temp, max_temp, min_temp, precipitation, source)
        VALUES (%s, %s, %s, %s, %s, %s, 'manual')
        ON CONFLICT (obs_date, location) DO UPDATE
          SET avg_temp=EXCLUDED.avg_temp, max_temp=EXCLUDED.max_temp,
              min_temp=EXCLUDED.min_temp, precipitation=EXCLUDED.precipitation
    """, [obs_date, location,
          float(avg_temp) if avg_temp else None,
          float(max_temp) if max_temp else None,
          float(min_temp) if min_temp else None,
          float(precip) if precip else 0])
    db.commit()
    flash(f'{obs_date} の気温データを登録しました。', 'success')
    return redirect(url_for('forecast.weather_data'))


@bp.route('/reports/weather/import', methods=['POST'])
@permission_required('reports')
def weather_import():
    db = get_db()
    f  = request.files.get('file')
    if not f or not f.filename:
        flash('ファイルを選択してください。', 'danger')
        return redirect(url_for('forecast.weather_data'))

    content = f.read().decode('utf-8-sig', errors='ignore')
    reader  = csv.DictReader(io.StringIO(content))
    created = 0
    errors  = []
    for i, row in enumerate(reader, 2):
        try:
            obs_date = (row.get('日付') or row.get('obs_date') or '').strip()
            avg_temp = row.get('平均気温(℃)') or row.get('平均気温') or row.get('avg_temp') or ''
            max_temp = row.get('最高気温(℃)') or row.get('最高気温') or row.get('max_temp') or ''
            min_temp = row.get('最低気温(℃)') or row.get('最低気温') or row.get('min_temp') or ''
            location = (row.get('地点') or row.get('location') or '東京').strip()
            precip   = row.get('降水量(mm)') or row.get('降水量') or 0
            if not obs_date or avg_temp == '':
                continue
            db.execute("""
                INSERT INTO weather_data (obs_date, location, avg_temp, max_temp, min_temp, precipitation, source)
                VALUES (%s, %s, %s, %s, %s, %s, 'csv')
                ON CONFLICT (obs_date, location) DO UPDATE
                  SET avg_temp=EXCLUDED.avg_temp, max_temp=EXCLUDED.max_temp,
                      min_temp=EXCLUDED.min_temp, precipitation=EXCLUDED.precipitation
            """, [obs_date, location,
                  float(avg_temp) if avg_temp else None,
                  float(max_temp) if max_temp else None,
                  float(min_temp) if min_temp else None,
                  float(precip) if precip else None])
            created += 1
        except Exception as e:
            errors.append(f'行{i}: {e}')
    db.commit()
    flash(f'{created}件の気温データをインポートしました。' + (f' エラー: {len(errors)}件' if errors else ''), 'success' if not errors else 'warning')
    return redirect(url_for('forecast.weather_data'))


@bp.route('/reports/weather/recalc_sensitivity', methods=['POST'])
@permission_required('reports')
def weather_recalc_sensitivity():
    db = get_db()

    # 事前チェック：原因診断
    weather_cnt = db.execute(
        "SELECT COUNT(*) AS c FROM weather_data WHERE avg_temp IS NOT NULL"
    ).fetchone()['c']
    sales_cnt = db.execute(
        "SELECT COUNT(DISTINCT jan) AS c FROM sales_history "
        "WHERE sale_date::date >= CURRENT_DATE - INTERVAL '365 days'"
    ).fetchone()['c']

    n = recalc_temp_sensitivity(db)

    if n > 0:
        _save_sens_log(db, n, 'manual')
        flash(f'気温感応度を {n} 商品分 再計算しました。', 'success')
    elif weather_cnt < 5:
        flash(
            f'再計算できませんでした。気象データが {weather_cnt} 件しかありません（5件以上必要）。'
            f' まず「気象データ取得」で地点を選択して取得してください。',
            'warning'
        )
    elif sales_cnt == 0:
        flash('再計算できませんでした。過去365日の売上データがありません。', 'warning')
    else:
        flash(
            f'再計算対象なし（0商品）。気象データ {weather_cnt} 件・売上商品 {sales_cnt} 種ありますが、'
            f' 日付が10日以上重複する商品がありません。気象データの期間と売上期間を合わせてください。',
            'warning'
        )
    return redirect(url_for('forecast.weather_data'))


@bp.route('/reports/weather/template')
@permission_required('reports')
def weather_template():
    import datetime
    sio = io.StringIO()
    w   = csv.writer(sio)
    w.writerow(['日付', '地点', '平均気温(℃)', '最高気温(℃)', '最低気温(℃)', '降水量(mm)'])
    today = datetime.date.today()
    for i in range(30):
        d = today - datetime.timedelta(days=29 - i)
        w.writerow([d.strftime('%Y-%m-%d'), '東京', '15.0', '20.0', '10.0', '0.0'])
    from urllib.parse import quote
    return Response(sio.getvalue().encode('utf-8-sig'), mimetype='text/csv',
                    headers={'Content-Disposition': "attachment; filename*=UTF-8''" + quote('気温データテンプレート.csv')})


@bp.route('/reports/weather/excel_template')
@permission_required('reports')
def weather_excel_template():
    """気象データ Excelテンプレートをダウンロード（days/locations パラメータ対応）"""
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment
    import io as _bio
    import datetime as _dt
    try:
        days = max(1, min(int(request.args.get('days', 30)), 365))
    except (ValueError, TypeError):
        days = 30
    # 地点リスト（クエリパラメータ locations=東京,大阪 で指定可、省略時は東京のみ）
    locs_param = request.args.get('locations', '東京')
    locations  = [l.strip() for l in locs_param.split(',') if l.strip()] or ['東京']

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '気象データ'
    headers = ['日付', '地点', '平均気温(℃)', '最高気温(℃)', '最低気温(℃)', '降水量(mm)']
    hfill = PatternFill(fill_type='solid', fgColor='1D4ED8')
    hfont = Font(color='FFFFFF', bold=True)
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = hfill
        c.font = hfont
        c.alignment = Alignment(horizontal='center')
    today = _dt.date.today()
    for loc in locations:
        for i in range(days):
            d = today - _dt.timedelta(days=days - 1 - i)
            ws.append([d.strftime('%Y-%m-%d'), loc, '', '', '', ''])
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 10
    for col in ['C', 'D', 'E', 'F']:
        ws.column_dimensions[col].width = 14
    buf = _bio.BytesIO()
    wb.save(buf)
    buf.seek(0)
    from urllib.parse import quote
    fname = quote('気象データ_テンプレート.xlsx')
    return Response(buf.read(),
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    headers={'Content-Disposition': f"attachment; filename*=UTF-8''{fname}"})


@bp.route('/reports/weather/import_page')
@permission_required('reports')
def weather_import_page():
    """過去気象データ 一括インポート専用ページ"""
    from flask import session
    result = session.pop('weather_import_result', None)
    return render_template('weather_import.html', result=result)


@bp.route('/reports/weather/excel_import', methods=['POST'])
@permission_required('reports')
def weather_excel_import():
    """気象データをExcel/CSVから一括インポート"""
    import openpyxl
    import datetime as _dt
    db = get_db()
    f  = request.files.get('excel_file')
    if not f or not f.filename:
        flash('ファイルを選択してください。', 'danger')
        return redirect(url_for('forecast.weather_data'))

    created = 0
    errors  = []
    filename = f.filename.lower()

    if filename.endswith('.csv'):
        # CSV パス
        content = f.read().decode('utf-8-sig', errors='ignore')
        reader  = csv.DictReader(io.StringIO(content))
        for i, row in enumerate(reader, 2):
            try:
                obs_date = (row.get('日付') or row.get('obs_date') or '').strip()
                location = (row.get('地点') or row.get('location') or '東京').strip()
                avg_temp = row.get('平均気温(℃)') or row.get('平均気温') or row.get('avg_temp') or ''
                max_temp = row.get('最高気温(℃)') or row.get('最高気温') or row.get('max_temp') or ''
                min_temp = row.get('最低気温(℃)') or row.get('最低気温') or row.get('min_temp') or ''
                precip   = row.get('降水量(mm)')  or row.get('降水量')  or 0
                if not obs_date or avg_temp == '':
                    continue
                db.execute("""
                    INSERT INTO weather_data
                      (obs_date, location, avg_temp, max_temp, min_temp, precipitation, source)
                    VALUES (%s, %s, %s, %s, %s, %s, 'import')
                    ON CONFLICT (obs_date, location) DO UPDATE
                      SET avg_temp=EXCLUDED.avg_temp, max_temp=EXCLUDED.max_temp,
                          min_temp=EXCLUDED.min_temp, precipitation=EXCLUDED.precipitation
                """, [obs_date, location,
                      float(avg_temp) if avg_temp else None,
                      float(max_temp) if max_temp else None,
                      float(min_temp) if min_temp else None,
                      float(precip)   if precip   else 0])
                created += 1
            except Exception as e:
                errors.append(f'行{i}: {e}')
    else:
        # Excel パス
        try:
            wb = openpyxl.load_workbook(f, data_only=True)
            ws = wb.active
        except Exception as e:
            flash(f'Excelファイルの読み込みに失敗しました: {e}', 'danger')
            return redirect(url_for('forecast.weather_data'))

        header_row = [str(c.value).strip() if c.value is not None else '' for c in ws[1]]

        def _col(names):
            for n in names:
                for j, h in enumerate(header_row):
                    if n in h:
                        return j
            return -1

        idx_date = _col(['日付', 'obs_date'])
        idx_loc  = _col(['地点', 'location'])
        idx_avg  = _col(['平均気温', 'avg_temp'])
        idx_max  = _col(['最高気温', 'max_temp'])
        idx_min  = _col(['最低気温', 'min_temp'])
        idx_pre  = _col(['降水量',   'precipitation'])

        if idx_date < 0 or idx_avg < 0:
            flash('必須列「日付」「平均気温」が見つかりません。テンプレートを確認してください。', 'danger')
            return redirect(url_for('forecast.weather_data'))

        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            try:
                obs_val = row[idx_date] if idx_date < len(row) else None
                if obs_val is None:
                    continue
                if isinstance(obs_val, (_dt.date, _dt.datetime)):
                    obs_date = obs_val.strftime('%Y-%m-%d')
                else:
                    obs_date = str(obs_val).strip()
                if not obs_date:
                    continue
                location = str(row[idx_loc]).strip() if idx_loc >= 0 and idx_loc < len(row) and row[idx_loc] else '東京'
                avg_temp = row[idx_avg] if idx_avg >= 0 and idx_avg < len(row) else None
                max_temp = row[idx_max] if idx_max >= 0 and idx_max < len(row) else None
                min_temp = row[idx_min] if idx_min >= 0 and idx_min < len(row) else None
                precip   = row[idx_pre] if idx_pre >= 0 and idx_pre < len(row) else 0
                if avg_temp is None:
                    continue
                db.execute("""
                    INSERT INTO weather_data
                      (obs_date, location, avg_temp, max_temp, min_temp, precipitation, source)
                    VALUES (%s, %s, %s, %s, %s, %s, 'import')
                    ON CONFLICT (obs_date, location) DO UPDATE
                      SET avg_temp=EXCLUDED.avg_temp, max_temp=EXCLUDED.max_temp,
                          min_temp=EXCLUDED.min_temp, precipitation=EXCLUDED.precipitation
                """, [obs_date, location,
                      float(avg_temp) if avg_temp is not None else None,
                      float(max_temp) if max_temp is not None else None,
                      float(min_temp) if min_temp is not None else None,
                      float(precip)   if precip   is not None else 0])
                created += 1
            except Exception as e:
                errors.append(f'行{i}: {e}')

    db.commit()

    # インポート後に気温感応度を自動再計算
    sens_n = 0
    if created > 0:
        try:
            sens_n = recalc_temp_sensitivity(db)
            if sens_n:
                _save_sens_log(db, sens_n, 'import')
        except Exception:
            sens_n = 0

    # 地点別インポート結果を集計してセッションへ
    from flask import session
    loc_summary = {}
    for row_data in db.execute("""
        SELECT location, COUNT(*) AS cnt, MIN(obs_date) AS oldest, MAX(obs_date) AS newest
        FROM weather_data WHERE source='import'
        GROUP BY location ORDER BY location
    """).fetchall():
        loc_summary[row_data['location']] = {
            'cnt':    row_data['cnt'],
            'oldest': str(row_data['oldest']),
            'newest': str(row_data['newest']),
        }
    session['weather_import_result'] = {
        'created':     created,
        'error_count': len(errors),
        'errors':      errors[:5],
        'loc_summary': loc_summary,
    }
    redirect_to = request.form.get('redirect_to', 'import_page')
    sens_msg = f' 気温感応度 {sens_n}商品 再計算済み。' if sens_n else ''
    if redirect_to == 'weather_data':
        flash(f'{created}件インポート完了。{sens_msg}', 'success' if not errors else 'warning')
        return redirect(url_for('forecast.weather_data'))
    session['weather_import_result']['sens_n'] = sens_n
    return redirect(url_for('forecast.weather_import_page'))


@bp.route('/reports/weather/fetch_api', methods=['POST'])
@permission_required('reports')
def weather_fetch_api():
    """Open-Meteo APIから気象データを取得してDBに保存"""
    import urllib.request
    import urllib.parse
    db = get_db()
    location = (request.form.get('location') or '東京').strip()
    try:
        lat = float(request.form.get('lat') or 35.6897)
        lon = float(request.form.get('lon') or 139.6922)
    except (ValueError, TypeError):
        lat, lon = 35.6897, 139.6922
    try:
        days = int(request.form.get('days') or 30)
        days = max(1, min(days, 365))
    except (ValueError, TypeError):
        days = 30

    import json as _json
    from datetime import date as _date, timedelta as _timedelta
    end_date   = _date.today()
    start_date = end_date - _timedelta(days=days - 1)

    if days <= 92:
        # forecast API（直近92日まで）
        params = urllib.parse.urlencode({
            'latitude':      lat,
            'longitude':     lon,
            'daily':         'temperature_2m_max,temperature_2m_min,temperature_2m_mean,precipitation_sum',
            'timezone':      'Asia/Tokyo',
            'past_days':     days,
            'forecast_days': 0,
        })
        url = f'https://api.open-meteo.com/v1/forecast?{params}'
    else:
        # archive API（93日〜365日）
        params = urllib.parse.urlencode({
            'latitude':   lat,
            'longitude':  lon,
            'daily':      'temperature_2m_max,temperature_2m_min,temperature_2m_mean,precipitation_sum',
            'timezone':   'Asia/Tokyo',
            'start_date': start_date.isoformat(),
            'end_date':   end_date.isoformat(),
        })
        url = f'https://archive-api.open-meteo.com/v1/archive?{params}'

    try:
        with urllib.request.urlopen(url, timeout=30) as resp:
            data = _json.loads(resp.read())
    except Exception as e:
        flash(f'API取得エラー: {e}', 'danger')
        return redirect(url_for('forecast.weather_data'))

    daily = data.get('daily', {})
    dates     = daily.get('time', [])
    max_temps = daily.get('temperature_2m_max', [])
    min_temps = daily.get('temperature_2m_min', [])
    avg_temps = daily.get('temperature_2m_mean', [])
    precips   = daily.get('precipitation_sum', [])

    created = 0
    errors  = []
    for i, obs_date in enumerate(dates):
        try:
            avg_t = avg_temps[i] if i < len(avg_temps) else None
            max_t = max_temps[i] if i < len(max_temps) else None
            min_t = min_temps[i] if i < len(min_temps) else None
            prec  = precips[i]   if i < len(precips)   else None
            if avg_t is None:
                continue
            db.execute("""
                INSERT INTO weather_data (obs_date, location, avg_temp, max_temp, min_temp, precipitation, source)
                VALUES (%s, %s, %s, %s, %s, %s, 'api')
                ON CONFLICT (obs_date, location) DO UPDATE
                  SET avg_temp=EXCLUDED.avg_temp, max_temp=EXCLUDED.max_temp,
                      min_temp=EXCLUDED.min_temp, precipitation=EXCLUDED.precipitation,
                      source='api'
            """, [obs_date, location, avg_t, max_t, min_t, prec])
            created += 1
        except Exception as e:
            errors.append(f'{obs_date}: {e}')
    db.commit()
    msg = f'{location}: {created}件の気象データをAPIから取得しました。'
    if errors:
        msg += f' エラー: {len(errors)}件'
    flash(msg, 'success' if not errors else 'warning')
    return redirect(url_for('forecast.weather_data'))


@bp.route('/reports/weather/fetch_multi', methods=['POST'])
@permission_required('reports')
def weather_fetch_multi():
    """複数地点の気象データを一括取得"""
    import urllib.request
    import urllib.parse
    db = get_db()
    locations_raw = request.form.getlist('locations')
    try:
        days = int(request.form.get('days') or 30)
        days = max(1, min(days, 92))
    except (ValueError, TypeError):
        days = 30

    if not locations_raw:
        flash('地点を選択してください。', 'warning')
        return redirect(url_for('forecast.weather_data'))

    total_created = 0
    total_errors  = 0
    failed_locs   = []

    for loc_str in locations_raw:
        parts = loc_str.split('|')
        if len(parts) != 3:
            continue
        location = parts[0].strip()
        try:
            lat = float(parts[1])
            lon = float(parts[2])
        except ValueError:
            continue

        params = urllib.parse.urlencode({
            'latitude':  lat,
            'longitude': lon,
            'daily':     'temperature_2m_max,temperature_2m_min,temperature_2m_mean,precipitation_sum',
            'timezone':  'Asia/Tokyo',
            'past_days': days,
            'forecast_days': 0,
        })
        url = f'https://api.open-meteo.com/v1/forecast?{params}'
        try:
            import json as _json
            with urllib.request.urlopen(url, timeout=15) as resp:
                data = _json.loads(resp.read())
        except Exception as e:
            failed_locs.append(f'{location}({e})')
            total_errors += 1
            continue

        daily     = data.get('daily', {})
        dates     = daily.get('time', [])
        max_temps = daily.get('temperature_2m_max', [])
        min_temps = daily.get('temperature_2m_min', [])
        avg_temps = daily.get('temperature_2m_mean', [])
        precips   = daily.get('precipitation_sum', [])

        for i, obs_date in enumerate(dates):
            try:
                avg_t = avg_temps[i] if i < len(avg_temps) else None
                max_t = max_temps[i] if i < len(max_temps) else None
                min_t = min_temps[i] if i < len(min_temps) else None
                prec  = precips[i]   if i < len(precips)   else None
                if avg_t is None:
                    continue
                db.execute("""
                    INSERT INTO weather_data
                      (obs_date, location, avg_temp, max_temp, min_temp, precipitation, source)
                    VALUES (%s, %s, %s, %s, %s, %s, 'api')
                    ON CONFLICT (obs_date, location) DO UPDATE
                      SET avg_temp=EXCLUDED.avg_temp, max_temp=EXCLUDED.max_temp,
                          min_temp=EXCLUDED.min_temp, precipitation=EXCLUDED.precipitation,
                          source='api'
                """, [obs_date, location, avg_t, max_t, min_t, prec])
                total_created += 1
            except Exception:
                total_errors += 1

    db.commit()
    loc_count = len(locations_raw) - len(failed_locs)
    msg = f'{loc_count}地点・{total_created}件の気象データを取得しました。'
    if failed_locs:
        msg += f' 取得失敗: {", ".join(failed_locs[:3])}{"他" if len(failed_locs)>3 else ""}'
    flash(msg, 'success' if not failed_locs else 'warning')
    return redirect(url_for('forecast.weather_data'))
