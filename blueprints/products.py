"""商品管理 Blueprint"""
from flask import Blueprint, render_template, request, redirect, url_for, flash, Response, session
import csv, io, logging
from db import get_db
from auth_helpers import login_required, admin_required, permission_required
from helpers import _PRODUCT_COLS, _EXPORT_ROW_LIMIT

logger = logging.getLogger('inventory.products')
bp = Blueprint('products', __name__)


_REORDER_AUTO_LABEL = {0: '0 手動', 1: '1 AIモード', 2: '2 前年実績'}

@bp.route('/products/export')
@admin_required
def product_export():
    fmt = request.args.get('fmt', 'xlsx')
    db  = get_db()
    total = db.execute("SELECT COUNT(*) AS c FROM products WHERE is_active=1").fetchone()['c']
    if total > _EXPORT_ROW_LIMIT:
        flash(f'件数が多いため先頭 {_EXPORT_ROW_LIMIT:,} 件のみエクスポートします（全 {total:,} 件）。', 'warning')
    rows = db.execute(
        "SELECT * FROM products WHERE is_active=1 ORDER BY CAST(NULLIF(regexp_replace(supplier_cd,'[^0-9]','','g'),'') AS BIGINT) NULLS LAST, CAST(NULLIF(regexp_replace(product_cd,'[^0-9]','','g'),'') AS BIGINT) NULLS LAST LIMIT %s",
        [_EXPORT_ROW_LIMIT]
    ).fetchall()

    headers = [c[1] for c in _PRODUCT_COLS]
    keys    = [c[0] for c in _PRODUCT_COLS]

    def _cell_val(k, v):
        """reorder_auto はラベル付き文字列で出力（インポート時も 0/1/2 数値として認識）"""
        if k == 'reorder_auto' and v is not None:
            return _REORDER_AUTO_LABEL.get(int(v), str(v))
        return v if v is not None else ''

    if fmt == 'csv':
        buf = io.StringIO()
        w = csv.writer(buf)
        w.writerow(headers)
        for r in rows:
            w.writerow([_cell_val(k, r[k]) for k in keys])
        buf.seek(0)
        _db2 = get_db()
        _exp_name = _db2.execute("SELECT value FROM settings WHERE key='product_export_name'").fetchone()
        _exp_name = (_exp_name['value'] if _exp_name else '商品マスタ') + '.csv'
        from urllib.parse import quote
        return Response(
            buf.getvalue().encode('utf-8-sig'),
            mimetype='text/csv',
            headers={'Content-Disposition': f"attachment; filename*=UTF-8''{quote(_exp_name)}"}
        )
    else:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.worksheet.datavalidation import DataValidation
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '商品マスタ'

        # ヘッダースタイル
        hfill  = PatternFill('solid', fgColor='1E3A8A')
        hfont  = Font(bold=True, color='FFFFFF', name='Meiryo UI', size=10)
        hali   = Alignment(horizontal='center', vertical='center')
        bdr    = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'),  bottom=Side(style='thin')
        )
        reorder_auto_col = next(
            (ci for ci, c in enumerate(_PRODUCT_COLS, 1) if c[0] == 'reorder_auto'), None
        )
        for ci, h in enumerate(headers, 1):
            cell = ws.cell(1, ci, h)
            cell.font  = hfont
            cell.fill  = hfill
            cell.alignment = hali
            cell.border = bdr

        # データ行
        rfill_odd  = PatternFill('solid', fgColor='FFFFFF')
        rfill_even = PatternFill('solid', fgColor='F0F4FF')
        rfont = Font(name='Meiryo UI', size=10)
        for ri, r in enumerate(rows, 2):
            fill = rfill_even if ri % 2 == 0 else rfill_odd
            for ci, k in enumerate(keys, 1):
                cell = ws.cell(ri, ci, _cell_val(k, r[k]))
                cell.font   = rfont
                cell.fill   = fill
                cell.border = bdr
                cell.alignment = Alignment(vertical='center')
                if ci == 1:
                    cell.number_format = '0'

        # 発注点自動更新列にドロップダウン検証
        if reorder_auto_col:
            col_letter = openpyxl.utils.get_column_letter(reorder_auto_col)
            dv = DataValidation(
                type='list',
                formula1='"0 手動,1 AIモード,2 前年実績"',
                allow_blank=True,
                showErrorMessage=False,
            )
            dv.sqref = f'{col_letter}2:{col_letter}100000'
            ws.add_data_validation(dv)

        # 列幅
        col_widths = [16,14,30,14,20,8,8,8,8,12,12,8,8,10,10,18,14,10,10,10,10]
        for i, w_val in enumerate(col_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w_val
        ws.row_dimensions[1].height = 20
        ws.freeze_panes = 'A2'

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        _db3 = get_db()
        _exp_name2 = _db3.execute("SELECT value FROM settings WHERE key='product_export_name'").fetchone()
        _exp_name2 = (_exp_name2['value'] if _exp_name2 else '商品マスタ') + '.xlsx'
        from urllib.parse import quote
        return Response(
            buf.getvalue(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': f"attachment; filename*=UTF-8''{quote(_exp_name2)}"}
        )


@bp.route('/products/template')
@admin_required
def product_template():
    """インポート用テンプレートExcelをダウンロード"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.worksheet.datavalidation import DataValidation
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '商品マスタ'

    headers  = [c[1] for c in _PRODUCT_COLS]
    notes    = {
        'JANコード':       '必須・UNIQUE キー（13桁）',
        '商品コード':      '必須',
        '商品名':          '必須',
        '仕入先コード':    '必須',
        '仕入先名':        '必須',
        '入数':            '1ケースあたりの個数',
        '発注単位':        '何ケース単位で発注するか',
        '発注数量':        '1回の発注ケース数',
        '発注点':          '在庫がこの数以下で自動発注',
        '発注点自動更新':  '0=手動 / 1=AIモード / 2=前年実績',
        'リードタイム日数':'仕入先から納品までの日数',
        '安全係数':        '例: 1.3 = 30%の余裕',
        'メーカーロット数':'0=未使用',
        '賞味期限日数':    '製造から賞味期限までの日数',
        '期限アラート日数':'期限X日前にアラートメール',
        '混載グループ名':  '同グループ名の商品を混載発注',
        '混載ロットルール':'gte=以上 / unit=単位',
        '混載ケース数':    '混載の基準ケース数',
        '強制発注日数':    'ロット未達でもX日後に強制発注',
        '原価':            '必須・円（例: 150.00）',
        '売価':            '必須・円（例: 200.00）',
    }

    hfill = PatternFill('solid', fgColor='1E3A8A')
    hfont = Font(bold=True, color='FFFFFF', name='Meiryo UI', size=10)
    nfill = PatternFill('solid', fgColor='FEF9C3')
    nfont = Font(name='Meiryo UI', size=9, color='78350F')
    bdr   = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'),  bottom=Side(style='thin')
    )

    # 発注点自動更新列のインデックスを特定
    reorder_auto_col = next(
        (ci for ci, c in enumerate(_PRODUCT_COLS, 1) if c[0] == 'reorder_auto'), None
    )

    for ci, h in enumerate(headers, 1):
        c1 = ws.cell(1, ci, h)
        c1.font = hfont; c1.fill = hfill
        c1.alignment = Alignment(horizontal='center', vertical='center')
        c1.border = bdr
        c2 = ws.cell(2, ci, notes.get(h, ''))
        c2.font = nfont; c2.fill = nfill
        c2.alignment = Alignment(wrap_text=True, vertical='center')
        c2.border = bdr

    # 発注点自動更新列にドロップダウン検証を追加（データ行 3行目〜）
    if reorder_auto_col:
        col_letter = openpyxl.utils.get_column_letter(reorder_auto_col)
        dv = DataValidation(
            type='list',
            formula1='"0,1,2"',
            allow_blank=True,
            showErrorMessage=True,
            errorTitle='入力エラー',
            error='0=手動 / 1=AIモード / 2=前年実績 のいずれかを入力してください',
            showInputMessage=True,
            promptTitle='発注点自動更新モード',
            prompt='0=手動（自動更新しない）\n1=AIモード自動更新\n2=前年実績モード自動更新'
        )
        dv.sqref = f'{col_letter}3:{col_letter}10000'
        ws.add_data_validation(dv)

    col_widths = [16,14,30,14,20,8,8,8,8,12,12,8,8,10,10,18,14,10,10,10,10]
    for i, w_val in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w_val
    ws.row_dimensions[1].height = 20
    ws.row_dimensions[2].height = 36
    ws.freeze_panes = 'A3'

    # 凡例シートを追加
    ws_legend = wb.create_sheet('凡例')
    legend_data = [
        ('列名', '値', '説明'),
        ('発注点自動更新', '0', '手動（自動更新しない）'),
        ('発注点自動更新', '1', 'AIモード自動更新（毎月1日、AI予測値で発注点を自動計算）'),
        ('発注点自動更新', '2', '前年実績モード自動更新（毎月1日、前年同月実績で発注点を自動計算）'),
        ('', '', ''),
        ('混載ロットルール', 'gte', '合計ケース数が混載ケース数以上になったら発注'),
        ('混載ロットルール', 'unit', '混載ケース数の倍数単位で発注'),
    ]
    lfill_h = PatternFill('solid', fgColor='1E3A8A')
    lfont_h = Font(bold=True, color='FFFFFF', name='Meiryo UI', size=10)
    lfont   = Font(name='Meiryo UI', size=10)
    for ri, row in enumerate(legend_data, 1):
        for ci, val in enumerate(row, 1):
            cell = ws_legend.cell(ri, ci, val)
            cell.font = lfont_h if ri == 1 else lfont
            if ri == 1:
                cell.fill = lfill_h
                cell.alignment = Alignment(horizontal='center')
    ws_legend.column_dimensions['A'].width = 18
    ws_legend.column_dimensions['B'].width = 8
    ws_legend.column_dimensions['C'].width = 55

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    _db = get_db()
    _tmpl_name = _db.execute("SELECT value FROM settings WHERE key='product_template_name'").fetchone()
    _tmpl_name = (_tmpl_name['value'] if _tmpl_name else '商品マスタ_テンプレート') + '.xlsx'
    from urllib.parse import quote
    return Response(
        buf.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f"attachment; filename*=UTF-8''{quote(_tmpl_name)}"}
    )


@bp.route('/products/import', methods=['GET','POST'])
@admin_required
def product_import():
    if request.method == 'GET':
        return render_template('product_import.html')

    f    = request.files.get('file')
    mode = request.form.get('mode', 'upsert')  # upsert or add_only
    if not f or not f.filename:
        flash('ファイルを選択してください。', 'danger')
        return redirect(url_for('products.product_import'))

    fname = f.filename.lower()
    rows_data = []

    try:
        if fname.endswith('.csv'):
            content = f.read().decode('utf-8-sig', errors='replace')
            reader  = csv.DictReader(io.StringIO(content))
            rows_data = list(reader)

        elif fname.endswith(('.xlsx', '.xls')):
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(f.read()), data_only=True)
            ws = wb.active
            headers = [str(ws.cell(1, c).value or '').strip() for c in range(1, ws.max_column+1)]
            for ri in range(2, ws.max_row+1):
                row = {}
                for ci, h in enumerate(headers, 1):
                    v = ws.cell(ri, ci).value
                    row[h] = str(v).strip() if v is not None else ''
                # 空行スキップ
                if any(v for v in row.values()):
                    rows_data.append(row)
        else:
            flash('xlsx または csv ファイルを選択してください。', 'danger')
            return redirect(url_for('products.product_import'))

    except Exception as e:
        flash(f'ファイル読み込みエラー: {e}', 'danger')
        return redirect(url_for('products.product_import'))

    # 列名マッピング（日本語ヘッダー → DBキー）
    col_map = {c[1]: c[0] for c in _PRODUCT_COLS}
    type_map = {c[0]: c[2] for c in _PRODUCT_COLS}

    # 発注点自動更新のテキストラベル → 数値マッピング
    _reorder_auto_map = {
        '0': 0, '手動': 0, 'manual': 0, '0 手動': 0, '自動更新しない': 0,
        '1': 1, 'ai': 1, 'aiモード': 1, 'aiモード自動': 1, '1 aiモード': 1,
        '前年実績': 2, '2': 2, 'ly': 2, '前年実績モード': 2, '2 前年実績': 2,
        '前年実績自動': 2, '前年実績モード自動': 2,
    }

    def cast(key, val):
        t = type_map.get(key, 'text')
        if val == '' or val is None:
            return None
        # 発注点自動更新はテキストラベルも受け付ける
        if key == 'reorder_auto':
            s = str(val).strip().lower()
            if s in _reorder_auto_map:
                return _reorder_auto_map[s]
            try: return int(float(s.replace(',','')))
            except: return None
        if t == 'int':
            try: return int(float(str(val).replace(',','')))
            except: return None
        if t == 'float':
            try: return float(str(val).replace(',',''))
            except: return None
        return str(val).strip()

    db = get_db()
    cnt_add = cnt_upd = cnt_skip = cnt_err = 0
    errors = []

    # グローバル設定から新規商品のデフォルト自動更新モードを取得
    _mode_row = db.execute("SELECT value FROM settings WHERE key='reorder_auto_mode'").fetchone()
    _mode_val = _mode_row['value'] if _mode_row else 'ai'
    _reorder_auto_default = {'ai': 1, 'ly': 2}.get(_mode_val, 0)  # manual=0, ai=1, ly=2

    for ri, row in enumerate(rows_data, 1):
        # テンプレートの説明行スキップ
        jan_raw = str(row.get('JANコード') or row.get('jan') or '').strip()
        if jan_raw in ('JANコード', 'jan', '必須・UNIQUE キー（13桁）'):
            cnt_skip += 1
            continue

        # 必須チェック
        product_name = str(row.get('商品名') or row.get('product_name') or '').strip()
        supplier_cd  = str(row.get('仕入先コード') or row.get('supplier_cd') or '').strip()
        supplier_name= str(row.get('仕入先名') or row.get('supplier_name') or '').strip()
        product_cd   = str(row.get('商品コード') or row.get('product_cd') or '').strip()

        # JANも商品CDもない行はスキップ
        if not jan_raw and not product_cd:
            cnt_skip += 1
            continue

        cost_price_raw = str(row.get('原価') or row.get('cost_price') or '').strip()
        sell_price_raw = str(row.get('売価') or row.get('sell_price') or '').strip()

        if not all([product_name, supplier_cd, supplier_name, product_cd, cost_price_raw, sell_price_raw]):
            missing = []
            if not product_name: missing.append('商品名')
            if not supplier_cd: missing.append('仕入先コード')
            if not supplier_name: missing.append('仕入先名')
            if not product_cd: missing.append('商品コード')
            if not cost_price_raw: missing.append('原価')
            if not sell_price_raw: missing.append('売価')
            errors.append(f'行{ri}: 必須項目が不足 ({", ".join(missing)}) JAN={jan_raw}')
            cnt_err += 1
            continue

        try:
            # JANまたは商品CDで既存チェック
            existing = None
            if jan_raw:
                existing = db.execute(
                    "SELECT id, jan FROM products WHERE jan=%s", [jan_raw]
                ).fetchone()
            if not existing and product_cd:
                existing = db.execute(
                    "SELECT id, jan FROM products WHERE product_cd=%s AND is_active=1", [product_cd]
                ).fetchone()
                if existing and not jan_raw:
                    jan_raw = existing['jan']  # 商品CDで見つかった場合JANを補完

            # 全カラムの値を準備
            vals = {}
            for jp_name, db_key in col_map.items():
                v = row.get(jp_name) or row.get(db_key)
                vals[db_key] = cast(db_key, v)

            # NULLのデフォルト値を補完
            defaults = {
                'unit_qty':1,'order_unit':1,'order_qty':1,'reorder_point':0,
                'reorder_auto':_reorder_auto_default,'lead_time_days':3,'safety_factor':1.3,
                'lot_size':0,'shelf_life_days':365,'expiry_alert_days':30,
                'mixed_group':'','mixed_lot_mode':'gte','mixed_lot_cases':3,'mixed_force_days':3,
                'cost_price':0,'sell_price':0,
            }
            for k, dv in defaults.items():
                if vals.get(k) is None:
                    vals[k] = dv

            if existing:
                if mode == 'add_only':
                    cnt_skip += 1
                    continue
                db.execute("""
                    UPDATE products SET
                    supplier_cd=%s,supplier_name=%s,product_cd=%s,product_name=%s,
                    unit_qty=%s,order_unit=%s,order_qty=%s,reorder_point=%s,reorder_auto=%s,
                    lead_time_days=%s,safety_factor=%s,lot_size=%s,
                    shelf_life_days=%s,expiry_alert_days=%s,
                    mixed_group=%s,mixed_lot_mode=%s,mixed_lot_cases=%s,mixed_force_days=%s,
                    cost_price=%s,sell_price=%s,
                    is_active=1
                    WHERE jan=%s
                """, [
                    vals['supplier_cd'],vals['supplier_name'],vals['product_cd'],vals['product_name'],
                    vals['unit_qty'],vals['order_unit'],vals['order_qty'],vals['reorder_point'],vals['reorder_auto'],
                    vals['lead_time_days'],vals['safety_factor'],vals['lot_size'],
                    vals['shelf_life_days'],vals['expiry_alert_days'],
                    vals['mixed_group'],vals['mixed_lot_mode'],vals['mixed_lot_cases'],vals['mixed_force_days'],
                    vals.get('cost_price',0) or 0, vals.get('sell_price',0) or 0,
                    jan_raw
                ])
                cnt_upd += 1
            else:
                if not jan_raw:
                    errors.append(f'行{ri}: JANコードなし・商品CDのみでは新規追加できません (商品CD={product_cd})')
                    cnt_err += 1
                    continue
                db.execute("""
                    INSERT INTO products
                    (jan,supplier_cd,supplier_name,product_cd,product_name,
                     unit_qty,order_unit,order_qty,reorder_point,reorder_auto,
                     lead_time_days,safety_factor,lot_size,
                     shelf_life_days,expiry_alert_days,
                     mixed_group,mixed_lot_mode,mixed_lot_cases,mixed_force_days,
                     cost_price,sell_price)
                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                """, [
                    jan_raw,vals['supplier_cd'],vals['supplier_name'],vals['product_cd'],vals['product_name'],
                    vals['unit_qty'],vals['order_unit'],vals['order_qty'],vals['reorder_point'],vals['reorder_auto'],
                    vals['lead_time_days'],vals['safety_factor'],vals['lot_size'],
                    vals['shelf_life_days'],vals['expiry_alert_days'],
                    vals['mixed_group'],vals['mixed_lot_mode'],vals['mixed_lot_cases'],vals['mixed_force_days'],
                    vals.get('cost_price',0) or 0, vals.get('sell_price',0) or 0,
                ])
                cnt_add += 1

        except Exception as e:
            errors.append(f'行{ri} JAN={jan_raw} 商品CD={product_cd}: {e}')
            cnt_err += 1

    db.commit()

    msg = f'完了：追加 {cnt_add}件 / 更新 {cnt_upd}件 / スキップ {cnt_skip}件'
    session.pop('import_errors', None)
    if cnt_err:
        msg += f' / エラー {cnt_err}件'
        flash(msg, 'warning')
        session['import_errors'] = [e[:120] for e in errors[:20]]
    else:
        flash(msg, 'success')

    return redirect(url_for('products.products'))

@bp.route('/products')
@permission_required('products')
def products():
    db = get_db()
    q = request.args.get('q','').strip()
    rows = db.execute("""
        SELECT p.*, COALESCE((SELECT SUM(quantity) FROM stocks WHERE jan=p.jan),0) as stock_qty
        FROM products p WHERE p.is_active=1 ORDER BY CAST(NULLIF(regexp_replace(p.supplier_cd,'[^0-9]','','g'),'') AS BIGINT) NULLS LAST, CAST(NULLIF(regexp_replace(p.product_cd,'[^0-9]','','g'),'') AS BIGINT) NULLS LAST
    """).fetchall()
    if q:
        rows = [r for r in rows if q.lower() in (r['jan'] or '').lower()
                or q.lower() in (r['product_cd'] or '').lower()
                or q.lower() in (r['product_name'] or '').lower()
                or q.lower() in (r['supplier_cd'] or '').lower()
                or q.lower() in (r['supplier_name'] or '').lower()]
    return render_template('products.html', products=rows, q=q)

@bp.route('/products/new', methods=['GET','POST'])
@admin_required
def product_new():
    if request.method == 'POST':
        f = request.form
        db = get_db()
        db.execute("""
            INSERT INTO products
            (supplier_cd,supplier_name,supplier_email,jan,product_cd,product_name,
             unit_qty,order_unit,order_qty,reorder_point,reorder_auto,lot_size,
             shelf_life_days,expiry_alert_days,safety_factor,lead_time_days,
             mixed_group,mixed_lot_mode,mixed_lot_cases,mixed_force_days,
             cost_price,sell_price,location_code,shelf_face_qty,shelf_replenish_point,
             manual_adj_factor)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
        """, [f['supplier_cd'],f['supplier_name'],f.get('supplier_email',''),
              f['jan'],f['product_cd'],f['product_name'],
              int(f.get('unit_qty',1)),int(f.get('order_unit',1)),
              int(f.get('order_qty',1)),int(f.get('reorder_point',0)),
              int(f.get('reorder_auto', 0) or 0),
              int(f.get('lot_size',0)),int(f.get('shelf_life_days',365)),
              int(f.get('expiry_alert_days',30)),
              float(f.get('safety_factor',1.3)),int(f.get('lead_time_days',3)),
              f.get('mixed_group',''),f.get('mixed_lot_mode','gte'),
              int(f.get('mixed_lot_cases',3)),int(f.get('mixed_force_days',3)),
              float(f.get('cost_price',0)),float(f.get('sell_price',0)),
              f.get('location_code',''),
              int(f.get('shelf_face_qty',0) or 0),
              int(f.get('shelf_replenish_point',0) or 0),
              float(f.get('manual_adj_factor',1.0) or 1.0)])
        db.commit()
        flash('商品を登録しました。', 'success')
        return redirect(url_for('products.products'))
    return render_template('product_form.html', product=None)

@bp.route('/products/inactive')
@login_required
def products_inactive():
    db = get_db()
    q = request.args.get('q','').strip()
    products = db.execute("""
        SELECT p.*, COALESCE(SUM(s.quantity),0) AS stock_qty
        FROM products p
        LEFT JOIN stocks s ON s.jan=p.jan
        WHERE p.is_active=0
        GROUP BY p.id
        ORDER BY CAST(NULLIF(regexp_replace(p.supplier_cd,'[^0-9]','','g'),'') AS BIGINT) NULLS LAST, CAST(NULLIF(regexp_replace(p.product_cd,'[^0-9]','','g'),'') AS BIGINT) NULLS LAST
    """).fetchall()
    if q:
        products = [r for r in products if q.lower() in (r['jan'] or '').lower()
                or q.lower() in (r['product_cd'] or '').lower()
                or q.lower() in (r['product_name'] or '').lower()
                or q.lower() in (r['supplier_cd'] or '').lower()
                or q.lower() in (r['supplier_name'] or '').lower()]
    return render_template('products_inactive.html', products=products, q=q)


@bp.route('/products/<int:pid>/delete', methods=['POST'])
@admin_required
def product_delete(pid):
    db = get_db()
    p = db.execute("SELECT jan, product_name FROM products WHERE id=%s", [pid]).fetchone()
    if not p:
        flash('商品が見つかりません', 'error')
        return redirect(url_for('products.products'))
    jan = p['jan']
    # 発注保留・発注済みフラグをクリア
    db.execute("DELETE FROM order_pending WHERE jan=%s", [jan])
    db.execute("UPDATE products SET is_active=0, ordered_at='' WHERE id=%s", [pid])
    db.commit()
    flash(f'「{p["product_name"]}」を無効化しました（復活させるには無効商品一覧から有効化できます）', 'success')
    return redirect(url_for('products.products'))

@bp.route('/products/<int:pid>/restore', methods=['POST'])
@admin_required
def product_restore(pid):
    db = get_db()
    p = db.execute("SELECT product_name FROM products WHERE id=%s", [pid]).fetchone()
    if not p:
        flash('商品が見つかりません', 'error')
        return redirect(url_for('products.products'))
    db.execute("UPDATE products SET is_active=1 WHERE id=%s", [pid])
    db.commit()
    flash(f'「{p["product_name"]}」を有効化しました', 'success')
    return redirect(url_for('products.products_inactive'))


@bp.route('/products/<int:pid>/edit', methods=['GET','POST'])
def product_edit(pid):
    db = get_db()
    product = db.execute("SELECT * FROM products WHERE id=%s", [pid]).fetchone()
    if not product:
        flash('商品が見つかりません。', 'danger')
        return redirect(url_for('products.products'))
    if request.method == 'POST':
        f = request.form
        db.execute("""
            UPDATE products SET
            supplier_cd=%s,supplier_name=%s,supplier_email=%s,product_cd=%s,product_name=%s,
            unit_qty=%s,order_unit=%s,order_qty=%s,reorder_point=%s,reorder_auto=%s,lot_size=%s,
            shelf_life_days=%s,expiry_alert_days=%s,safety_factor=%s,lead_time_days=%s,
            mixed_group=%s,mixed_lot_mode=%s,mixed_lot_cases=%s,mixed_force_days=%s,
            cost_price=%s,sell_price=%s,location_code=%s,shelf_face_qty=%s,shelf_replenish_point=%s,
            manual_adj_factor=%s
            WHERE id=%s
        """, [f['supplier_cd'],f['supplier_name'],f.get('supplier_email',''),
              f['product_cd'],f['product_name'],
              int(f.get('unit_qty',1)),int(f.get('order_unit',1)),
              int(f.get('order_qty',1)),int(f.get('reorder_point',0)),
              int(f.get('reorder_auto', 0) or 0),
              int(f.get('lot_size',0)),int(f.get('shelf_life_days',365)),
              int(f.get('expiry_alert_days',30)),
              float(f.get('safety_factor',1.3)),int(f.get('lead_time_days',3)),
              f.get('mixed_group',''),f.get('mixed_lot_mode','gte'),
              int(f.get('mixed_lot_cases',3)),int(f.get('mixed_force_days',3)),
              float(f.get('cost_price',0)),float(f.get('sell_price',0)),
              f.get('location_code',''),
              int(f.get('shelf_face_qty',0) or 0),
              int(f.get('shelf_replenish_point',0) or 0),
              float(f.get('manual_adj_factor',1.0) or 1.0),
              pid])
        db.commit()
        flash('商品情報を更新しました。', 'success')
        return redirect(url_for('products.products'))
    return render_template('product_form.html', product=product)
